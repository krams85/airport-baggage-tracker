"""
Persistent SQLite storage for ReID match events.

Every match found in production mode is automatically saved to a local
SQLite database (matches.db by default). The file survives restarts,
so you can query the full history across multiple shifts.

Thread-safe: insert() is called from Qt's main thread via signals,
but the class uses a lock for safety if ever called from a worker.
"""

import csv
import logging
import sqlite3
import threading
from datetime import datetime
from pathlib import Path
from typing import List, Optional, Tuple

logger = logging.getLogger("BaggageTracker.Storage")

_SCHEMA = """
CREATE TABLE IF NOT EXISTS matches (
    id           INTEGER PRIMARY KEY AUTOINCREMENT,
    timestamp    REAL    NOT NULL,
    transit_sec  REAL    NOT NULL,
    verdict      TEXT    NOT NULL,
    similarity   REAL    NOT NULL,
    source_cam   TEXT    NOT NULL,
    source_desk  INTEGER NOT NULL,
    source_track INTEGER NOT NULL,
    query_cam    TEXT    NOT NULL,
    query_desk   INTEGER NOT NULL,
    query_track  INTEGER NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_ts ON matches (timestamp);
"""


class MatchStorage:
    """SQLite-backed match history with CSV/Excel export."""

    def __init__(self, db_path: str = "matches.db"):
        self._path = db_path
        self._lock = threading.Lock()
        self._con: Optional[sqlite3.Connection] = None
        self._open()

    # ── Lifecycle ────────────────────────────────────────────────────────────

    def _open(self):
        try:
            self._con = sqlite3.connect(
                self._path, check_same_thread=False, timeout=10
            )
            self._con.executescript(_SCHEMA)
            self._con.commit()
            logger.info("MatchStorage: база открыта (%s)", self._path)
        except Exception as exc:
            logger.error("MatchStorage: не удалось открыть БД: %s", exc)
            self._con = None

    def close(self):
        if self._con:
            try:
                self._con.close()
            except Exception:
                pass
            self._con = None

    # ── Write ────────────────────────────────────────────────────────────────

    def save(self, mr) -> None:
        """Persist one MatchResult. Silent on error."""
        if self._con is None:
            return
        try:
            transit = max(0.0, mr.timestamp - mr.source_entry.timestamp)
            with self._lock:
                self._con.execute(
                    """INSERT INTO matches
                       (timestamp, transit_sec, verdict, similarity,
                        source_cam, source_desk, source_track,
                        query_cam,  query_desk,  query_track)
                       VALUES (?,?,?,?,?,?,?,?,?,?)""",
                    (
                        mr.timestamp,
                        transit,
                        mr.verdict,
                        mr.similarity,
                        mr.source_entry.cam_name,
                        mr.source_entry.counter_id,
                        mr.source_entry.track_id,
                        mr.query_cam_name,
                        mr.query_counter_id,
                        mr.query_track_id,
                    ),
                )
                self._con.commit()
        except Exception as exc:
            logger.error("MatchStorage.save: %s", exc)

    # ── Read ─────────────────────────────────────────────────────────────────

    def get_all(self) -> List[Tuple]:
        """Return all rows ordered by timestamp desc."""
        if self._con is None:
            return []
        try:
            with self._lock:
                cur = self._con.execute(
                    "SELECT * FROM matches ORDER BY timestamp DESC"
                )
                return cur.fetchall()
        except Exception as exc:
            logger.error("MatchStorage.get_all: %s", exc)
            return []

    def count(self) -> int:
        if self._con is None:
            return 0
        try:
            with self._lock:
                cur = self._con.execute("SELECT COUNT(*) FROM matches")
                return cur.fetchone()[0]
        except Exception:
            return 0

    def clear_session(self) -> None:
        """Delete all rows (called on new session/shift start)."""
        if self._con is None:
            return
        try:
            with self._lock:
                self._con.execute("DELETE FROM matches")
                self._con.commit()
        except Exception as exc:
            logger.error("MatchStorage.clear_session: %s", exc)

    # ── Export ───────────────────────────────────────────────────────────────

    _HEADERS = [
        "ID", "Дата/время", "Время в пути (сек)", "Вердикт", "Сходство %",
        "Источник камера", "Источник стол", "Источник трек",
        "Запрос камера",   "Запрос стол",   "Запрос трек",
    ]

    def stats_summary(self) -> dict:
        """Агрегированная статистика для вкладки Статистика."""
        empty = {
            "total": 0, "today": 0,
            "avg_sim": 0.0, "avg_transit": 0.0,
            "verdict_high": 0, "verdict_mid": 0, "verdict_low": 0,
            "hourly": [],   # list of (hour_str, count) for last 24h
        }
        if self._con is None:
            return empty
        try:
            import time
            from datetime import datetime, timedelta
            with self._lock:
                cur = self._con.execute("SELECT COUNT(*) FROM matches")
                total = cur.fetchone()[0]

                today_ts = datetime.now().replace(
                    hour=0, minute=0, second=0, microsecond=0
                ).timestamp()
                cur = self._con.execute(
                    "SELECT COUNT(*) FROM matches WHERE timestamp >= ?", (today_ts,)
                )
                today = cur.fetchone()[0]

                cur = self._con.execute(
                    "SELECT AVG(similarity), AVG(transit_sec) FROM matches"
                )
                row = cur.fetchone()
                avg_sim     = (row[0] or 0.0) * 100
                avg_transit = row[1] or 0.0

                cur = self._con.execute(
                    "SELECT verdict, COUNT(*) FROM matches GROUP BY verdict"
                )
                v_high = v_mid = v_low = 0
                for verdict, cnt in cur.fetchall():
                    if "Тот же" in verdict:
                        v_high += cnt
                    elif "Вероятно" in verdict:
                        v_mid += cnt
                    else:
                        v_low += cnt

                # Hourly for last 24h
                since = time.time() - 86400
                cur = self._con.execute(
                    "SELECT timestamp FROM matches WHERE timestamp >= ? ORDER BY timestamp",
                    (since,),
                )
                from collections import defaultdict
                buckets: dict = defaultdict(int)
                for (ts,) in cur.fetchall():
                    h = datetime.fromtimestamp(ts).strftime("%H:00")
                    buckets[h] += 1
                hourly = sorted(buckets.items())

            return {
                "total": total, "today": today,
                "avg_sim": avg_sim, "avg_transit": avg_transit,
                "verdict_high": v_high, "verdict_mid": v_mid, "verdict_low": v_low,
                "hourly": hourly,
            }
        except Exception as exc:
            logger.error("MatchStorage.stats_summary: %s", exc)
            return empty

    def export_csv(self, path: str) -> int:
        """Export all rows to CSV. Returns number of rows written."""
        rows = self.get_all()
        with open(path, "w", newline="", encoding="utf-8-sig") as fh:
            w = csv.writer(fh)
            w.writerow(self._HEADERS)
            for row in rows:
                rid, ts, transit, verdict, sim, sc, sd, st, qc, qd, qt = row
                w.writerow([
                    rid,
                    datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S"),
                    f"{transit:.1f}",
                    verdict,
                    f"{sim * 100:.1f}",
                    sc, sd, st,
                    qc, qd, qt,
                ])
        return len(rows)

    def export_xlsx(self, path: str) -> int:
        """Export all rows to Excel (.xlsx). Returns number of rows written."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            raise RuntimeError("openpyxl не установлен: pip install openpyxl")

        rows = self.get_all()
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Совпадения"

        # Header row
        header_fill = PatternFill("solid", fgColor="065A82")
        header_font = Font(bold=True, color="FFFFFF")
        for col, hdr in enumerate(self._HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=hdr)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Verdict colours
        _fill = {
            "Тот же":   PatternFill("solid", fgColor="D4EDDA"),
            "Вероятно": PatternFill("solid", fgColor="FFF3CD"),
            "Другой":   PatternFill("solid", fgColor="F8D7DA"),
        }

        for r, row in enumerate(rows, 2):
            rid, ts, transit, verdict, sim, sc, sd, st, qc, qd, qt = row
            values = [
                rid,
                datetime.fromtimestamp(ts).strftime("%Y-%m-%d %H:%M:%S"),
                round(transit, 1),
                verdict,
                round(sim * 100, 1),
                sc, sd, st,
                qc, qd, qt,
            ]
            fill_key = next(
                (k for k in _fill if k in verdict), None
            )
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=r, column=col, value=val)
                if fill_key and col == 4:   # verdict column
                    cell.fill = _fill[fill_key]

        # Column widths
        for col, width in enumerate(
            [6, 19, 18, 22, 12, 20, 14, 14, 20, 14, 14], 1
        ):
            ws.column_dimensions[
                openpyxl.utils.get_column_letter(col)
            ].width = width

        wb.save(path)
        return len(rows)
