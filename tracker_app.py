#!/usr/bin/env python3
"""
Airport Baggage Tracker — GUI Application
Run:  python tracker_app.py [--config config.yaml]
"""

# Критично для Windows: импортируем onnxruntime ДО torch чтобы избежать DLL-конфликта.
# onnxruntime и PyTorch CUDA конфликтуют по DLL если torch загружается первым.
try:
    import onnxruntime as _ort_preload  # noqa — должен быть до любого импорта torch
    del _ort_preload
except Exception:
    pass

import argparse
import csv
import logging
import os
import subprocess
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import cv2
import numpy as np
from PyQt5.QtCore import (
    Qt, QSize, QTimer, pyqtSlot, pyqtSignal, QPoint, QMimeData, QEvent,
)
from PyQt5.QtGui import (
    QColor, QDrag, QFont, QIcon, QImage, QPainter,
    QPen, QPixmap,
)
from PyQt5.QtWidgets import (
    QAction, QApplication, QCheckBox, QComboBox,
    QDialog, QDialogButtonBox, QDoubleSpinBox,
    QFileDialog, QFormLayout, QFrame, QGridLayout,
    QGroupBox, QHBoxLayout, QHeaderView, QInputDialog, QLabel,
    QLineEdit, QListWidget, QListWidgetItem,
    QMainWindow, QMenu, QMessageBox,
    QPlainTextEdit, QPushButton, QRadioButton,
    QScrollArea, QSizePolicy, QSlider, QSpinBox,
    QSplitter, QStackedWidget, QStatusBar, QSystemTrayIcon,
    QTabWidget, QTableWidget, QTableWidgetItem, QToolBar,
    QVBoxLayout, QWidget,
)

from tracker_core import (
    AppConfig, CameraEntry, CameraProcessor,
    ProcessorStats, gpu_info, load_config, save_config, COCO_NAMES,
)
from reid_engine import ReIDDatabase, ReIDFeatureExtractor, MatchResult
from web_server import WebDashboard
from batch_engine import BatchInferenceEngine
from match_storage import MatchStorage

logger = logging.getLogger("BaggageTracker.GUI")

def app_dir() -> Path:
    """Папка приложения: рядом с .exe (PyInstaller) или рядом с .py (обычный запуск)."""
    if getattr(sys, "frozen", False):
        return Path(sys.executable).parent
    return Path(__file__).parent

CONFIG_PATH  = str(app_dir() / "config.yaml")
SNAPSHOT_DIR = app_dir() / "Snapshots"


# ── Dark theme QSS ────────────────────────────────────────────────────────────

DARK_QSS = """
QMainWindow, QDialog, QWidget {
    background: #1e1e2e;
    color: #cdd6f4;
    font-family: "Segoe UI", sans-serif;
    font-size: 13px;
}
QTabWidget::pane { border: 1px solid #313244; border-radius: 4px; }
QTabBar::tab {
    background: #313244; color: #a6adc8;
    padding: 7px 18px; border-radius: 4px 4px 0 0; margin-right: 2px;
}
QTabBar::tab:selected { background: #89b4fa; color: #1e1e2e; font-weight: bold; }
QTabBar::tab:hover:!selected { background: #45475a; color: #cdd6f4; }

QPushButton {
    background: #313244; color: #cdd6f4;
    border: 1px solid #45475a; border-radius: 5px;
    padding: 5px 14px; min-height: 26px;
}
QPushButton:hover   { background: #45475a; border-color: #89b4fa; }
QPushButton:pressed { background: #181825; }
QPushButton:disabled { color: #585b70; border-color: #313244; }
QPushButton#btn_primary {
    background: #89b4fa; color: #1e1e2e; font-weight: bold; border: none;
}
QPushButton#btn_primary:hover { background: #b4d0fb; }
QPushButton#btn_danger  { background: #f38ba8; color: #1e1e2e; font-weight: bold; border: none; }
QPushButton#btn_danger:hover  { background: #f5a3b7; }
QPushButton#btn_success { background: #a6e3a1; color: #1e1e2e; font-weight: bold; border: none; }
QPushButton#btn_success:hover { background: #b9edb5; }
QPushButton#btn_warn   { background: #f9e2af; color: #1e1e2e; font-weight: bold; border: none; }
QPushButton#btn_warn:hover    { background: #faeec4; }

QLineEdit, QSpinBox, QDoubleSpinBox, QComboBox {
    background: #313244; color: #cdd6f4;
    border: 1px solid #45475a; border-radius: 4px;
    padding: 4px 8px; min-height: 26px;
}
QLineEdit:focus, QSpinBox:focus, QDoubleSpinBox:focus, QComboBox:focus {
    border-color: #89b4fa;
}
QComboBox::drop-down { border: none; width: 22px; }
QComboBox QAbstractItemView {
    background: #313244; color: #cdd6f4;
    selection-background-color: #45475a;
}

QSlider::groove:horizontal {
    height: 4px; background: #45475a; border-radius: 2px;
}
QSlider::handle:horizontal {
    background: #89b4fa; width: 14px; height: 14px;
    margin: -5px 0; border-radius: 7px;
}
QSlider::sub-page:horizontal { background: #89b4fa; border-radius: 2px; }

QTableWidget {
    background: #1e1e2e; gridline-color: #313244;
    border: 1px solid #313244;
}
QTableWidget::item { padding: 4px 8px; }
QTableWidget::item:selected { background: #313244; color: #89b4fa; }
QHeaderView::section {
    background: #313244; color: #a6adc8;
    padding: 5px 8px; border: none;
    border-bottom: 1px solid #45475a;
}

QGroupBox {
    border: 1px solid #45475a; border-radius: 6px;
    margin-top: 10px; padding-top: 6px;
    font-weight: bold; color: #89b4fa;
}
QGroupBox::title { subcontrol-origin: margin; left: 10px; padding: 0 4px; }

QScrollArea { border: none; }
QScrollBar:vertical {
    background: #1e1e2e; width: 8px; border-radius: 4px;
}
QScrollBar::handle:vertical {
    background: #45475a; border-radius: 4px; min-height: 20px;
}
QScrollBar::handle:vertical:hover { background: #585b70; }
QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical { height: 0; }

QCheckBox::indicator {
    width: 16px; height: 16px; border-radius: 3px;
    border: 1px solid #45475a; background: #313244;
}
QCheckBox::indicator:checked { background: #89b4fa; border-color: #89b4fa; }

QRadioButton::indicator {
    width: 14px; height: 14px; border-radius: 7px;
    border: 1px solid #45475a; background: #313244;
}
QRadioButton::indicator:checked { background: #89b4fa; border-color: #89b4fa; }

QPlainTextEdit {
    background: #181825; color: #a6e3a1;
    font-family: "Consolas", monospace; font-size: 12px;
    border: 1px solid #313244; border-radius: 4px;
}

QToolBar {
    background: #181825; border-bottom: 1px solid #313244;
    spacing: 4px; padding: 3px;
}
QToolBar QToolButton {
    background: transparent; color: #cdd6f4;
    border: 1px solid transparent; border-radius: 4px;
    padding: 4px 10px; font-size: 12px;
}
QToolBar QToolButton:hover   { background: #313244; border-color: #45475a; }
QToolBar QToolButton:pressed { background: #181825; }
QToolBar::separator { background: #45475a; width: 1px; margin: 4px 6px; }

QStatusBar { background: #181825; color: #6c7086; border-top: 1px solid #313244; }
QLabel#hdr { color: #89b4fa; font-weight: bold; font-size: 14px; }
"""


# ── Status dot widget ──────────────────────────────────────────────────────────

STATUS_COLORS = {
    "stopped":    "#6c7086",
    "connecting": "#f9e2af",
    "running":    "#a6e3a1",
    "paused":     "#fab387",
    "error":      "#f38ba8",
}
STATUS_LABELS = {
    "stopped": "Остановлен", "connecting": "Подключение…",
    "running": "Работает",   "paused": "Пауза", "error": "Ошибка",
}


class StatusDot(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._color = QColor(STATUS_COLORS["stopped"])
        self.setFixedSize(12, 12)

    def set_status(self, status: str):
        self._color = QColor(STATUS_COLORS.get(status, "#6c7086"))
        self.update()

    def paintEvent(self, _):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        p.setBrush(self._color)
        p.setPen(Qt.NoPen)
        p.drawEllipse(1, 1, 10, 10)


# ── Drag handle widget ────────────────────────────────────────────────────────

class _DragHandle(QLabel):
    """⠿ grip widget that initiates a QDrag for its parent CameraFeedWidget."""

    def __init__(self, cam_widget: "CameraFeedWidget"):
        super().__init__("⠿")
        self._w = cam_widget
        self._start: Optional[QPoint] = None
        self.setCursor(Qt.SizeAllCursor)
        self.setFixedSize(20, 28)
        self.setToolTip("Перетащить камеру")
        self.setStyleSheet("color:#585b70; font-size:16px; padding:0;")

    def mousePressEvent(self, e):
        if e.button() == Qt.LeftButton:
            self._start = e.pos()

    def mouseMoveEvent(self, e):
        if not (e.buttons() & Qt.LeftButton) or self._start is None:
            return
        if (e.pos() - self._start).manhattanLength() < QApplication.startDragDistance():
            return
        drag = QDrag(self)
        mime = QMimeData()
        mime.setText(str(self._w._cam_index))
        drag.setMimeData(mime)
        # Thumbnail preview
        pm = QPixmap(120, 80)
        pm.fill(QColor("#313244"))
        if self._w._last_raw_frame is not None:
            h, w = self._w._last_raw_frame.shape[:2]
            rgb = cv2.cvtColor(self._w._last_raw_frame, cv2.COLOR_BGR2RGB)
            qi  = QImage(rgb.data, w, h, w * 3, QImage.Format_RGB888)
            pm  = QPixmap.fromImage(qi).scaled(
                120, 80, Qt.KeepAspectRatio, Qt.SmoothTransformation
            )
        drag.setPixmap(pm)
        drag.setHotSpot(QPoint(pm.width() // 2, pm.height() // 2))
        drag.exec_(Qt.MoveAction)
        self._start = None

    def mouseReleaseEvent(self, _):
        self._start = None


# ── Scalable video label ───────────────────────────────────────────────────────

class VideoLabel(QLabel):
    """QLabel that rescales its stored QPixmap on every resize.
    ROI mode: кликай по видео — каждый ЛКМ добавляет вершину полигона.
    ПКМ — удаляет последнюю вершину. Двойной ЛКМ — закрывает и сохраняет
    полигон (нужно минимум 3 точки). ESC / set_roi_mode(False) — сброс."""

    roi_drawn = pyqtSignal(list)   # [[x,y], [x,y], ...] нормализованные [0..1]

    def __init__(self, parent=None):
        super().__init__(parent)
        self._source_pmap: Optional[QPixmap] = None
        self._src_w = 0
        self._src_h = 0
        self._roi_mode = False
        self._poly_pts: List[QPoint] = []   # вершины рисуемого полигона (widget-coords)
        self.setAlignment(Qt.AlignCenter)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.setMinimumSize(1, 1)
        self.setText("Нет сигнала")
        self.setStyleSheet("color:#45475a; background:#000; border-radius:4px; font-size:14px;")

    def set_pixmap(self, pmap: QPixmap):
        self._source_pmap = pmap
        self._src_w = pmap.width()
        self._src_h = pmap.height()
        self._redraw()

    def set_roi_mode(self, on: bool):
        self._roi_mode = on
        if not on:
            self._poly_pts.clear()
            self.update()
        self.setCursor(Qt.CrossCursor if on else Qt.ArrowCursor)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._redraw()

    def _redraw(self):
        if self._source_pmap and not self._source_pmap.isNull():
            scaled = self._source_pmap.scaled(
                self.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation
            )
            super().setPixmap(scaled)

    # ── Polygon ROI drawing ───────────────────────────────────────────────────

    def mousePressEvent(self, e):
        if self._roi_mode:
            if e.button() == Qt.LeftButton:
                self._poly_pts.append(e.pos())
                self.update()
            elif e.button() == Qt.RightButton and self._poly_pts:
                self._poly_pts.pop()          # отменить последнюю точку
                self.update()
        else:
            super().mousePressEvent(e)

    def mouseDoubleClickEvent(self, e):
        if self._roi_mode and e.button() == Qt.LeftButton:
            # двойной клик добавляет ещё одну точку — убираем дубликат
            if self._poly_pts:
                self._poly_pts.pop()
            if len(self._poly_pts) >= 3:
                self._emit_polygon()
                self._poly_pts.clear()
                self.update()
        else:
            super().mouseDoubleClickEvent(e)

    def keyPressEvent(self, e):
        if self._roi_mode and e.key() == Qt.Key_Escape:
            self._poly_pts.clear()
            self.update()
        else:
            super().keyPressEvent(e)

    def paintEvent(self, e):
        super().paintEvent(e)
        if not self._roi_mode or not self._poly_pts:
            return
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        color = QColor("#f9e2af")
        pen = QPen(color, 2)
        p.setPen(pen)

        # Рёбра полигона
        n = len(self._poly_pts)
        for i in range(n - 1):
            p.drawLine(self._poly_pts[i], self._poly_pts[i + 1])
        # Замыкающее ребро (серое, предпросмотр)
        if n >= 3:
            close_pen = QPen(QColor("#a6adc8"), 1, Qt.DashLine)
            p.setPen(close_pen)
            p.drawLine(self._poly_pts[-1], self._poly_pts[0])
            p.setPen(pen)

        # Вершины
        p.setBrush(color)
        p.setPen(Qt.NoPen)
        for pt in self._poly_pts:
            p.drawEllipse(pt.x() - 4, pt.y() - 4, 8, 8)

        # Подсказка
        p.setPen(QPen(QColor("#cdd6f4")))
        hint = (f"  {n} точек — двойной клик чтобы закрыть  |  ПКМ — отмена"
                if n >= 3 else f"  {n} точек — нужно минимум 3")
        p.drawText(4, self.height() - 6, hint)

    def _emit_polygon(self):
        if not self._src_w or not self._src_h or len(self._poly_pts) < 3:
            return
        W, H = self.width(), self.height()
        scale  = min(W / self._src_w, H / self._src_h)
        disp_w = self._src_w * scale
        disp_h = self._src_h * scale
        off_x  = (W - disp_w) / 2
        off_y  = (H - disp_h) / 2

        normalized = []
        for pt in self._poly_pts:
            nx = max(0.0, min(1.0, (pt.x() - off_x) / disp_w))
            ny = max(0.0, min(1.0, (pt.y() - off_y) / disp_h))
            normalized.append([round(nx, 4), round(ny, 4)])
        self.roi_drawn.emit(normalized)


# ── Fullscreen single-camera dialog ───────────────────────────────────────────

class FullscreenCameraDialog(QDialog):
    """Maximized window showing one camera feed. ESC or double-click to close."""

    def __init__(self, feed_widget: "CameraFeedWidget", parent=None):
        super().__init__(parent)
        self.setWindowTitle(feed_widget.cam.name)
        self.setWindowFlags(Qt.Window | Qt.WindowMaximizeButtonHint | Qt.WindowCloseButtonHint)
        self.setStyleSheet("background:#000;")
        self.showMaximized()

        self._feed = feed_widget
        self._video = VideoLabel(self)
        self._video.setStyleSheet("background:#000;")

        # Copy last frame if available
        if feed_widget._last_raw_frame is not None:
            feed_widget.on_frame(feed_widget._last_raw_frame)

        # Connect to live feed
        self._feed._fullscreen_target = self._video

        # Status bar at bottom
        bar = QLabel()
        bar.setFixedHeight(28)
        bar.setStyleSheet(
            "background:#181825; color:#a6adc8; font-size:12px; padding: 0 12px;"
        )
        bar.setText(f"  {feed_widget.cam.name}  —  двойной клик или ESC для закрытия")
        bar.setAlignment(Qt.AlignVCenter | Qt.AlignLeft)

        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)
        lay.addWidget(self._video, 1)
        lay.addWidget(bar)

    def mouseDoubleClickEvent(self, _):
        self.close()

    def keyPressEvent(self, e):
        if e.key() == Qt.Key_Escape:
            self.close()
        else:
            super().keyPressEvent(e)

    def closeEvent(self, event):
        self._feed._fullscreen_target = None
        super().closeEvent(event)


# ── Camera feed widget (one tile in the monitor grid) ─────────────────────────

class CameraFeedWidget(QFrame):
    """Shows a single camera: live annotated feed + stats. Scalable tile."""

    double_clicked = pyqtSignal()
    roi_changed    = pyqtSignal(object)   # emits CameraEntry after ROI set/cleared

    def __init__(self, cam: CameraEntry, parent=None):
        super().__init__(parent)
        self.cam = cam
        self._paused   = False
        self._compact  = False
        self._hidden   = False
        self._side_visible = True
        self._cam_index = 0          # set by MonitorTab
        self._last_raw_frame: Optional[np.ndarray] = None
        self._fullscreen_target: Optional[VideoLabel] = None

        self.setFrameShape(QFrame.StyledPanel)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.setMinimumSize(240, 180)
        self.setAcceptDrops(True)
        self.setStyleSheet(
            "CameraFeedWidget { background:#11111b; border:1px solid #313244; border-radius:6px; }"
        )
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Header bar ────────────────────────────────────────────────────────
        self._hdr = QWidget()
        self._hdr.setFixedHeight(34)
        self._hdr.setStyleSheet("background:#181825; border-radius:6px 6px 0 0;")
        hl = QHBoxLayout(self._hdr)
        hl.setContentsMargins(8, 0, 6, 0)
        hl.setSpacing(6)

        hl.addWidget(_DragHandle(self))

        self._dot = StatusDot()
        hl.addWidget(self._dot)

        self._name_lbl = QLabel(self.cam.name)
        self._name_lbl.setStyleSheet("color:#cdd6f4; font-weight:bold; font-size:13px;")
        hl.addWidget(self._name_lbl, 1)

        mode_badge = QLabel("📁 FILE" if self.cam.mode == "file" else "📡 RTSP")
        mode_badge.setStyleSheet(
            "color:#1e1e2e; background:#f9e2af; border-radius:3px;"
            "padding:1px 5px; font-size:10px; font-weight:bold;"
            if self.cam.mode == "file" else
            "color:#1e1e2e; background:#89dceb; border-radius:3px;"
            "padding:1px 5px; font-size:10px; font-weight:bold;"
        )
        hl.addWidget(mode_badge)

        self._roi_btn = QPushButton("⬚")
        self._roi_btn.setFixedSize(28, 28)
        self._roi_btn.setCheckable(True)
        self._roi_btn.setToolTip(
            "Нарисовать ROI-полигон:\n"
            "  ЛКМ — добавить вершину\n"
            "  ПКМ — удалить последнюю\n"
            "  Двойной клик — закрыть полигон (мин. 3 точки)\n"
            "  ESC — сбросить"
        )
        self._roi_btn.clicked.connect(self._toggle_roi_mode)
        hl.addWidget(self._roi_btn)

        self._hide_btn = QPushButton("👁")
        self._hide_btn.setFixedSize(28, 28)
        self._hide_btn.setCheckable(True)
        self._hide_btn.setToolTip("Скрыть видео (камера продолжает работать)")
        self._hide_btn.clicked.connect(self._toggle_hidden)
        hl.addWidget(self._hide_btn)

        self._side_btn = QPushButton("◀")
        self._side_btn.setFixedSize(24, 24)
        self._side_btn.setToolTip("Скрыть/Показать панель статистики")
        self._side_btn.clicked.connect(self._toggle_side)
        hl.addWidget(self._side_btn)

        self._pause_btn = QPushButton("⏸")
        self._pause_btn.setFixedSize(28, 28)
        self._pause_btn.setToolTip("Пауза / Продолжить")
        self._pause_btn.clicked.connect(self._toggle_pause)
        hl.addWidget(self._pause_btn)

        self._snap_btn = QPushButton("📷")
        self._snap_btn.setFixedSize(28, 28)
        self._snap_btn.setToolTip("Сохранить снимок")
        self._snap_btn.clicked.connect(self._snapshot)
        hl.addWidget(self._snap_btn)

        root.addWidget(self._hdr)

        # ── Video + side panel row ────────────────────────────────────────────
        self._mid = QHBoxLayout()
        self._mid.setContentsMargins(4, 4, 4, 4)
        self._mid.setSpacing(6)

        self._video_lbl = VideoLabel()
        self._video_lbl.roi_drawn.connect(self._on_roi_drawn)
        self._mid.addWidget(self._video_lbl, 1)

        self._hidden_ph = QLabel("👁 Видео скрыто\n(камера работает)")
        self._hidden_ph.setAlignment(Qt.AlignCenter)
        self._hidden_ph.setStyleSheet(
            "color:#45475a; background:#000; border-radius:4px; font-size:13px; line-height:1.6;"
        )
        self._hidden_ph.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self._hidden_ph.setVisible(False)
        self._mid.addWidget(self._hidden_ph, 1)

        # ── Side panel ────────────────────────────────────────────────────────
        self._side = QFrame()
        self._side.setFixedWidth(114)
        self._side.setStyleSheet("QFrame{background:#181825;border-radius:4px;}")
        sv = QVBoxLayout(self._side)
        sv.setContentsMargins(6, 6, 6, 6)
        sv.setSpacing(4)

        self._crop_lbl = QLabel()
        self._crop_lbl.setFixedSize(98, 98)
        self._crop_lbl.setAlignment(Qt.AlignCenter)
        self._crop_lbl.setStyleSheet(
            "background:#11111b; border:1px solid #313244; border-radius:4px;"
        )
        sv.addWidget(self._crop_lbl)

        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet("color:#313244;")
        sv.addWidget(sep)

        self._desk_lbl   = QLabel(f"Стол #{self.cam.counter_id}")
        self._active_lbl = QLabel("Активно:  0")
        self._seen_lbl   = QLabel("Треков:   0")
        self._yolo_lbl   = QLabel("YOLO:     0")
        self._reid_lbl   = QLabel("ReID:     0")
        self._fps_lbl    = QLabel("FPS:      —")
        # Счётчик отсортированного багажа — виден только на query-камерах
        self._sorted_lbl = QLabel("Багаж:    0")

        self._desk_lbl.setStyleSheet("color:#89b4fa; font-size:11px; font-weight:bold;")
        for lbl in (self._active_lbl, self._seen_lbl,
                    self._yolo_lbl, self._reid_lbl, self._fps_lbl):
            lbl.setStyleSheet("color:#a6adc8; font-size:11px; font-family:Consolas;")
            lbl.setAlignment(Qt.AlignLeft)
        self._sorted_lbl.setStyleSheet(
            "color:#a6e3a1; font-size:11px; font-family:Consolas; font-weight:bold;"
        )
        self._sorted_lbl.setAlignment(Qt.AlignLeft)
        self._sorted_lbl.setVisible(self.cam.role == "query")

        for lbl in (self._desk_lbl, self._active_lbl, self._seen_lbl,
                    self._yolo_lbl, self._reid_lbl, self._fps_lbl, self._sorted_lbl):
            sv.addWidget(lbl)

        sv.addStretch()
        self._mid.addWidget(self._side)
        root.addLayout(self._mid, 1)

        # ── Footer status bar ─────────────────────────────────────────────────
        ftr = QWidget()
        ftr.setFixedHeight(24)
        ftr.setStyleSheet("background:#181825; border-radius:0 0 6px 6px;")
        fl = QHBoxLayout(ftr)
        fl.setContentsMargins(10, 0, 10, 0)
        self._status_lbl = QLabel("● Остановлен")
        self._status_lbl.setStyleSheet("color:#6c7086; font-size:11px;")
        fl.addWidget(self._status_lbl, 1)
        self._mode_lbl = QLabel("")
        self._mode_lbl.setStyleSheet("color:#45475a; font-size:10px;")
        fl.addWidget(self._mode_lbl)
        root.addWidget(ftr)

    # ── Public API ────────────────────────────────────────────────────────────

    def set_compact(self, compact: bool):
        """Hide header and footer for compact/thumbnail view."""
        self._compact = compact
        self._hdr.setVisible(not compact)

    def toggle_side_panel(self):
        self._toggle_side()

    # ── Mouse events ──────────────────────────────────────────────────────────

    def mouseDoubleClickEvent(self, event):
        if event.button() == Qt.LeftButton:
            dlg = FullscreenCameraDialog(self, self.window())
            dlg.exec_()
        super().mouseDoubleClickEvent(event)

    def contextMenuEvent(self, event: "QContextMenuEvent"):
        menu = QMenu(self)
        menu.setStyleSheet(
            "QMenu{background:#313244;color:#cdd6f4;border:1px solid #45475a;border-radius:4px;}"
            "QMenu::item{padding:5px 20px;}"
            "QMenu::item:selected{background:#45475a;}"
        )
        act_fs      = menu.addAction("⛶  Открыть во весь экран")
        act_pause   = menu.addAction("⏸  Пауза" if not self._paused else "▶  Продолжить")
        act_snap    = menu.addAction("📷  Сохранить снимок")
        menu.addSeparator()
        act_side    = menu.addAction("◀  Скрыть статистику" if self._side_visible else "▶  Показать статистику")
        has_roi = bool(self.cam.roi)
        act_roi     = menu.addAction("⬚  Нарисовать ROI")
        act_roi_clr = menu.addAction("✖  Сбросить ROI")
        act_roi_clr.setEnabled(has_roi)
        act_hide    = menu.addAction("🚫  Скрыть видео" if not self._hidden else "👁  Показать видео")
        act_compact = menu.addAction("▦  Компактный вид" if not self._compact else "▣  Обычный вид")
        menu.addSeparator()
        act_info    = menu.addAction(f"ℹ  {self.cam.name}  (Стол #{self.cam.counter_id})")
        act_info.setEnabled(False)

        chosen = menu.exec_(event.globalPos())
        if chosen == act_fs:
            dlg = FullscreenCameraDialog(self, self.window())
            dlg.exec_()
        elif chosen == act_pause:
            self._toggle_pause()
        elif chosen == act_snap:
            self._snapshot()
        elif chosen == act_roi:
            self._roi_btn.setChecked(True)
            self._toggle_roi_mode()
        elif chosen == act_roi_clr:
            self.cam.roi = None
            self.roi_changed.emit(self.cam)
        elif chosen == act_side:
            self._toggle_side()
        elif chosen == act_hide:
            self._toggle_hidden()
        elif chosen == act_compact:
            self.set_compact(not self._compact)

    # ── Slots ─────────────────────────────────────────────────────────────────

    @pyqtSlot(object)
    def on_frame(self, frame: np.ndarray):
        self._last_raw_frame = frame
        if self._hidden:
            return
        h, w = frame.shape[:2]
        rgb  = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        qimg = QImage(rgb.data, w, h, w * 3, QImage.Format_RGB888)
        pmap = QPixmap.fromImage(qimg)
        self._video_lbl.set_pixmap(pmap)
        if self._fullscreen_target is not None:
            self._fullscreen_target.set_pixmap(pmap)

    @pyqtSlot(object)
    def on_crop(self, crop: np.ndarray):
        h, w = crop.shape[:2]
        rgb  = cv2.cvtColor(crop, cv2.COLOR_BGR2RGB)
        qimg = QImage(rgb.data, w, h, w * 3, QImage.Format_RGB888)
        pmap = QPixmap.fromImage(qimg).scaled(
            QSize(96, 96), Qt.KeepAspectRatio, Qt.SmoothTransformation
        )
        self._crop_lbl.setPixmap(pmap)

    @pyqtSlot(object)
    def on_stats(self, stats: ProcessorStats):
        self._dot.set_status(stats.status)
        color = STATUS_COLORS.get(stats.status, "#6c7086")
        label = STATUS_LABELS.get(stats.status, stats.status)
        self._status_lbl.setText(f"● {label}")
        self._status_lbl.setStyleSheet(f"color:{color}; font-size:11px;")
        self._active_lbl.setText(f"Активно:  {stats.active_count}")
        self._seen_lbl.setText  (f"Треков:   {stats.total_seen}")
        self._yolo_lbl.setText  (f"YOLO:     {stats.yolo_saved}")
        self._reid_lbl.setText  (f"ReID:     {stats.reid_saved}")
        self._fps_lbl.setText   (f"FPS:      {stats.fps:.1f}")
        if self.cam.role == "query":
            self._sorted_lbl.setText(f"Багаж:    {stats.bags_sorted}")

    def refresh_cam_info(self, cam: CameraEntry):
        self.cam = cam
        self._name_lbl.setText(cam.name)
        self._desk_lbl.setText(f"Стол #{cam.counter_id}")
        self._mode_lbl.setText("📁 Файл" if cam.mode == "file" else "📡 RTSP")
        self._sorted_lbl.setVisible(cam.role == "query")

    # ── Drag & drop target ────────────────────────────────────────────────────

    def dragEnterEvent(self, e):
        if e.mimeData().hasText():
            e.acceptProposedAction()
            self.setStyleSheet(
                "CameraFeedWidget { background:#11111b; border:2px solid #89b4fa; border-radius:6px; }"
            )

    def dragLeaveEvent(self, _):
        self._restore_border()

    def dropEvent(self, e):
        self._restore_border()
        txt = e.mimeData().text()
        if txt.isdigit():
            src = int(txt)
            if src != self._cam_index:
                # Walk up to find MonitorTab
                p = self.parent()
                while p and not isinstance(p, MonitorTab):
                    p = p.parent()
                if p:
                    p.swap_cameras(src, self._cam_index)
        e.acceptProposedAction()

    def _restore_border(self):
        self.setStyleSheet(
            "CameraFeedWidget { background:#11111b; border:1px solid #313244; border-radius:6px; }"
        )

    # ── Private ───────────────────────────────────────────────────────────────

    def _toggle_roi_mode(self):
        on = self._roi_btn.isChecked()
        self._video_lbl.set_roi_mode(on)
        self._roi_btn.setStyleSheet(
            "background:#f9e2af; color:#1e1e2e;" if on else ""
        )
        if on:
            QMessageBox.information(
                self, "Рисование ROI-полигона",
                "Кликайте ЛКМ по видео — каждый клик добавляет вершину.\n"
                "ПКМ — удалить последнюю вершину.\n"
                "Двойной клик — закрыть полигон (минимум 3 точки).\n"
                "ESC — сбросить все точки.\n\n"
                "Правый клик на камере → «Сбросить ROI» для удаления."
            )

    @pyqtSlot(list)
    def _on_roi_drawn(self, roi: list):
        self.cam.roi = roi
        self._video_lbl.set_roi_mode(False)
        self._roi_btn.setChecked(False)
        self._roi_btn.setStyleSheet("")
        # Notify parent to save config
        self.roi_changed.emit(self.cam)

    def _toggle_hidden(self):
        self._hidden = not self._hidden
        self._video_lbl.setVisible(not self._hidden)
        self._hidden_ph.setVisible(self._hidden)
        self._hide_btn.setChecked(self._hidden)
        if not self._hidden and self._last_raw_frame is not None:
            # Refresh immediately when un-hiding
            self.on_frame(self._last_raw_frame)

    def _toggle_side(self):
        self._side_visible = not self._side_visible
        self._side.setVisible(self._side_visible)
        self._side_btn.setText("▶" if not self._side_visible else "◀")

    def _toggle_pause(self):
        self._paused = not self._paused
        self._pause_btn.setText("▶" if self._paused else "⏸")
        self.parent_pause_cb(self._paused)

    def parent_pause_cb(self, paused: bool):
        pass

    def _snapshot(self):
        if self._last_raw_frame is None:
            QMessageBox.information(self, "Снимок", "Нет кадра для сохранения.")
            return
        SNAPSHOT_DIR.mkdir(exist_ok=True)
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = SNAPSHOT_DIR / f"{self.cam.name.replace(' ', '_')}_{ts}.jpg"
        cv2.imwrite(str(path), self._last_raw_frame, [cv2.IMWRITE_JPEG_QUALITY, 95])
        QMessageBox.information(self, "Снимок сохранён", str(path))


# ── Add / Edit camera dialog ───────────────────────────────────────────────────

class AddEditCameraDialog(QDialog):
    def __init__(
        self,
        cam: Optional[CameraEntry] = None,
        all_cameras: list = None,
        parent=None,
    ):
        super().__init__(parent)
        self.setWindowTitle("Добавить камеру" if cam is None else "Редактировать камеру")
        self.setMinimumWidth(500)
        self.setModal(True)
        self._entry = cam or CameraEntry()
        # Other cameras for receives_from selector (exclude self)
        self._other_cams = [c for c in (all_cameras or []) if c is not self._entry]
        self._build_ui()
        self._populate()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setSpacing(8)

        tabs = QTabWidget()
        root.addWidget(tabs)

        # ── Tab 1: Основное ───────────────────────────────────────────────────
        basic_w = QWidget()
        basic_layout = QVBoxLayout(basic_w)
        basic_layout.setSpacing(12)

        form = QFormLayout()
        form.setSpacing(8)

        # Camera name
        self._name_edit = QLineEdit()
        form.addRow("Название камеры:", self._name_edit)

        # Camera ID (unique routing key)
        self._camid_spin = QSpinBox()
        self._camid_spin.setRange(1, 9999)
        self._camid_spin.setToolTip(
            "Уникальный ID камеры — используется для маршрутизации ReID.\n"
            "Каждая камера должна иметь отдельный ID."
        )
        form.addRow("ID камеры:", self._camid_spin)

        # Role
        self._role_combo = QComboBox()
        self._role_combo.addItem("Source — стойка регистрации (добавляет в ReID БД)", "source")
        self._role_combo.addItem("Transit — промежуточная камера (добавляет в ReID БД)", "transit")
        self._role_combo.addItem("Query  — лента сортировки (определяет стойку по ReID)", "query")
        form.addRow("Роль камеры:", self._role_combo)

        # Desk # — only for source cameras
        self._desk_label = QLabel("Номер стойки (Desk #):")
        self._desk_spin  = QSpinBox()
        self._desk_spin.setRange(1, 999)
        form.addRow(self._desk_label, self._desk_spin)

        # Receives From — for transit and query cameras
        self._from_label = QLabel("Получать данные от камеры:")
        self._from_combo = QComboBox()
        self._from_combo.addItem("— Любая (не фильтровать) —", None)
        for c in self._other_cams:
            label = f"ID {c.cam_id}  «{c.name}»  [{c.role}]"
            self._from_combo.addItem(label, c.cam_id)
        self._from_label.setToolTip(
            "Матчинг только против записей от этой камеры.\n"
            "Оставьте «Любая» если нет транзитных камер."
        )
        self._from_combo.setToolTip(self._from_label.toolTip())
        form.addRow(self._from_label, self._from_combo)

        basic_layout.addLayout(form)

        # Mode selection
        mode_grp = QGroupBox("Режим источника")
        ml = QVBoxLayout(mode_grp)
        ml.setSpacing(6)
        self._rtsp_radio = QRadioButton("📡  RTSP камера (PoE)")
        self._file_radio = QRadioButton("📁  Видеофайл (тестовый режим)")
        self._rtsp_radio.setChecked(True)
        ml.addWidget(self._rtsp_radio)
        ml.addWidget(self._file_radio)
        basic_layout.addWidget(mode_grp)

        # Stacked panels
        self._stack = QStackedWidget()

        rtsp_w = QWidget()
        rl = QFormLayout(rtsp_w)
        rl.setSpacing(8)
        self._url_edit = QLineEdit()
        self._url_edit.setPlaceholderText("rtsp://user:pass@192.168.1.100:554/stream1")
        rl.addRow("RTSP URL:", self._url_edit)
        self._test_btn = QPushButton("🔗  Проверить подключение")
        self._test_btn.setObjectName("btn_primary")
        self._test_btn.clicked.connect(self._test_connection)
        rl.addRow("", self._test_btn)
        self._stack.addWidget(rtsp_w)

        file_w = QWidget()
        fl = QFormLayout(file_w)
        fl.setSpacing(8)
        fp_row = QHBoxLayout()
        self._file_edit = QLineEdit()
        self._file_edit.setPlaceholderText("Путь к видеофайлу…")
        browse_btn = QPushButton("📂  Обзор")
        browse_btn.clicked.connect(self._browse_file)
        fp_row.addWidget(self._file_edit)
        fp_row.addWidget(browse_btn)
        fl.addRow("Видеофайл:", fp_row)
        self._loop_chk = QCheckBox("Зациклить воспроизведение")
        self._loop_chk.setChecked(True)
        fl.addRow("", self._loop_chk)
        self._stack.addWidget(file_w)

        basic_layout.addWidget(self._stack)

        self._rtsp_radio.toggled.connect(lambda on: self._stack.setCurrentIndex(0 if on else 1))
        self._file_radio.toggled.connect(lambda on: self._stack.setCurrentIndex(1 if on else 0))
        self._role_combo.currentIndexChanged.connect(self._on_role_changed)

        tabs.addTab(basic_w, "Основное")

        # ── Tab 2: Настройки камеры ───────────────────────────────────────────
        cam_settings_w = self._build_cam_settings_tab()
        tabs.addTab(cam_settings_w, "Настройки камеры")

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.button(QDialogButtonBox.Ok).setText("✔  Сохранить")
        btns.button(QDialogButtonBox.Cancel).setText("Отмена")
        btns.accepted.connect(self._accept)
        btns.rejected.connect(self.reject)
        root.addWidget(btns)

    def _make_override_row(self, label_text: str, widget: QWidget, chk_attr: str):
        """
        Returns (row_layout, global_checkbox).
        The global_checkbox disables the widget when checked (= use global value).
        """
        row = QHBoxLayout()
        lbl = QLabel(label_text)
        lbl.setMinimumWidth(200)
        chk = QCheckBox("Глобальные")
        chk.setToolTip("Использовать глобальное значение из Настройки → Глобальные")
        chk.setChecked(True)
        widget.setEnabled(False)
        chk.toggled.connect(lambda on: widget.setEnabled(not on))
        setattr(self, chk_attr, chk)
        row.addWidget(lbl)
        row.addWidget(chk)
        row.addWidget(widget)
        row.addStretch()
        return row, chk

    def _build_cam_settings_tab(self) -> QWidget:
        w = QWidget()
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea{border:none;}")
        inner = QWidget()
        v = QVBoxLayout(inner)
        v.setSpacing(8)
        v.setContentsMargins(12, 12, 12, 12)

        grp = QGroupBox("Переопределения для этой камеры")
        gv = QVBoxLayout(grp)
        gv.setSpacing(6)

        hint = QLabel("Оставьте 'Глобальные' чтобы использовать настройки из Настройки → Глобальные")
        hint.setStyleSheet("color:#a6adc8; font-size:11px;")
        hint.setWordWrap(True)
        gv.addWidget(hint)

        # Confidence
        self._cam_conf_spin = QDoubleSpinBox()
        self._cam_conf_spin.setRange(0.00, 1.00)
        self._cam_conf_spin.setSingleStep(0.05)
        self._cam_conf_spin.setDecimals(2)
        row, _ = self._make_override_row("Confidence:", self._cam_conf_spin, "_cam_conf_chk")
        gv.addLayout(row)

        # IoU
        self._cam_iou_spin = QDoubleSpinBox()
        self._cam_iou_spin.setRange(0.00, 1.00)
        self._cam_iou_spin.setSingleStep(0.05)
        self._cam_iou_spin.setDecimals(2)
        row, _ = self._make_override_row("IoU:", self._cam_iou_spin, "_cam_iou_chk")
        gv.addLayout(row)

        # Infer every N
        self._cam_infer_n_spin = QSpinBox()
        self._cam_infer_n_spin.setRange(1, 20)
        row, _ = self._make_override_row("YOLO каждые N кадров:", self._cam_infer_n_spin, "_cam_infer_n_chk")
        gv.addLayout(row)

        # YOLO imgsz
        self._cam_imgsz_combo = QComboBox()
        for sz in (256, 320, 416, 512, 640, 736, 800, 1280):
            self._cam_imgsz_combo.addItem(str(sz), sz)
        row, _ = self._make_override_row("YOLO imgsz:", self._cam_imgsz_combo, "_cam_imgsz_chk")
        gv.addLayout(row)

        # ReID every N
        self._cam_reid_n_spin = QSpinBox()
        self._cam_reid_n_spin.setRange(1, 30)
        row, _ = self._make_override_row("ReID каждые N:", self._cam_reid_n_spin, "_cam_reid_n_chk")
        gv.addLayout(row)

        # ReID min crop px
        self._cam_reid_min_px_spin = QSpinBox()
        self._cam_reid_min_px_spin.setRange(16, 256)
        self._cam_reid_min_px_spin.setSuffix(" px")
        row, _ = self._make_override_row("Мин. размер кропа ReID (px):", self._cam_reid_min_px_spin, "_cam_reid_min_px_chk")
        gv.addLayout(row)

        # Motion detect
        self._cam_motion_chk_val = QCheckBox("Включить")
        row, _ = self._make_override_row("Детектор движения:", self._cam_motion_chk_val, "_cam_motion_chk")
        gv.addLayout(row)

        # Motion min area
        self._cam_motion_area_spin = QSpinBox()
        self._cam_motion_area_spin.setRange(100, 100000)
        self._cam_motion_area_spin.setSingleStep(500)
        self._cam_motion_area_spin.setSuffix(" px²")
        row, _ = self._make_override_row("Мин. площадь движения:", self._cam_motion_area_spin, "_cam_motion_area_chk")
        gv.addLayout(row)

        # Classes filter
        self._cam_classes_edit = QLineEdit()
        self._cam_classes_edit.setPlaceholderText("24, 26, 28, 80")
        self._cam_classes_edit.setToolTip(
            "ID классов через запятую. Custom: '80: box, 81: cart'"
        )
        classes_col = QVBoxLayout()
        classes_row, _ = self._make_override_row("Фильтр классов:", self._cam_classes_edit, "_cam_classes_chk")
        classes_hint = QLabel("ID классов через запятую. Custom: '80: box, 81: cart'")
        classes_hint.setStyleSheet("color:#6c7086; font-size:10px;")
        gv.addLayout(classes_row)
        gv.addWidget(classes_hint)

        v.addWidget(grp)
        v.addStretch()
        scroll.setWidget(inner)

        outer = QVBoxLayout(w)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.addWidget(scroll)
        return w

    def _on_role_changed(self):
        role = self._role_combo.currentData()
        is_source = (role == "source")
        self._desk_label.setVisible(is_source)
        self._desk_spin.setVisible(is_source)
        self._from_label.setVisible(not is_source)
        self._from_combo.setVisible(not is_source)

    def _populate(self):
        e = self._entry
        self._name_edit.setText(e.name)
        self._camid_spin.setValue(e.cam_id if e.cam_id > 0 else 1)
        self._desk_spin.setValue(e.counter_id)
        self._url_edit.setText(e.rtsp_url)
        self._file_edit.setText(e.file_path)
        self._loop_chk.setChecked(e.loop_video)
        role_idx = self._role_combo.findData(e.role)
        self._role_combo.setCurrentIndex(max(0, role_idx))
        # Populate receives_from
        if e.receives_from is not None:
            idx = self._from_combo.findData(e.receives_from)
            if idx >= 0:
                self._from_combo.setCurrentIndex(idx)
        self._on_role_changed()
        if e.mode == "file":
            self._file_radio.setChecked(True)
            self._stack.setCurrentIndex(1)
        else:
            self._rtsp_radio.setChecked(True)
            self._stack.setCurrentIndex(0)

        # ── Per-camera settings tab ───────────────────────────────────────────
        def _set_override(chk, widget, value, setter):
            if value is None:
                chk.setChecked(True)   # global
                widget.setEnabled(False)
            else:
                chk.setChecked(False)  # override
                widget.setEnabled(True)
                setter(value)

        _set_override(self._cam_conf_chk, self._cam_conf_spin,
                      e.cam_confidence, self._cam_conf_spin.setValue)
        _set_override(self._cam_iou_chk, self._cam_iou_spin,
                      e.cam_iou, self._cam_iou_spin.setValue)
        _set_override(self._cam_infer_n_chk, self._cam_infer_n_spin,
                      e.cam_infer_every_n, self._cam_infer_n_spin.setValue)
        # imgsz combo
        if e.cam_infer_imgsz is None:
            self._cam_imgsz_chk.setChecked(True)
            self._cam_imgsz_combo.setEnabled(False)
        else:
            self._cam_imgsz_chk.setChecked(False)
            self._cam_imgsz_combo.setEnabled(True)
            idx = self._cam_imgsz_combo.findData(e.cam_infer_imgsz)
            if idx >= 0:
                self._cam_imgsz_combo.setCurrentIndex(idx)
        _set_override(self._cam_reid_n_chk, self._cam_reid_n_spin,
                      e.cam_reid_every_n, self._cam_reid_n_spin.setValue)
        _set_override(self._cam_reid_min_px_chk, self._cam_reid_min_px_spin,
                      e.cam_reid_min_crop_px, self._cam_reid_min_px_spin.setValue)
        _set_override(self._cam_motion_chk, self._cam_motion_chk_val,
                      e.cam_motion_detect, self._cam_motion_chk_val.setChecked)
        _set_override(self._cam_motion_area_chk, self._cam_motion_area_spin,
                      e.cam_motion_min_area, self._cam_motion_area_spin.setValue)
        # classes
        if e.cam_classes is None:
            self._cam_classes_chk.setChecked(True)
            self._cam_classes_edit.setEnabled(False)
        else:
            self._cam_classes_chk.setChecked(False)
            self._cam_classes_edit.setEnabled(True)
            self._cam_classes_edit.setText(", ".join(str(c) for c in e.cam_classes))

    def _browse_file(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Выбрать видеофайл", "",
            "Видео (*.mp4 *.avi *.mkv *.mov *.ts *.m4v);;Все файлы (*)"
        )
        if path:
            self._file_edit.setText(path)

    def _test_connection(self):
        url = self._url_edit.text().strip()
        if not url:
            QMessageBox.warning(self, "Ошибка", "Введите RTSP URL.")
            return
        self._test_btn.setText("Проверка…")
        self._test_btn.setEnabled(False)
        QApplication.processEvents()

        cap = cv2.VideoCapture(url, cv2.CAP_FFMPEG)
        cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
        ok = cap.isOpened()
        if ok:
            ok2, _ = cap.read()
            ok = ok and ok2
        cap.release()

        self._test_btn.setText("🔗  Проверить подключение")
        self._test_btn.setEnabled(True)
        if ok:
            QMessageBox.information(self, "Успех", "✔  Подключение успешно!")
        else:
            QMessageBox.critical(self, "Ошибка", "✘  Не удалось подключиться.")

    def _accept(self):
        name = self._name_edit.text().strip()
        if not name:
            QMessageBox.warning(self, "Ошибка", "Введите название камеры.")
            return
        mode = "file" if self._file_radio.isChecked() else "rtsp"
        src  = (self._file_edit.text().strip() if mode == "file"
                else self._url_edit.text().strip())
        if not src:
            QMessageBox.warning(self, "Ошибка", "Укажите источник (URL или файл).")
            return

        role = self._role_combo.currentData()
        self._entry.name          = name
        self._entry.cam_id        = self._camid_spin.value()
        self._entry.counter_id    = self._desk_spin.value() if role == "source" else 0
        self._entry.mode          = mode
        self._entry.rtsp_url      = self._url_edit.text().strip()
        self._entry.file_path     = self._file_edit.text().strip()
        self._entry.loop_video    = self._loop_chk.isChecked()
        self._entry.role          = role
        self._entry.receives_from = (self._from_combo.currentData()
                                     if role != "source" else None)

        # ── Per-camera overrides ──────────────────────────────────────────────
        self._entry.cam_confidence = (
            None if self._cam_conf_chk.isChecked()
            else self._cam_conf_spin.value()
        )
        self._entry.cam_iou = (
            None if self._cam_iou_chk.isChecked()
            else self._cam_iou_spin.value()
        )
        self._entry.cam_infer_every_n = (
            None if self._cam_infer_n_chk.isChecked()
            else self._cam_infer_n_spin.value()
        )
        self._entry.cam_infer_imgsz = (
            None if self._cam_imgsz_chk.isChecked()
            else self._cam_imgsz_combo.currentData()
        )
        self._entry.cam_reid_every_n = (
            None if self._cam_reid_n_chk.isChecked()
            else self._cam_reid_n_spin.value()
        )
        self._entry.cam_reid_min_crop_px = (
            None if self._cam_reid_min_px_chk.isChecked()
            else self._cam_reid_min_px_spin.value()
        )
        self._entry.cam_motion_detect = (
            None if self._cam_motion_chk.isChecked()
            else self._cam_motion_chk_val.isChecked()
        )
        self._entry.cam_motion_min_area = (
            None if self._cam_motion_area_chk.isChecked()
            else self._cam_motion_area_spin.value()
        )
        if self._cam_classes_chk.isChecked():
            self._entry.cam_classes = None
        else:
            raw_cls = self._cam_classes_edit.text().strip()
            parsed = []
            for part in raw_cls.split(","):
                part = part.strip()
                # Support "80: box" format — take just the ID
                if ":" in part:
                    part = part.split(":", 1)[0].strip()
                try:
                    parsed.append(int(part))
                except ValueError:
                    pass
            self._entry.cam_classes = parsed if parsed else []

        self.accept()

    def result_entry(self) -> CameraEntry:
        return self._entry


# ── Tracking overview tab ──────────────────────────────────────────────────────

class TrackTile(QFrame):
    """Small widget: thumbnail + camera name + track ID + status badge."""

    def __init__(self, cam_name: str, cam_id: int, tid: int,
                 cls_name: str, parent=None):
        super().__init__(parent)
        self.cam_name = cam_name
        self.cam_id   = cam_id
        self.tid      = tid
        self.cls_name = cls_name
        self._status  = "active"
        self.setFixedSize(110, 130)
        self.setFrameShape(QFrame.StyledPanel)
        self._build()

    def _build(self):
        v = QVBoxLayout(self)
        v.setContentsMargins(4, 4, 4, 4)
        v.setSpacing(2)

        self._img_lbl = QLabel()
        self._img_lbl.setFixedSize(100, 80)
        self._img_lbl.setAlignment(Qt.AlignCenter)
        self._img_lbl.setStyleSheet("background:#11111b; border-radius:4px;")
        v.addWidget(self._img_lbl)

        self._id_lbl = QLabel(f"#{self.tid}")
        self._id_lbl.setAlignment(Qt.AlignCenter)
        self._id_lbl.setStyleSheet("font-size:11px; font-weight:bold; color:#cdd6f4;")
        v.addWidget(self._id_lbl)

        self._status_lbl = QLabel("● active")
        self._status_lbl.setAlignment(Qt.AlignCenter)
        self._status_lbl.setStyleSheet("font-size:10px; color:#a6e3a1;")
        v.addWidget(self._status_lbl)

        self._update_style()

    def set_crop(self, crop: np.ndarray):
        if crop is None or crop.size == 0:
            return
        h, w = crop.shape[:2]
        rgb = cv2.cvtColor(crop, cv2.COLOR_BGR2RGB)
        qimg = QImage(rgb.data, w, h, w * 3, QImage.Format_RGB888)
        pix = QPixmap.fromImage(qimg).scaled(
            100, 80, Qt.KeepAspectRatio, Qt.SmoothTransformation
        )
        self._img_lbl.setPixmap(pix)

    def set_status(self, status: str, match_info: str = ""):
        self._status = status
        if status == "matched":
            self._status_lbl.setText(f"✔ {match_info}")
            self._status_lbl.setStyleSheet("font-size:10px; color:#a6e3a1;")
        elif status == "lost":
            self._status_lbl.setText("✘ lost")
            self._status_lbl.setStyleSheet("font-size:10px; color:#f38ba8;")
        else:
            self._status_lbl.setText("● active")
            self._status_lbl.setStyleSheet("font-size:10px; color:#89dceb;")
        self._update_style()

    def _update_style(self):
        colors = {"active": "#313244", "matched": "#1e3a2f", "lost": "#2a1a1a"}
        border = {"active": "#45475a", "matched": "#a6e3a1", "lost": "#f38ba8"}
        bg  = colors.get(self._status, "#313244")
        br  = border.get(self._status, "#45475a")
        self.setStyleSheet(
            f"TrackTile {{ background:{bg}; border:1px solid {br}; "
            f"border-radius:6px; }}"
        )


class CameraTrackPanel(QWidget):
    """Column showing all active tracks for one camera."""

    def __init__(self, cam_name: str, cam_id: int, role: str, parent=None):
        super().__init__(parent)
        self.cam_name = cam_name
        self.cam_id   = cam_id
        self.role     = role
        self._tiles: Dict[int, TrackTile] = {}
        self._build()

    def _build(self):
        v = QVBoxLayout(self)
        v.setContentsMargins(4, 4, 4, 4)
        v.setSpacing(4)

        role_colors = {"source": "#89b4fa", "query": "#a6e3a1", "transit": "#f9e2af"}
        role_names  = {"source": "SOURCE", "query": "QUERY", "transit": "TRANSIT"}
        clr = role_colors.get(self.role, "#cdd6f4")
        tag = role_names.get(self.role, self.role.upper())

        hdr = QLabel(f"<b>{self.cam_name}</b>  <span style='color:{clr};font-size:10px;'>{tag}</span>")
        hdr.setStyleSheet("background:#181825; padding:4px; border-radius:4px; color:#cdd6f4;")
        hdr.setWordWrap(True)
        v.addWidget(hdr)

        self._count_lbl = QLabel("Треков: 0")
        self._count_lbl.setStyleSheet("color:#6c7086; font-size:10px;")
        v.addWidget(self._count_lbl)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea{border:none; background:transparent;}")
        self._inner = QWidget()
        self._inner.setStyleSheet("background:transparent;")
        self._flow = QGridLayout(self._inner)
        self._flow.setSpacing(4)
        self._flow.setContentsMargins(0, 0, 0, 0)
        scroll.setWidget(self._inner)
        v.addWidget(scroll, 1)

    def update_track(self, tid: int, crop, cls_name: str, status: str):
        if tid not in self._tiles:
            tile = TrackTile(self.cam_name, self.cam_id, tid, cls_name)
            self._tiles[tid] = tile
            self._re_layout()
        tile = self._tiles[tid]
        if crop is not None:
            tile.set_crop(crop)
        tile.set_status(status)
        self._count_lbl.setText(f"Треков: {sum(1 for t in self._tiles.values() if t._status == 'active')}")

    def mark_matched(self, tid: int, match_info: str):
        if tid in self._tiles:
            self._tiles[tid].set_status("matched", match_info)

    def clear_old(self, max_tiles: int = 20):
        lost = [tid for tid, t in self._tiles.items() if t._status == "lost"]
        for tid in lost[:-max_tiles]:
            tile = self._tiles.pop(tid)
            tile.deleteLater()
        self._re_layout()

    def _re_layout(self):
        for i, tile in enumerate(self._tiles.values()):
            self._flow.removeWidget(tile)
            r, c = divmod(i, 3)
            self._flow.addWidget(tile, r, c)
        self._inner.setMinimumHeight(
            (len(self._tiles) // 3 + 1) * 138
        )


class TrackingTab(QWidget):
    """
    Real-time overview of all tracked objects across cameras.
    Shows track thumbnails grouped by camera role, plus recent match log.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self._panels: Dict[int, CameraTrackPanel] = {}   # cam_id → panel
        self._match_log_rows: list = []
        self._build_ui()
        # Auto-clean timer
        self._clean_timer = QTimer(self)
        self._clean_timer.timeout.connect(self._auto_clean)
        self._clean_timer.start(10_000)  # every 10s

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # Toolbar
        tb = QWidget()
        tb.setFixedHeight(38)
        tb.setStyleSheet("background:#181825; border-bottom:1px solid #313244;")
        tbl = QHBoxLayout(tb)
        tbl.setContentsMargins(8, 0, 8, 0)

        tbl.addWidget(QLabel("🎯  Треки в реальном времени"))

        tbl.addStretch()

        clr_btn = QPushButton("🗑  Очистить")
        clr_btn.setFixedHeight(26)
        clr_btn.clicked.connect(self.clear_all)
        tbl.addWidget(clr_btn)

        root.addWidget(tb)

        # Main splitter: top = cameras, bottom = match log
        splitter = QSplitter(Qt.Vertical)
        splitter.setStyleSheet("QSplitter::handle{background:#313244; height:3px;}")

        # Cameras area (horizontal scroll)
        self._cam_scroll = QScrollArea()
        self._cam_scroll.setWidgetResizable(True)
        self._cam_scroll.setStyleSheet("QScrollArea{border:none; background:#1e1e2e;}")
        self._cam_container = QWidget()
        self._cam_container.setStyleSheet("background:#1e1e2e;")
        self._cam_layout = QHBoxLayout(self._cam_container)
        self._cam_layout.setSpacing(8)
        self._cam_layout.setContentsMargins(8, 8, 8, 8)
        self._cam_placeholder = QLabel("Камеры не запущены. Нажмите ▶ Старт.")
        self._cam_placeholder.setStyleSheet("color:#45475a; font-size:14px;")
        self._cam_placeholder.setAlignment(Qt.AlignCenter)
        self._cam_layout.addWidget(self._cam_placeholder)
        self._cam_layout.addStretch()
        self._cam_scroll.setWidget(self._cam_container)
        splitter.addWidget(self._cam_scroll)

        # Match log
        log_w = QWidget()
        log_w.setStyleSheet("background:#181825;")
        lv = QVBoxLayout(log_w)
        lv.setContentsMargins(8, 4, 8, 4)
        lv.setSpacing(4)
        log_hdr = QLabel("Последние совпадения")
        log_hdr.setStyleSheet("color:#a6adc8; font-size:11px; font-weight:bold;")
        lv.addWidget(log_hdr)
        self._match_list = QListWidget()
        self._match_list.setStyleSheet(
            "QListWidget{background:#11111b; border:1px solid #313244; border-radius:4px; color:#cdd6f4; font-size:11px;}"
            "QListWidget::item{padding:3px 6px; border-bottom:1px solid #1e1e2e;}"
        )
        self._match_list.setMaximumHeight(160)
        lv.addWidget(self._match_list)
        splitter.addWidget(log_w)

        splitter.setSizes([400, 170])
        root.addWidget(splitter, 1)

    # ── Public API ────────────────────────────────────────────────────────────

    def rebuild(self, cameras, processors):
        """Connect to camera processors. Called when cameras change."""
        # Disconnect old
        for panel in self._panels.values():
            panel.deleteLater()
        self._panels.clear()
        while self._cam_layout.count() > 0:
            item = self._cam_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        self._cam_placeholder.setParent(None)
        if not cameras:
            self._cam_layout.addWidget(self._cam_placeholder)
            self._cam_layout.addStretch()
            return

        for cam, proc in zip(cameras, processors):
            panel = CameraTrackPanel(cam.name, cam.cam_id, cam.role)
            panel.setMinimumWidth(260)
            self._panels[cam.cam_id] = panel
            self._cam_layout.addWidget(panel)
            if proc:
                proc.track_updated.connect(
                    lambda tid, crop, cls, status, cid=cam.cam_id:
                        self._on_track_update(cid, tid, crop, cls, status)
                )
        self._cam_layout.addStretch()

    def on_match(self, mr):
        """Called when a ReID match is found — update tiles and log."""
        src_panel = self._panels.get(mr.source_entry.cam_id)
        # Find query panel by cam name
        qry_panel = None
        for cid, p in self._panels.items():
            if p.cam_name == mr.query_cam_name:
                qry_panel = p
                break

        info = f"{mr.similarity*100:.1f}%"
        if src_panel:
            src_panel.mark_matched(mr.source_entry.track_id, f"→{mr.query_cam_name} {info}")
        if qry_panel:
            qry_panel.mark_matched(mr.query_track_id, f"←{mr.source_entry.cam_name} {info}")

        ts = time.strftime("%H:%M:%S", time.localtime(mr.timestamp))
        verdict_icon = {"✔  Тот же багаж": "✔", "?  Вероятно тот же": "?", "✘  Другой": "✘"}.get(mr.verdict, "?")
        item = QListWidgetItem(
            f"  {verdict_icon}  {ts}  "
            f"{mr.source_entry.cam_name} #{mr.source_entry.track_id}"
            f"  →  {mr.query_cam_name} #{mr.query_track_id}"
            f"    {info}"
        )
        colors = {"✔": "#a6e3a1", "?": "#f9e2af", "✘": "#f38ba8"}
        item.setForeground(QColor(colors.get(verdict_icon, "#cdd6f4")))
        self._match_list.insertItem(0, item)
        while self._match_list.count() > 100:
            self._match_list.takeItem(self._match_list.count() - 1)

    def clear_all(self):
        for panel in self._panels.values():
            panel._tiles.clear()
            panel._re_layout()
        self._match_list.clear()

    # ── Private ───────────────────────────────────────────────────────────────

    def _on_track_update(self, cam_id: int, tid: int,
                         crop, cls_name: str, status: str):
        panel = self._panels.get(cam_id)
        if panel:
            panel.update_track(tid, crop, cls_name, status)

    def _auto_clean(self):
        for panel in self._panels.values():
            panel.clear_old(max_tiles=30)


# ── Monitor tab ────────────────────────────────────────────────────────────────

class MonitorTab(QWidget):
    """Responsive camera grid with column selector, compact mode, and drag-to-reorder."""

    cameras_reordered = pyqtSignal(list)   # emits new List[CameraEntry] order
    roi_changed       = pyqtSignal(object) # emits CameraEntry whose ROI changed

    def __init__(self, parent=None):
        super().__init__(parent)
        self._columns = 2
        self._compact = False
        self._widgets: List[CameraFeedWidget] = []
        self._processors: List[Optional[CameraProcessor]] = []
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # ── Monitor toolbar ───────────────────────────────────────────────────
        toolbar = QWidget()
        toolbar.setFixedHeight(38)
        toolbar.setStyleSheet(
            "background:#181825; border-bottom:1px solid #313244;"
        )
        tl = QHBoxLayout(toolbar)
        tl.setContentsMargins(8, 0, 8, 0)
        tl.setSpacing(4)

        tl.addWidget(QLabel("Колонки:"))

        self._col_btns: List[QPushButton] = []
        for n in (1, 2, 3, 4):
            btn = QPushButton(str(n))
            btn.setFixedSize(30, 26)
            btn.setCheckable(True)
            btn.setChecked(n == self._columns)
            btn.clicked.connect(lambda checked, cols=n: self._set_columns(cols))
            self._col_btns.append(btn)
            tl.addWidget(btn)

        tl.addWidget(_vline())

        self._compact_btn = QPushButton("▦ Компактно")
        self._compact_btn.setFixedHeight(26)
        self._compact_btn.setCheckable(True)
        self._compact_btn.setToolTip("Скрыть заголовки плиток")
        self._compact_btn.clicked.connect(self._toggle_compact)
        tl.addWidget(self._compact_btn)

        tl.addWidget(_vline())

        snap_all_btn = QPushButton("📷 Все снимки")
        snap_all_btn.setFixedHeight(26)
        snap_all_btn.setToolTip("Сохранить снимки со всех камер")
        snap_all_btn.clicked.connect(self._snapshot_all)
        tl.addWidget(snap_all_btn)

        tl.addStretch()

        self._cam_count_lbl = QLabel("Нет камер")
        self._cam_count_lbl.setStyleSheet("color:#6c7086; font-size:11px;")
        tl.addWidget(self._cam_count_lbl)

        root.addWidget(toolbar)

        # ── Scrollable grid area ──────────────────────────────────────────────
        self._scroll = QScrollArea()
        self._scroll.setWidgetResizable(True)
        self._scroll.setStyleSheet("QScrollArea{border:none;}")

        self._container = QWidget()
        self._container.setStyleSheet("background:#1e1e2e;")
        self._grid = QGridLayout(self._container)
        self._grid.setContentsMargins(6, 6, 6, 6)
        self._grid.setSpacing(8)
        self._scroll.setWidget(self._container)
        root.addWidget(self._scroll, 1)

        # Placeholder
        self._ph = QLabel("Камеры не добавлены.\nИспользуйте «＋ Добавить камеру».")
        self._ph.setAlignment(Qt.AlignCenter)
        self._ph.setStyleSheet("color:#45475a; font-size:16px;")
        self._grid.addWidget(self._ph, 0, 0)

    # ── Public API ────────────────────────────────────────────────────────────

    def rebuild(self, cameras: List[CameraEntry], processors: List[Optional[CameraProcessor]]):
        # Detach old widgets
        for w in self._widgets:
            self._grid.removeWidget(w)
            w.hide()
            w.deleteLater()
        self._widgets.clear()
        self._processors = processors

        self._ph.setVisible(len(cameras) == 0)

        for i, cam in enumerate(cameras):
            fw = CameraFeedWidget(cam)
            fw._cam_index = i
            fw.set_compact(self._compact)
            proc = processors[i] if i < len(processors) else None
            fw.roi_changed.connect(self._on_roi_changed)
            if proc:
                proc.frame_ready.connect(fw.on_frame)
                proc.crop_ready.connect(fw.on_crop)
                proc.stats_updated.connect(fw.on_stats)
                fw.parent_pause_cb = proc.set_paused
            row, col = divmod(i, self._columns)
            self._grid.addWidget(fw, row, col)
            self._widgets.append(fw)

        self._apply_stretch()
        n = len(cameras)
        self._cam_count_lbl.setText(f"Камер: {n}" if n else "Нет камер")

    def widget_for_index(self, idx: int) -> Optional[CameraFeedWidget]:
        return self._widgets[idx] if idx < len(self._widgets) else None

    def _on_roi_changed(self, cam: CameraEntry):
        self.roi_changed.emit(cam)

    def swap_cameras(self, i1: int, i2: int):
        """Swap two tiles by their cam_index and emit reorder signal."""
        if i1 == i2 or i1 >= len(self._widgets) or i2 >= len(self._widgets):
            return
        self._widgets[i1], self._widgets[i2] = self._widgets[i2], self._widgets[i1]
        # Update indices
        for i, w in enumerate(self._widgets):
            w._cam_index = i
        self._re_grid()
        self.cameras_reordered.emit([w.cam for w in self._widgets])

    def _re_grid(self):
        for w in self._widgets:
            self._grid.removeWidget(w)
        for i, w in enumerate(self._widgets):
            r, c = divmod(i, self._columns)
            self._grid.addWidget(w, r, c)
        self._apply_stretch()

    # ── Private ───────────────────────────────────────────────────────────────

    def _set_columns(self, n: int):
        self._columns = n
        for btn in self._col_btns:
            btn.setChecked(int(btn.text()) == n)
        self._re_grid()

    def _apply_stretch(self):
        for c in range(self._columns):
            self._grid.setColumnStretch(c, 1)
        # Clear stretch for columns beyond current count
        for c in range(self._columns, 5):
            self._grid.setColumnStretch(c, 0)
        rows = max(1, (len(self._widgets) + self._columns - 1) // self._columns)
        for r in range(rows):
            self._grid.setRowStretch(r, 1)

    def _toggle_compact(self):
        self._compact = not self._compact
        self._compact_btn.setChecked(self._compact)
        for w in self._widgets:
            w.set_compact(self._compact)

    def _snapshot_all(self):
        SNAPSHOT_DIR.mkdir(exist_ok=True)
        count = 0
        for fw in self._widgets:
            if fw._last_raw_frame is not None:
                ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
                path = SNAPSHOT_DIR / f"{fw.cam.name.replace(' ', '_')}_{ts}.jpg"
                cv2.imwrite(str(path), fw._last_raw_frame, [cv2.IMWRITE_JPEG_QUALITY, 95])
                count += 1
        QMessageBox.information(
            self, "Снимки",
            f"Сохранено {count} снимков в папку {SNAPSHOT_DIR}." if count
            else "Нет активных кадров для сохранения."
        )


def _vline() -> QFrame:
    """Thin vertical separator for toolbars."""
    f = QFrame()
    f.setFrameShape(QFrame.VLine)
    f.setFixedWidth(1)
    f.setStyleSheet("color:#45475a;")
    return f


# ── Cameras tab ────────────────────────────────────────────────────────────────

class CamerasTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._build_ui()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(8)

        lbl = QLabel("Управление камерами")
        lbl.setObjectName("hdr")
        root.addWidget(lbl)

        self._table = QTableWidget(0, 7)
        self._table.setHorizontalHeaderLabels(
            ["Название", "Режим", "Источник", "ID / Стол", "Роль", "Статус", "Активно"]
        )
        self._table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self._table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Stretch)
        self._table.setColumnWidth(4, 90)
        self._table.setSelectionBehavior(QTableWidget.SelectRows)
        self._table.setEditTriggers(QTableWidget.NoEditTriggers)
        self._table.verticalHeader().setVisible(False)
        self._table.setAlternatingRowColors(True)
        self._table.setStyleSheet(
            "QTableWidget { alternate-background-color: #181825; }"
        )
        root.addWidget(self._table, 1)

        # Button bar
        btn_row = QHBoxLayout()
        self._add_btn  = QPushButton("＋  Добавить камеру")
        self._add_btn.setObjectName("btn_primary")
        self._edit_btn = QPushButton("✏  Редактировать")
        self._del_btn  = QPushButton("✖  Удалить")
        self._del_btn.setObjectName("btn_danger")
        self._start_btn = QPushButton("▶  Запустить")
        self._start_btn.setObjectName("btn_success")
        self._stop_btn  = QPushButton("■  Остановить")
        self._stop_btn.setObjectName("btn_warn")
        self._test_btn  = QPushButton("🔗  Тест связи")

        self._scan_btn = QPushButton("🔍  Сканировать сеть")
        self._scan_btn.setToolTip("Найти камеры в локальной сети (ONVIF / RTSP)")

        for b in (self._add_btn, self._edit_btn, self._del_btn,
                  self._start_btn, self._stop_btn, self._test_btn):
            btn_row.addWidget(b)
        btn_row.addStretch()
        btn_row.addWidget(self._scan_btn)
        root.addLayout(btn_row)

    def populate(self, cameras: List[CameraEntry],
                 processors: List[Optional[CameraProcessor]]):
        self._table.setRowCount(len(cameras))
        for i, cam in enumerate(cameras):
            proc  = processors[i] if i < len(processors) else None
            stats = proc.stats if proc and proc.isRunning() else None
            status = stats.status if stats else "stopped"

            dot = StatusDot()
            dot.set_status(status)
            dot_wrap = QWidget()
            dot_layout = QHBoxLayout(dot_wrap)
            dot_layout.setContentsMargins(8, 0, 0, 0)
            dot_layout.addWidget(dot)
            dot_layout.addStretch()

            _role_labels = {"source": "Source", "query": "Query", "transit": "Transit"}
            _role_colors = {"source": "#89b4fa", "query": "#a6e3a1", "transit": "#cba6f7"}
            role_icon = _role_labels.get(cam.role, cam.role)
            role_item = QTableWidgetItem(role_icon)
            role_item.setForeground(QColor(_role_colors.get(cam.role, "#cdd6f4")))

            if cam.role == "source":
                id_cell = f"#{cam.cam_id} / стол {cam.counter_id}"
            elif cam.receives_from:
                id_cell = f"#{cam.cam_id} ← #{cam.receives_from}"
            else:
                id_cell = f"#{cam.cam_id}"
            self._table.setItem(i, 0, QTableWidgetItem(cam.name))
            self._table.setItem(i, 1, QTableWidgetItem("Файл" if cam.mode == "file" else "RTSP"))
            self._table.setItem(i, 2, QTableWidgetItem(cam.short_source()))
            self._table.setItem(i, 3, QTableWidgetItem(id_cell))
            self._table.setItem(i, 4, role_item)
            self._table.setCellWidget(i, 5, dot_wrap)
            self._table.setItem(i, 6, QTableWidgetItem(
                str(stats.active_count) if stats else "—"
            ))

    def selected_row(self) -> int:
        rows = self._table.selectedItems()
        return self._table.row(rows[0]) if rows else -1


# ── Settings tab ───────────────────────────────────────────────────────────────

class SettingsTab(QScrollArea):
    settings_saved = pyqtSignal()   # ← MainWindow слушает, чтобы обновить режим

    def __init__(self, cfg: AppConfig, parent=None):
        super().__init__(parent)
        self.setWidgetResizable(True)
        self._cfg = cfg
        container = QWidget()
        self.setWidget(container)
        self._build_ui(container)
        self._load()

    def _build_ui(self, container: QWidget):
        root = QVBoxLayout(container)
        root.setContentsMargins(16, 16, 16, 16)
        root.setSpacing(14)

        lbl = QLabel("Настройки")
        lbl.setObjectName("hdr")
        root.addWidget(lbl)

        # ── Model ─────────────────────────────────────────────────────────────
        mdl_grp = QGroupBox("Модель и трекер")
        mf = QFormLayout(mdl_grp)
        mf.setSpacing(10)

        # Быстрый выбор модели
        preset_row = QHBoxLayout()
        preset_row.setSpacing(4)
        _PRESETS = [
            ("11n", "yolo11n.pt"), ("11s ★", "yolo11s.pt"),
            ("11m",  "yolo11m.pt"), ("11l",  "yolo11l.pt"),
            ("11x",  "yolo11x.pt"), ("──", None),
            ("8n",  "yolov8n.pt"),  ("8s",  "yolov8s.pt"),
            ("8m",  "yolov8m.pt"),
        ]
        for label, name in _PRESETS:
            if name is None:
                sep = QLabel("│")
                sep.setStyleSheet("color:#45475a;")
                preset_row.addWidget(sep)
                continue
            btn = QPushButton(label)
            btn.setFixedSize(46, 24)
            btn.setStyleSheet(
                "font-size:11px; padding:0; border-radius:3px;"
                + ("background:#89b4fa;color:#1e1e2e;font-weight:bold;"
                   if "★" in label else "")
            )
            btn.setToolTip(name)
            btn.clicked.connect(lambda _=False, n=name: self._model_edit.setText(n))
            preset_row.addWidget(btn)
        preset_row.addStretch()
        mf.addRow("Быстрый выбор:", preset_row)

        mp_row = QHBoxLayout()
        self._model_edit = QLineEdit()
        self._model_edit.setPlaceholderText("yolo11s.pt  или  /path/to/custom.pt")
        browse_m = QPushButton("📂")
        browse_m.setFixedWidth(36)
        browse_m.setToolTip("Выбрать файл модели")
        browse_m.clicked.connect(self._browse_model)
        dl_btn = QPushButton("⬇ Скачать")
        dl_btn.setFixedWidth(90)
        dl_btn.setToolTip("Скачать / проверить выбранную модель через Ultralytics")
        dl_btn.clicked.connect(self._download_model)
        mp_row.addWidget(self._model_edit)
        mp_row.addWidget(browse_m)
        mp_row.addWidget(dl_btn)
        mf.addRow("Файл модели (.pt):", mp_row)

        self._tracker_combo = QComboBox()
        self._tracker_combo.addItems(["botsort.yaml", "bytetrack.yaml", "strongsort.yaml"])
        self._tracker_combo.setToolTip(
            "BoT-SORT: лучшая точность для сложных сцен\n"
            "ByteTrack: самый быстрый, хуже при окклюзиях\n"
            "StrongSORT: лучший при долгих окклюзиях (требует StrongSORT в ultralytics)"
        )
        mf.addRow("Трекер:", self._tracker_combo)

        self._conf_slider = QSlider(Qt.Horizontal)
        self._conf_slider.setRange(10, 95)
        self._conf_lbl = QLabel()
        self._conf_slider.valueChanged.connect(
            lambda v: self._conf_lbl.setText(f"{v/100:.2f}")
        )
        conf_row = QHBoxLayout()
        conf_row.addWidget(self._conf_slider)
        conf_row.addWidget(self._conf_lbl)
        mf.addRow("Порог уверенности:", conf_row)

        self._iou_slider = QSlider(Qt.Horizontal)
        self._iou_slider.setRange(10, 95)
        self._iou_lbl = QLabel()
        self._iou_slider.valueChanged.connect(
            lambda v: self._iou_lbl.setText(f"{v/100:.2f}")
        )
        iou_row = QHBoxLayout()
        iou_row.addWidget(self._iou_slider)
        iou_row.addWidget(self._iou_lbl)
        mf.addRow("IoU (NMS):", iou_row)

        # ── GPU / Device ──────────────────────────────────────────────────────
        self._device_combo = QComboBox()
        self._device_combo.addItem("🔄  Авто (cuda:0 если есть, иначе CPU)", "auto")
        self._device_combo.addItem("🖥  CPU", "cpu")
        try:
            import torch
            for i in range(torch.cuda.device_count()):
                props = torch.cuda.get_device_properties(i)
                total = props.total_memory / 1024 ** 3
                self._device_combo.addItem(
                    f"⚡  CUDA:{i}  —  {props.name}  ({total:.1f} GB)",
                    f"cuda:{i}",
                )
        except Exception:
            pass
        mf.addRow("Устройство (device):", self._device_combo)

        self._half_chk = QCheckBox(
            "FP16 / половинная точность  —  ~2× быстрее на GPU (RTX/A-series)"
        )
        self._half_chk.setToolTip(
            "Включать только при device = cuda:X.\n"
            "На CPU вызовет ошибку."
        )
        mf.addRow("", self._half_chk)

        # GPU info label (read-only)
        self._gpu_info_lbl = QLabel()
        self._gpu_info_lbl.setWordWrap(True)
        self._gpu_info_lbl.setStyleSheet("color:#a6adc8; font-size:11px;")
        self._refresh_gpu_label()
        mf.addRow("Доступные GPU:", self._gpu_info_lbl)

        root.addWidget(mdl_grp)

        # ── Classes — все 80 COCO классов ────────────────────────────────────
        cls_grp = QGroupBox("Обнаруживаемые классы YOLO (все 80 COCO)")
        cl_v = QVBoxLayout(cls_grp)
        cl_v.setSpacing(4)

        # Быстрые пресеты
        preset_row = QHBoxLayout()
        preset_row.setSpacing(6)
        _presets = [
            ("Багаж", [24, 25, 26, 28]),
            ("Люди", [0]),
            ("Всё", list(range(80))),
            ("Сбросить", []),
        ]
        for label, ids in _presets:
            btn = QPushButton(label)
            btn.setFixedHeight(22)
            btn.setStyleSheet("font-size:11px; padding:0 6px;")
            btn.clicked.connect(lambda _, _ids=ids: self._set_classes_preset(_ids))
            preset_row.addWidget(btn)
        preset_row.addStretch()
        cl_v.addLayout(preset_row)

        # Скроллируемая сетка всех классов
        cls_scroll = QScrollArea()
        cls_scroll.setWidgetResizable(True)
        cls_scroll.setFixedHeight(150)
        cls_scroll.setStyleSheet("QScrollArea{border:1px solid #45475a; border-radius:4px;}")
        cls_inner = QWidget()
        cls_inner.setStyleSheet("background:#181825;")
        cls_grid = QGridLayout(cls_inner)
        cls_grid.setSpacing(1)
        cls_grid.setContentsMargins(4, 4, 4, 4)
        cls_scroll.setWidget(cls_inner)

        self._class_chks: Dict[int, QCheckBox] = {}
        _COLS = 4
        for idx, (cid, cname) in enumerate(sorted(COCO_NAMES.items())):
            chk = QCheckBox(f"{cid}: {cname}")
            chk.setChecked(cid in (24, 26, 28))
            chk.setStyleSheet("font-size:10px; color:#a6adc8;")
            row, col = divmod(idx, _COLS)
            cls_grid.addWidget(chk, row, col)
            self._class_chks[cid] = chk

        cl_v.addWidget(cls_scroll)

        # Custom classes section
        custom_lbl = QLabel("Пользовательские классы (для своих моделей YOLO):")
        custom_lbl.setStyleSheet("color:#a6adc8; font-size:11px;")
        cl_v.addWidget(custom_lbl)
        self._custom_classes_edit = QLineEdit()
        self._custom_classes_edit.setPlaceholderText('Пример: "80: box, 81: cart, 82: pallet" — ID: название, через запятую')
        self._custom_classes_edit.setToolTip(
            "Классы за пределами COCO80 (ID 80+) для пользовательских YOLO-моделей.\n"
            "Формат: ID: название, ID: название\n"
            "Пример: 80: box, 81: cart"
        )
        cl_v.addWidget(self._custom_classes_edit)

        root.addWidget(cls_grp)

        # ── Dataset ───────────────────────────────────────────────────────────
        ds_grp = QGroupBox("Датасет")
        df = QFormLayout(ds_grp)
        df.setSpacing(10)

        def path_row(placeholder):
            row = QHBoxLayout()
            edit = QLineEdit()
            edit.setPlaceholderText(placeholder)
            btn = QPushButton("📂")
            btn.setFixedWidth(36)
            row.addWidget(edit)
            row.addWidget(btn)
            return row, edit, btn

        row1, self._yolo_img_edit, b1 = path_row("Dataset/datasetyolo/images")
        row2, self._yolo_lbl_edit, b2 = path_row("Dataset/datasetyolo/labels")
        row3, self._reid_edit,     b3 = path_row("Dataset/datasetReID")

        b1.clicked.connect(lambda: self._browse_dir(self._yolo_img_edit))
        b2.clicked.connect(lambda: self._browse_dir(self._yolo_lbl_edit))
        b3.clicked.connect(lambda: self._browse_dir(self._reid_edit))

        df.addRow("Изображения YOLO:", row1)
        df.addRow("Метки YOLO:",       row2)
        df.addRow("Папка ReID:",       row3)

        self._yolo_n_spin = QSpinBox()
        self._yolo_n_spin.setRange(1, 100)
        self._reid_k_spin = QSpinBox()
        self._reid_k_spin.setRange(1, 50)
        df.addRow("Сохранять кадр YOLO каждые N:", self._yolo_n_spin)
        df.addRow("Сохранять кроп ReID каждые K:", self._reid_k_spin)

        root.addWidget(ds_grp)

        # ── Performance ───────────────────────────────────────────────────────
        perf_grp = QGroupBox("Производительность / Очередь кадров")
        pff = QFormLayout(perf_grp)
        pff.setSpacing(8)

        self._infer_n_spin = QSpinBox()
        self._infer_n_spin.setRange(1, 10)
        self._infer_n_spin.setToolTip(
            "1 = YOLO на каждом кадре (максимум точности).\n"
            "2 = каждый второй, 3 = каждый третий и т.д.\n"
            "Уменьшает нагрузку на CPU/GPU при медленном железе."
        )
        pff.addRow("YOLO каждые N кадров:", self._infer_n_spin)

        self._disp_fps_spin = QSpinBox()
        self._disp_fps_spin.setRange(5, 60)
        self._disp_fps_spin.setSuffix(" fps")
        self._disp_fps_spin.setToolTip(
            "Максимальный FPS отправки кадров в интерфейс.\n"
            "Снижение разгружает Qt-очередь при многих камерах."
        )
        pff.addRow("Лимит FPS в интерфейсе:", self._disp_fps_spin)

        self._disp_w_spin = QSpinBox()
        self._disp_w_spin.setRange(0, 1920)
        self._disp_w_spin.setSingleStep(160)
        self._disp_w_spin.setSpecialValueText("Без ресайза")
        self._disp_w_spin.setToolTip(
            "0 = без масштабирования.\n"
            "640 = кадр уменьшается до 640px перед отправкой в GUI.\n"
            "Существенно снижает потребление памяти."
        )
        pff.addRow("Ширина GUI-кадра (px):", self._disp_w_spin)

        self._imgsz_combo = QComboBox()
        for sz in [256, 320, 416, 512, 640, 736, 800, 1280]:
            self._imgsz_combo.addItem(f"{sz} px", sz)
        self._imgsz_combo.setToolTip(
            "Размер входного изображения YOLO.\n"
            "Меньше = быстрее инференс, но хуже точность на мелких объектах.\n"
            "320 px: ~4× быстрее чем 640 px (CPU).\n"
            "416 px: ~2× быстрее — хороший баланс.\n"
            "640 px: максимальная точность (по умолчанию)."
        )
        pff.addRow("Размер YOLO-входа (imgsz):", self._imgsz_combo)

        self._shared_yolo_chk = QCheckBox(
            "Общая модель YOLO для всех камер  (рекомендуется при 8+ камерах)"
        )
        self._shared_yolo_chk.setToolTip(
            "Включите при 8+ камерах:\n"
            "Одна YOLO-модель загружается один раз и используется всеми потоками.\n"
            "Экономит VRAM (~100 MB на камеру), но инференс становится последовательным.\n"
            "Не включайте при 1-4 камерах — там параллельный инференс быстрее."
        )
        pff.addRow("", self._shared_yolo_chk)

        root.addWidget(perf_grp)

        # ── Анализ кадров ─────────────────────────────────────────────────────
        anal_grp = QGroupBox("Анализ кадров и трекинг")
        anf = QFormLayout(anal_grp)
        anf.setSpacing(8)

        self._reid_every_n_spin = QSpinBox()
        self._reid_every_n_spin.setRange(1, 30)
        self._reid_every_n_spin.setToolTip(
            "Запускать ReID каждые N детекций трека.\n"
            "1 = каждый раз (макс. точность, больше нагрузка).\n"
            "3–5 = хороший баланс скорость/точность.\n"
            "10+ = минимальная нагрузка (только первый кроп)."
        )
        anf.addRow("ReID каждые N детекций:", self._reid_every_n_spin)

        self._track_min_hits_spin = QSpinBox()
        self._track_min_hits_spin.setRange(1, 10)
        self._track_min_hits_spin.setToolTip(
            "Минимум последовательных детекций до подтверждения трека.\n"
            "2 — стандарт (убирает случайные срабатывания).\n"
            "1 — трек создаётся с первой детекции."
        )
        anf.addRow("Мин. детекций трека:", self._track_min_hits_spin)

        self._track_max_age_spin = QSpinBox()
        self._track_max_age_spin.setRange(1, 300)
        self._track_max_age_spin.setSuffix(" кадров")
        self._track_max_age_spin.setToolTip(
            "Через сколько кадров без детекции трек удаляется.\n"
            "30 — стандарт для 25 FPS (~1.2 сек).\n"
            "Увеличьте при частичных окклюзиях."
        )
        anf.addRow("Макс. возраст трека:", self._track_max_age_spin)

        self._motion_chk = QCheckBox(
            "Детектор движения  —  пропускать YOLO при статичной сцене"
        )
        self._motion_chk.setToolTip(
            "Экономит CPU/GPU при неподвижной камере.\n"
            "YOLO не запускается пока не обнаружено движение."
        )
        anf.addRow("", self._motion_chk)

        self._motion_area_spin = QSpinBox()
        self._motion_area_spin.setRange(100, 100000)
        self._motion_area_spin.setSingleStep(500)
        self._motion_area_spin.setSuffix(" px²")
        self._motion_area_spin.setToolTip(
            "Минимальная площадь движущегося контура для запуска YOLO.\n"
            "500–2000 — для крупных объектов (багаж, люди).\n"
            "100–500 — если нужно реагировать на мелкие объекты."
        )
        anf.addRow("Мин. площадь движения:", self._motion_area_spin)

        self._emb_cache_chk = QCheckBox(
            "Кешировать эмбеддинги трека  —  не пересчитывать уже совпавшие"
        )
        self._emb_cache_chk.setToolTip(
            "При включении: после первого успешного ReID-матча\n"
            "эмбеддинг данного трека не пересчитывается снова.\n"
            "Экономит ресурсы; отключите для мониторинга изменений."
        )
        anf.addRow("", self._emb_cache_chk)

        root.addWidget(anal_grp)

        # ── GROUP C: Кропы и ReID-сборщик ────────────────────────────────────
        crop_grp = QGroupBox("Кропы и ReID-сборщик")
        cpf2 = QFormLayout(crop_grp)
        cpf2.setSpacing(8)

        self._crop_pad_spin = QDoubleSpinBox()
        self._crop_pad_spin.setRange(0.0, 0.5)
        self._crop_pad_spin.setSingleStep(0.01)
        self._crop_pad_spin.setDecimals(2)
        self._crop_pad_spin.setToolTip(
            "Отступ вокруг bbox при вырезке кропа (доля от размера bbox).\n"
            "0.0 = строго по bbox.\n"
            "0.15 = +15% отступ с каждой стороны (рекомендуется для ReID).\n"
            "0.3+ = большой отступ, может включать фон."
        )
        cpf2.addRow("Отступ кропа (crop_pad):", self._crop_pad_spin)

        self._roi_crop_chk = QCheckBox("YOLO только в зоне ROI (roi_crop_infer)")
        self._roi_crop_chk.setToolTip(
            "Если включено: YOLO-инференс запускается только на вырезанной области ROI,\n"
            "а не на полном кадре. Значительно повышает FPS при малом ROI."
        )
        cpf2.addRow("", self._roi_crop_chk)

        self._reid_min_age_spin = QDoubleSpinBox()
        self._reid_min_age_spin.setRange(0.0, 30.0)
        self._reid_min_age_spin.setSingleStep(0.5)
        self._reid_min_age_spin.setDecimals(1)
        self._reid_min_age_spin.setSuffix(" сек")
        self._reid_min_age_spin.setSpecialValueText("Без ожидания")
        self._reid_min_age_spin.setToolTip(
            "Минимальное время существования трека перед первым ReID.\n"
            "Позволяет накопить несколько кропов прежде чем запускать идентификацию.\n"
            "0 = сразу при обнаружении.\n"
            "2–5 сек — рекомендуется для точной идентификации."
        )
        cpf2.addRow("Мин. возраст трека для ReID:", self._reid_min_age_spin)

        self._batch_crops_chk = QCheckBox("Пакетный ReID-инференс (batch crops)")
        self._batch_crops_chk.setToolTip(
            "Объединяет кропы от всех треков в один батч для OSNet.\n"
            "Быстрее на GPU при 4+ треках одновременно."
        )
        cpf2.addRow("", self._batch_crops_chk)

        root.addWidget(crop_grp)

        # ── Оверлей ───────────────────────────────────────────────────────────
        ovl_grp = QGroupBox("Оверлей на видео")
        ovf = QFormLayout(ovl_grp)
        ovf.setSpacing(8)

        self._ovl_bbox_chk = QCheckBox("Рисовать bounding box")
        self._ovl_bbox_chk.setToolTip("Прямоугольник вокруг обнаруженного объекта.")
        self._ovl_tid_chk  = QCheckBox("Показывать Track ID")
        self._ovl_tid_chk.setToolTip("Уникальный номер трека над объектом.")
        self._ovl_cls_chk  = QCheckBox("Показывать класс объекта")
        self._ovl_cls_chk.setToolTip("Название класса COCO (backpack, suitcase…)")
        self._ovl_conf_chk = QCheckBox("Показывать уверенность (conf)")
        self._ovl_conf_chk.setToolTip("Значение уверенности YOLO (0.00–1.00).")

        ovl_row1 = QHBoxLayout()
        ovl_row1.addWidget(self._ovl_bbox_chk)
        ovl_row1.addWidget(self._ovl_tid_chk)
        ovl_row2 = QHBoxLayout()
        ovl_row2.addWidget(self._ovl_cls_chk)
        ovl_row2.addWidget(self._ovl_conf_chk)
        ovf.addRow("", ovl_row1)
        ovf.addRow("", ovl_row2)
        root.addWidget(ovl_grp)

        # ── Снимки совпадений ─────────────────────────────────────────────────
        snap_grp = QGroupBox("Снимки совпадений ReID")
        snf = QFormLayout(snap_grp)
        snf.setSpacing(8)

        self._snap_chk = QCheckBox(
            "Автоматически сохранять снимок при обнаружении совпадения"
        )
        self._snap_chk.setToolTip(
            "При каждом ReID-совпадении (query → source) сохраняется\n"
            "коллаж из двух кропов с метаданными в папку снимков."
        )
        snf.addRow("", self._snap_chk)

        snap_path_row = QHBoxLayout()
        self._snap_dir_edit = QLineEdit()
        self._snap_dir_edit.setPlaceholderText("Snapshots")
        snap_browse = QPushButton("📂")
        snap_browse.setFixedWidth(36)
        snap_browse.clicked.connect(lambda: self._browse_dir(self._snap_dir_edit))
        snap_path_row.addWidget(self._snap_dir_edit)
        snap_path_row.addWidget(snap_browse)
        snf.addRow("Папка снимков:", snap_path_row)

        root.addWidget(snap_grp)

        # ── Connection ────────────────────────────────────────────────────────
        conn_grp = QGroupBox("Соединение и сеть")
        cf = QFormLayout(conn_grp)
        cf.setSpacing(8)
        self._reconnect_spin = QDoubleSpinBox()
        self._reconnect_spin.setRange(1, 60)
        self._reconnect_spin.setSuffix(" сек")
        self._reconnect_spin.setToolTip(
            "Задержка перед попыткой переподключения к RTSP-потоку при обрыве."
        )
        cf.addRow("Задержка переподключения:", self._reconnect_spin)

        self._stream_buf_spin = QSpinBox()
        self._stream_buf_spin.setRange(1, 10)
        self._stream_buf_spin.setSuffix(" кадр(а)")
        self._stream_buf_spin.setToolTip(
            "Размер внутреннего буфера OpenCV для RTSP-потока.\n"
            "1 = минимальная задержка (рекомендуется для онлайн-мониторинга).\n"
            "3–5 = сглаживание при нестабильной сети."
        )
        cf.addRow("Буфер RTSP-потока:", self._stream_buf_spin)

        self._web_port_spin = QSpinBox()
        self._web_port_spin.setRange(1024, 65535)
        self._web_port_spin.setToolTip(
            "Порт веб-дашборда (http://localhost:ПОРТ).\n"
            "По умолчанию: 8765.\n"
            "Требует перезапуска приложения для применения."
        )
        cf.addRow("Порт веб-дашборда:", self._web_port_spin)

        root.addWidget(conn_grp)

        # ── Режим работы ──────────────────────────────────────────────────────
        mode_grp = QGroupBox("Режим работы")
        mode_grp.setStyleSheet(
            "QGroupBox { border-color: #cba6f7; color: #cba6f7; }"
        )
        ml = QVBoxLayout(mode_grp)
        ml.setSpacing(8)

        self._mode_train_radio = QRadioButton(
            "🎓  Режим обучения  —  собирает датасет YOLO + ReID кропы на диск"
        )
        self._mode_prod_radio = QRadioButton(
            "🔍  Рабочий режим  —  ReID-сопоставление багажа в реальном времени"
        )
        self._mode_train_radio.setChecked(True)
        ml.addWidget(self._mode_train_radio)
        ml.addWidget(self._mode_prod_radio)

        # Training-only settings
        self._train_settings = QWidget()
        tf2 = QFormLayout(self._train_settings)
        tf2.setContentsMargins(16, 4, 0, 4)
        tf2.setSpacing(8)

        train_hint = QLabel(
            "Настройки сбора датасета для обучения ReID.\n"
            "Чемодан снимается с интервалом, затем новый чемодан только после cooldown."
        )
        train_hint.setStyleSheet("color:#6c7086; font-size:11px;")
        tf2.addRow("", train_hint)

        self._train_save_interval_spin = QDoubleSpinBox()
        self._train_save_interval_spin.setRange(0.05, 10.0)
        self._train_save_interval_spin.setSingleStep(0.1)
        self._train_save_interval_spin.setDecimals(2)
        self._train_save_interval_spin.setSuffix(" сек")
        self._train_save_interval_spin.setToolTip(
            "Минимальный интервал между сохранениями кропа одного чемодана.\n"
            "0.3 = раз в 0.3 сек (~3 кропа в секунду)."
        )
        tf2.addRow("Интервал сохранения кропа:", self._train_save_interval_spin)

        self._train_bag_cooldown_spin = QDoubleSpinBox()
        self._train_bag_cooldown_spin.setRange(0.1, 30.0)
        self._train_bag_cooldown_spin.setSingleStep(0.5)
        self._train_bag_cooldown_spin.setDecimals(1)
        self._train_bag_cooldown_spin.setSuffix(" сек")
        self._train_bag_cooldown_spin.setToolTip(
            "Пауза после исчезновения трека перед началом сбора нового.\n"
            "Нужно чтобы не смешивать разные чемоданы в одну папку."
        )
        tf2.addRow("Cooldown между чемоданами:", self._train_bag_cooldown_spin)

        self._train_link_timeout_spin = QDoubleSpinBox()
        self._train_link_timeout_spin.setRange(5.0, 300.0)
        self._train_link_timeout_spin.setSingleStep(5.0)
        self._train_link_timeout_spin.setDecimals(0)
        self._train_link_timeout_spin.setSuffix(" сек")
        self._train_link_timeout_spin.setToolTip(
            "Время ожидания связки source-трека с query-треком.\n"
            "Если query-трек появился в течение N сек после source — они считаются одним объектом.\n"
            "Увеличьте если путь от source до query камеры длинный."
        )
        tf2.addRow("Таймаут связки треков:", self._train_link_timeout_spin)

        ml.addWidget(self._train_settings)

        # Production-only settings (скрыты в режиме обучения)
        self._prod_settings = QWidget()
        pf = QFormLayout(self._prod_settings)
        pf.setContentsMargins(16, 4, 0, 4)
        pf.setSpacing(8)

        # Hint
        hint = QLabel(
            "Каждой камере задайте роль в диалоге «Редактировать»:\n"
            "  Source (📤) — стойка регистрации\n"
            "  Query  (📥) — лента сортировки"
        )
        hint.setStyleSheet("color:#6c7086; font-size:11px;")
        pf.addRow("", hint)

        ml.addWidget(self._prod_settings)
        root.addWidget(mode_grp)

        # Скрываем/показываем prod_settings
        def _update_mode_visibility(train_on):
            self._prod_settings.setVisible(not train_on)
            self._train_settings.setVisible(train_on)
        self._mode_train_radio.toggled.connect(_update_mode_visibility)
        self._prod_settings.setVisible(False)
        self._train_settings.setVisible(True)

        # ── Save / Reset ──────────────────────────────────────────────────────
        btn_row = QHBoxLayout()
        save_btn = QPushButton("💾  Сохранить настройки")
        save_btn.setObjectName("btn_primary")
        save_btn.clicked.connect(self._save)
        reset_btn = QPushButton("↺  Сбросить")
        reset_btn.clicked.connect(self._load)
        btn_row.addWidget(save_btn)
        btn_row.addWidget(reset_btn)
        btn_row.addStretch()
        root.addLayout(btn_row)

        root.addStretch()

    def _refresh_gpu_label(self):
        lines = gpu_info()
        if not lines:
            lines = ["GPU не обнаружены"]
        self._gpu_info_lbl.setText("\n".join(lines))

    def _set_classes_preset(self, ids: list):
        for cid, chk in self._class_chks.items():
            chk.setChecked(cid in ids)

    def _load(self):
        c = self._cfg
        self._model_edit.setText(c.model_path)
        idx = self._tracker_combo.findText(c.tracking_config)
        if idx >= 0:
            self._tracker_combo.setCurrentIndex(idx)
        self._conf_slider.setValue(int(c.confidence * 100))
        self._iou_slider.setValue(int(c.iou * 100))
        for cid, chk in self._class_chks.items():
            chk.setChecked(cid in c.classes)
        # Custom classes
        parts = [f"{k}: {v}" for k, v in sorted(c.custom_class_names.items())]
        self._custom_classes_edit.setText(", ".join(parts))
        self._yolo_img_edit.setText(c.yolo_images_dir)
        self._yolo_lbl_edit.setText(c.yolo_labels_dir)
        self._reid_edit.setText(c.reid_dir)
        self._yolo_n_spin.setValue(c.yolo_save_every_n)
        self._reid_k_spin.setValue(c.reid_save_every_k)
        self._infer_n_spin.setValue(c.infer_every_n)
        self._disp_fps_spin.setValue(c.display_fps_limit)
        self._disp_w_spin.setValue(c.display_width)
        idx = self._imgsz_combo.findData(c.infer_imgsz)
        if idx < 0:
            idx = self._imgsz_combo.findData(640)
        self._imgsz_combo.setCurrentIndex(max(0, idx))
        self._shared_yolo_chk.setChecked(c.shared_yolo)
        self._reconnect_spin.setValue(c.reconnect_delay)
        self._stream_buf_spin.setValue(c.stream_buffer)
        self._web_port_spin.setValue(c.web_port)
        dev_idx = self._device_combo.findData(c.yolo_device if c.yolo_device != "auto" else c.device)
        self._device_combo.setCurrentIndex(max(0, dev_idx))
        self._half_chk.setChecked(c.yolo_half)
        # Overlay
        self._ovl_bbox_chk.setChecked(c.overlay_bbox)
        self._ovl_tid_chk.setChecked(c.overlay_track_id)
        self._ovl_cls_chk.setChecked(c.overlay_class)
        self._ovl_conf_chk.setChecked(c.overlay_conf)
        # Snapshots
        self._snap_chk.setChecked(c.snapshot_on_match)
        self._snap_dir_edit.setText(c.snapshots_dir)
        # Анализ кадров
        self._reid_every_n_spin.setValue(c.reid_every_n)
        self._track_min_hits_spin.setValue(c.track_min_hits)
        self._track_max_age_spin.setValue(c.track_max_age)
        self._motion_chk.setChecked(c.motion_detect)
        self._motion_area_spin.setValue(c.motion_min_area)
        self._emb_cache_chk.setChecked(c.reid_embedding_cache)
        # Crop settings
        self._crop_pad_spin.setValue(getattr(c, 'crop_pad', 0.15))
        self._roi_crop_chk.setChecked(getattr(c, 'roi_crop_infer', True))
        self._reid_min_age_spin.setValue(getattr(c, 'reid_min_age_sec', 3.0))
        self._batch_crops_chk.setChecked(getattr(c, 'reid_batch_crops', True))
        # Training settings
        self._train_save_interval_spin.setValue(getattr(c, 'training_save_interval', 0.3))
        self._train_bag_cooldown_spin.setValue(getattr(c, 'training_bag_cooldown', 1.0))
        self._train_link_timeout_spin.setValue(getattr(c, 'training_link_timeout', 60.0))
        # Mode
        if c.app_mode == "production":
            self._mode_prod_radio.setChecked(True)
            self._prod_settings.setVisible(True)
            self._train_settings.setVisible(False)
        else:
            self._mode_train_radio.setChecked(True)
            self._prod_settings.setVisible(False)
            self._train_settings.setVisible(True)
        pass  # Reid TTL + threshold are in the ReID tab

    def _save(self):
        classes = sorted(cid for cid, chk in self._class_chks.items() if chk.isChecked())
        if not classes:
            QMessageBox.warning(self, "Ошибка", "Выберите хотя бы один класс.")
            return

        c = self._cfg
        c.model_path      = self._model_edit.text().strip()
        c.tracking_config = self._tracker_combo.currentText()
        c.confidence      = self._conf_slider.value() / 100
        c.iou             = self._iou_slider.value() / 100
        c.classes         = classes
        # Parse custom class names
        custom_text = self._custom_classes_edit.text().strip()
        custom_names = {}
        if custom_text:
            for part in custom_text.split(","):
                part = part.strip()
                if ":" in part:
                    id_str, name = part.split(":", 1)
                    try:
                        custom_names[int(id_str.strip())] = name.strip()
                    except ValueError:
                        pass
        c.custom_class_names = custom_names
        c.yolo_images_dir = self._yolo_img_edit.text().strip()
        c.yolo_labels_dir = self._yolo_lbl_edit.text().strip()
        c.reid_dir        = self._reid_edit.text().strip()
        c.yolo_save_every_n = self._yolo_n_spin.value()
        c.reid_save_every_k = self._reid_k_spin.value()
        c.infer_every_n     = self._infer_n_spin.value()
        c.display_fps_limit = self._disp_fps_spin.value()
        c.display_width     = self._disp_w_spin.value()
        c.infer_imgsz       = self._imgsz_combo.currentData()
        c.shared_yolo       = self._shared_yolo_chk.isChecked()
        c.reconnect_delay   = self._reconnect_spin.value()
        c.stream_buffer     = self._stream_buf_spin.value()
        c.web_port          = self._web_port_spin.value()
        c.yolo_device       = self._device_combo.currentData()
        c.device            = c.yolo_device   # sync legacy
        c.yolo_half         = self._half_chk.isChecked()
        c.half              = c.yolo_half      # sync legacy
        c.overlay_bbox      = self._ovl_bbox_chk.isChecked()
        c.overlay_track_id  = self._ovl_tid_chk.isChecked()
        c.overlay_class     = self._ovl_cls_chk.isChecked()
        c.overlay_conf      = self._ovl_conf_chk.isChecked()
        c.snapshot_on_match = self._snap_chk.isChecked()
        c.snapshots_dir     = self._snap_dir_edit.text().strip() or "Snapshots"
        c.reid_every_n         = self._reid_every_n_spin.value()
        c.track_min_hits       = self._track_min_hits_spin.value()
        c.track_max_age        = self._track_max_age_spin.value()
        c.motion_detect        = self._motion_chk.isChecked()
        c.motion_min_area      = self._motion_area_spin.value()
        c.reid_embedding_cache = self._emb_cache_chk.isChecked()
        c.crop_pad             = self._crop_pad_spin.value()
        c.roi_crop_infer       = self._roi_crop_chk.isChecked()
        c.reid_min_age_sec     = self._reid_min_age_spin.value()
        c.reid_batch_crops     = self._batch_crops_chk.isChecked()
        c.training_save_interval = self._train_save_interval_spin.value()
        c.training_bag_cooldown  = self._train_bag_cooldown_spin.value()
        c.training_link_timeout  = self._train_link_timeout_spin.value()
        c.app_mode          = "production" if self._mode_prod_radio.isChecked() else "training"

        # Предупреждение: FP16 на CPU — ошибка
        if c.yolo_half and c.yolo_device == "cpu":
            QMessageBox.warning(
                self, "Предупреждение",
                "FP16 (половинная точность) не работает на CPU.\n"
                "Отключите FP16 или выберите CUDA-устройство."
            )
            return

        QMessageBox.information(
            self, "Сохранено",
            "Настройки применены.\nПерезапустите камеры, чтобы изменения вступили в силу."
        )
        self.settings_saved.emit()

    def _browse_model(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Выбрать модель", "", "PyTorch модели (*.pt);;Все файлы (*)"
        )
        if path:
            self._model_edit.setText(path)

    def _download_model(self):
        name = self._model_edit.text().strip()
        if not name:
            QMessageBox.warning(self, "Ошибка", "Сначала введите имя модели.")
            return
        from PyQt5.QtWidgets import QProgressDialog
        pd = QProgressDialog(f"Загрузка / проверка  {name}…", "Отмена", 0, 0, self)
        pd.setWindowModality(Qt.WindowModal)
        pd.setMinimumDuration(0)
        pd.show()
        QApplication.processEvents()
        try:
            from ultralytics import YOLO
            YOLO(name)   # загружает если нет локально
            pd.close()
            QMessageBox.information(self, "Готово", f"✔  Модель {name!r} готова к использованию.")
        except Exception as e:
            pd.close()
            QMessageBox.critical(self, "Ошибка загрузки", str(e))

    def _browse_dir(self, edit: QLineEdit):
        d = QFileDialog.getExistingDirectory(self, "Выбрать папку", edit.text() or ".")
        if d:
            edit.setText(d)


# ── ReID tab ───────────────────────────────────────────────────────────────────

class ReIDTab(QScrollArea):
    """Dedicated tab for OSNet ReID configuration."""
    settings_saved = pyqtSignal()
    db_action = pyqtSignal(str, str)   # action, path

    def __init__(self, cfg: AppConfig, parent=None):
        super().__init__(parent)
        self.setWidgetResizable(True)
        self._cfg = cfg
        container = QWidget()
        self.setWidget(container)
        self._build_ui(container)
        self._load()

    def _build_ui(self, container: QWidget):
        root = QVBoxLayout(container)
        root.setContentsMargins(16, 16, 16, 16)
        root.setSpacing(14)

        lbl = QLabel("OSNet ReID — Настройки")
        lbl.setObjectName("hdr")
        root.addWidget(lbl)

        # ── Engine / Model ────────────────────────────────────────────────────
        model_grp = QGroupBox("Модель и движок — OSNet x1.0 (torchreid)")
        mf = QFormLayout(model_grp)
        mf.setSpacing(10)

        self._engine_combo = QComboBox()
        self._engine_combo.addItem("PyTorch  (CPU / GPU)  — работает везде", "pytorch")
        self._engine_combo.addItem("ONNX Runtime — GPU  (быстрее на NVIDIA, нужен onnxruntime-gpu)", "onnx_gpu")
        self._engine_combo.addItem("ONNX Runtime — CPU  (нужен onnxruntime)", "onnx_cpu")
        self._engine_combo.setToolTip(
            "Бэкенд для ReID-модели OSNet x1.0.\n\n"
            "PyTorch:  работает везде, использует device из вкладки «Настройки».\n"
            "ONNX GPU: быстрее на NVIDIA GTX/RTX; требует: pip install onnxruntime-gpu.\n"
            "          ONNX-файл (~6 МБ) создаётся автоматически при первом запуске.\n"
            "ONNX CPU: лучше чем PyTorch на чистом CPU; требует: pip install onnxruntime.\n\n"
            "Рекомендация: ONNX GPU при наличии NVIDIA, иначе PyTorch."
        )
        mf.addRow("Движок ReID:", self._engine_combo)

        # ReID model path
        reid_path_row = QHBoxLayout()
        self._reid_model_path_edit = QLineEdit()
        self._reid_model_path_edit.setPlaceholderText("По умолчанию: osnet_x1_0_256x128.onnx рядом со скриптом")
        self._reid_model_path_edit.setToolTip(
            "Путь к пользовательской ONNX-модели ReID.\n"
            "Оставьте пустым — программа ищет osnet_x1_0_256x128.onnx рядом со скриптом.\n"
            "Укажите путь если хотите использовать свою обученную модель с другим именем файла."
        )
        reid_path_browse = QPushButton("📂")
        reid_path_browse.setFixedWidth(36)
        reid_path_browse.clicked.connect(
            lambda: self._browse_reid_model()
        )
        reid_path_row.addWidget(self._reid_model_path_edit)
        reid_path_row.addWidget(reid_path_browse)
        mf.addRow("Путь к ReID модели (.onnx):", reid_path_row)

        self._status_lbl = QLabel("Загрузка при первом запуске камеры…")
        self._status_lbl.setStyleSheet("color:#a6adc8; font-size:11px;")
        self._status_lbl.setWordWrap(True)
        mf.addRow("Статус:", self._status_lbl)

        onnx_btn = QPushButton("⚙  Экспортировать ONNX-модель сейчас")
        onnx_btn.setToolTip(
            "Принудительно экспортирует OSNet → ONNX-файл (~6 МБ).\n"
            "Обычно происходит автоматически при первом запуске ONNX-движка.\n"
            "Нужен: pip install torchreid  (и  pip install onnx)"
        )
        onnx_btn.clicked.connect(self._export_onnx)
        mf.addRow("", onnx_btn)

        root.addWidget(model_grp)

        # ── Match thresholds ──────────────────────────────────────────────────
        thresh_grp = QGroupBox("Пороги сопоставления")
        tf = QFormLayout(thresh_grp)
        tf.setSpacing(10)

        def _slider_row(lo, hi):
            sl = QSlider(Qt.Horizontal)
            sl.setRange(lo, hi)
            lbl_val = QLabel()
            sl.valueChanged.connect(lambda v, l=lbl_val: l.setText(f"{v/100:.2f}"))
            row = QHBoxLayout()
            row.addWidget(sl)
            row.addWidget(lbl_val)
            return sl, row

        self._threshold_slider, th_row = _slider_row(40, 99)
        self._threshold_slider.setToolTip(
            "Минимальное косинусное сходство для регистрации совпадения.\n"
            "Ниже этого значения совпадение полностью игнорируется."
        )
        tf.addRow("Мин. сходство (порог матча):", th_row)

        self._verdict_high_slider, vh_row = _slider_row(60, 99)
        self._verdict_high_slider.setToolTip(
            "Порог уверенного совпадения — зелёная карточка «✔ Тот же багаж»."
        )
        tf.addRow("Порог  ✔ «Тот же багаж»:", vh_row)

        self._verdict_mid_slider, vm_row = _slider_row(40, 90)
        self._verdict_mid_slider.setToolTip(
            "Порог возможного совпадения — жёлтая карточка «? Вероятно тот же».\n"
            "Ниже этого порога (но выше мин. сходства) — красная карточка."
        )
        tf.addRow("Порог  ? «Вероятно тот же»:", vm_row)

        hint_thresh = QLabel(
            "Значения по умолчанию оптимальны для OSNet x1.0.\n"
            "Увеличивайте пороги, если много ложных совпадений;\n"
            "уменьшайте, если реальные совпадения пропускаются."
        )
        hint_thresh.setStyleSheet("color:#6c7086; font-size:11px;")
        tf.addRow("", hint_thresh)

        root.addWidget(thresh_grp)

        # ── Кроп-фильтрация ───────────────────────────────────────────────────
        crop_grp = QGroupBox("Фильтрация кропов")
        cpf = QFormLayout(crop_grp)
        cpf.setSpacing(10)

        self._min_crop_spin = QSpinBox()
        self._min_crop_spin.setRange(8, 256)
        self._min_crop_spin.setSuffix(" px")
        self._min_crop_spin.setToolTip(
            "Минимальный размер меньшей стороны кропа.\n"
            "Кропы меньше этого значения не отправляются в OSNet.\n"
            "Маленькие кропы содержат мало деталей и ухудшают точность.\n"
            "Рекомендуется: 32–64 px."
        )
        cpf.addRow("Мин. размер кропа:", self._min_crop_spin)

        root.addWidget(crop_grp)

        # ── Стратегия матчинга ────────────────────────────────────────────────
        match_grp = QGroupBox("Стратегия матчинга")
        maf = QFormLayout(match_grp)
        maf.setSpacing(10)

        self._aggregation_combo = QComboBox()
        self._aggregation_combo.addItem("max  — лучшее одиночное совпадение", "max")
        self._aggregation_combo.addItem("mean — среднее по всем наблюдениям", "mean")
        self._aggregation_combo.setToolTip(
            "Как объединять несколько векторов одного трека при матчинге.\n\n"
            "max:  берётся совпадение с максимальным сходством.\n"
            "      Быстрее, устойчивее при движении камеры.\n\n"
            "mean: усредняет все эмбеддинги трека перед сравнением.\n"
            "      Точнее при стабильной съёмке, медленнее при большой БД."
        )
        maf.addRow("Агрегация эмбеддингов:", self._aggregation_combo)

        self._top_k_spin = QSpinBox()
        self._top_k_spin.setRange(1, 10)
        self._top_k_spin.setToolTip(
            "Сколько лучших совпадений возвращать из БД.\n"
            "1 = только самое похожее (обычный режим).\n"
            "2–5 = показывать несколько кандидатов (режим аналитики)."
        )
        maf.addRow("Top-K совпадений:", self._top_k_spin)

        root.addWidget(match_grp)

        # ── ReID Database ─────────────────────────────────────────────────────
        db_grp = QGroupBox("База данных ReID")
        df = QFormLayout(db_grp)
        df.setSpacing(10)

        self._ttl_spin = QDoubleSpinBox()
        self._ttl_spin.setRange(1, 60)
        self._ttl_spin.setSuffix(" мин")
        self._ttl_spin.setToolTip(
            "Через сколько минут запись исчезает из временной БД ReID.\n"
            "Увеличьте, если путь багажа между точками контроля длинный."
        )
        df.addRow("Время жизни в БД:", self._ttl_spin)

        self._max_db_spin = QSpinBox()
        self._max_db_spin.setRange(0, 10000)
        self._max_db_spin.setSpecialValueText("Без ограничений")
        self._max_db_spin.setSingleStep(100)
        self._max_db_spin.setToolTip(
            "Максимальное число записей в БД.\n"
            "0 = без ограничений (очистка только по TTL).\n"
            "При достижении лимита самые старые записи удаляются."
        )
        df.addRow("Макс. записей в БД:", self._max_db_spin)

        cam_hint = QLabel(
            "Каждой камере задайте роль в диалоге «Редактировать»:\n"
            "  Source (📤) — стойка регистрации  (отправитель багажа)\n"
            "  Query  (📥) — лента сортировки     (получатель)"
        )
        cam_hint.setStyleSheet("color:#6c7086; font-size:11px;")
        df.addRow("", cam_hint)

        # Persistent DB
        db_file_row = QHBoxLayout()
        self._db_path_edit = QLineEdit()
        self._db_path_edit.setPlaceholderText("reid_db.json  — путь к сохранённой БД")
        db_browse = QPushButton("📂")
        db_browse.setFixedWidth(36)
        db_browse.clicked.connect(self._browse_db_path)
        db_file_row.addWidget(self._db_path_edit)
        db_file_row.addWidget(db_browse)
        df.addRow("Файл БД (JSON):", db_file_row)

        self._autosave_spin = QSpinBox()
        self._autosave_spin.setRange(0, 300)
        self._autosave_spin.setSuffix(" сек")
        self._autosave_spin.setSpecialValueText("Отключено")
        self._autosave_spin.setValue(0)
        self._autosave_spin.setToolTip("Автосохранение БД ReID каждые N секунд. 0 = отключено.")
        df.addRow("Автосохранение:", self._autosave_spin)

        db_btn_row = QHBoxLayout()
        self._save_db_btn = QPushButton("💾  Сохранить БД сейчас")
        self._save_db_btn.clicked.connect(self._save_db_now)
        self._load_db_btn = QPushButton("📂  Загрузить БД")
        self._load_db_btn.clicked.connect(self._load_db_now)
        self._clear_db_btn = QPushButton("🗑  Очистить БД")
        self._clear_db_btn.setObjectName("btn_danger")
        self._clear_db_btn.clicked.connect(self._clear_db)
        db_btn_row.addWidget(self._save_db_btn)
        db_btn_row.addWidget(self._load_db_btn)
        db_btn_row.addWidget(self._clear_db_btn)
        db_btn_row.addStretch()
        df.addRow("", db_btn_row)

        self._db_size_lbl = QLabel("—")
        self._db_size_lbl.setStyleSheet("color:#a6adc8; font-size:11px;")
        df.addRow("Записей в БД:", self._db_size_lbl)

        root.addWidget(db_grp)

        # ── GROUP A: Голосование N×M ──────────────────────────────────────────
        vote_grp = QGroupBox("Голосование N×M (voting)")
        vf = QFormLayout(vote_grp)
        vf.setSpacing(8)

        self._voting_chk = QCheckBox("Использовать голосование N×M эмбеддингов")
        self._voting_chk.setToolTip(
            "Вместо сравнения одного вектора — сравниваются все N накопленных "
            "эмбеддингов query-трека против всех K эмбеддингов gallery.\n"
            "Голос засчитывается если пара > vote_threshold.\n"
            "score = 0.5*max_sim + 0.3*mean_sim + 0.2*vote_ratio\n"
            "Значительно точнее но немного медленнее."
        )
        vf.addRow("", self._voting_chk)

        self._vote_threshold_slider = QSlider(Qt.Horizontal)
        self._vote_threshold_slider.setRange(40, 99)
        _vt_lbl = QLabel()
        self._vote_threshold_slider.valueChanged.connect(
            lambda v, l=_vt_lbl: l.setText(f"{v/100:.2f}")
        )
        _vt_lbl.setText(f"{self._vote_threshold_slider.value()/100:.2f}")
        vt_row = QHBoxLayout()
        vt_row.addWidget(self._vote_threshold_slider)
        vt_row.addWidget(_vt_lbl)
        self._vote_threshold_slider.setToolTip("Минимальное сходство пары эмбеддингов для засчитывания голоса.")
        vf.addRow("Порог голоса (vote_threshold):", vt_row)

        self._min_votes_spin = QSpinBox()
        self._min_votes_spin.setRange(1, 20)
        self._min_votes_spin.setToolTip("Минимум засчитанных голосов для регистрации совпадения.\n3–5 — хороший баланс.")
        vf.addRow("Мин. голосов:", self._min_votes_spin)

        self._vote_every_n_spin = QSpinBox()
        self._vote_every_n_spin.setRange(1, 30)
        self._vote_every_n_spin.setToolTip(
            "Запускать голосование каждые N детекций трека.\n"
            "1 = каждый раз (медленнее), 5 = каждые 5 детекций (быстрее)."
        )
        vf.addRow("Голосование каждые N детекций:", self._vote_every_n_spin)

        root.addWidget(vote_grp)

        # ── GROUP B: Дополнительные алгоритмы ────────────────────────────────
        adv_grp = QGroupBox("Дополнительные алгоритмы")
        af = QFormLayout(adv_grp)
        af.setSpacing(8)

        self._color_prefilter_slider = QSlider(Qt.Horizontal)
        self._color_prefilter_slider.setRange(0, 50)
        _cp_lbl = QLabel()
        self._color_prefilter_slider.valueChanged.connect(
            lambda v, l=_cp_lbl: l.setText(f"{v/100:.2f}")
        )
        _cp_lbl.setText(f"{self._color_prefilter_slider.value()/100:.2f}")
        cp_row = QHBoxLayout()
        cp_row.addWidget(self._color_prefilter_slider)
        cp_row.addWidget(_cp_lbl)
        self._color_prefilter_slider.setToolTip(
            "Минимальное пересечение HSV-гистограмм для допуска к OSNet.\n"
            "0 = отключено (всегда запускать OSNet).\n"
            "0.05–0.10 = отфильтровывает объекты с разными цветами (рекомендуется).\n"
            "0.20+ = жёсткий цветовой фильтр (возможны пропуски)."
        )
        af.addRow("Цветовой пре-фильтр (color_sim_min):", cp_row)

        self._adaptive_thresh_chk = QCheckBox("Адаптивный порог (повышается при однородной БД)")
        self._adaptive_thresh_chk.setToolTip(
            "Если все записи в БД похожи (малое разнообразие),\n"
            "порог автоматически повышается чтобы снизить ложные совпадения.\n"
            "Работает только в production-режиме."
        )
        af.addRow("", self._adaptive_thresh_chk)

        self._adaptive_boost_spin = QDoubleSpinBox()
        self._adaptive_boost_spin.setRange(0.0, 0.2)
        self._adaptive_boost_spin.setSingleStep(0.01)
        self._adaptive_boost_spin.setDecimals(3)
        self._adaptive_boost_spin.setToolTip("Максимальное повышение порога при низком разнообразии БД.")
        af.addRow("Макс. повышение порога:", self._adaptive_boost_spin)

        self._transit_spin = QDoubleSpinBox()
        self._transit_spin.setRange(0.0, 600.0)
        self._transit_spin.setSingleStep(1.0)
        self._transit_spin.setDecimals(1)
        self._transit_spin.setSuffix(" сек")
        self._transit_spin.setSpecialValueText("Отключено")
        self._transit_spin.setToolTip(
            "Ожидаемое время прохождения между source и query камерами.\n"
            "0 = отключено.\n"
            "Если задать, совпадению даётся бонус +0.02 если время перехода\n"
            "близко к указанному значению (±20%).\n"
            "Пример: 30 сек — если конвейер проходит за ~30 секунд."
        )
        af.addRow("Ожидаемое время перехода:", self._transit_spin)

        root.addWidget(adv_grp)

        # ── Save / Reset ──────────────────────────────────────────────────────
        btn_row = QHBoxLayout()
        save_btn = QPushButton("💾  Сохранить настройки ReID")
        save_btn.setObjectName("btn_primary")
        save_btn.clicked.connect(self._save)
        reset_btn = QPushButton("↺  Сбросить")
        reset_btn.clicked.connect(self._load)
        btn_row.addWidget(save_btn)
        btn_row.addWidget(reset_btn)
        btn_row.addStretch()
        root.addLayout(btn_row)

        root.addStretch()

    # ── Load / Save ────────────────────────────────────────────────────────────

    def _load(self):
        c = self._cfg
        idx = self._engine_combo.findData(c.reid_engine)
        self._engine_combo.setCurrentIndex(max(0, idx))
        self._reid_model_path_edit.setText(c.reid_model_path)
        self._threshold_slider.setValue(int(c.reid_threshold * 100))
        self._verdict_high_slider.setValue(int(c.reid_verdict_high * 100))
        self._verdict_mid_slider.setValue(int(c.reid_verdict_mid * 100))
        self._ttl_spin.setValue(c.reid_ttl_minutes)
        self._min_crop_spin.setValue(c.reid_min_crop_px)
        idx2 = self._aggregation_combo.findData(c.reid_aggregation)
        self._aggregation_combo.setCurrentIndex(max(0, idx2))
        self._top_k_spin.setValue(c.reid_top_k)
        self._max_db_spin.setValue(c.reid_max_db_size)
        p = getattr(c, 'reid_db_path', '')
        self._db_path_edit.setText(p)
        autosave = getattr(c, 'reid_autosave_interval', 0)
        self._autosave_spin.setValue(autosave)
        # Voting
        use_voting = getattr(c, 'reid_min_votes', 3) > 0
        self._voting_chk.setChecked(use_voting)
        self._vote_threshold_slider.setValue(int(getattr(c, 'reid_vote_threshold', 0.65) * 100))
        self._min_votes_spin.setValue(getattr(c, 'reid_min_votes', 3))
        self._vote_every_n_spin.setValue(getattr(c, 'reid_vote_every_n', 5))
        # Advanced
        self._color_prefilter_slider.setValue(int(getattr(c, 'color_prefilter_threshold', 0.08) * 100))
        self._adaptive_thresh_chk.setChecked(getattr(c, 'reid_adaptive_threshold', True))
        self._adaptive_boost_spin.setValue(getattr(c, 'reid_adaptive_max_boost', 0.08))
        self._transit_spin.setValue(getattr(c, 'transit_time_hint', 0.0))

    def _save(self):
        c = self._cfg
        c.reid_engine       = self._engine_combo.currentData()
        c.reid_model_path   = self._reid_model_path_edit.text().strip()
        c.reid_threshold    = self._threshold_slider.value() / 100
        c.reid_verdict_high = self._verdict_high_slider.value() / 100
        c.reid_verdict_mid  = self._verdict_mid_slider.value() / 100
        c.reid_ttl_minutes  = self._ttl_spin.value()
        c.reid_min_crop_px  = self._min_crop_spin.value()
        c.reid_aggregation  = self._aggregation_combo.currentData()
        c.reid_top_k        = self._top_k_spin.value()
        c.reid_max_db_size  = self._max_db_spin.value()
        c.reid_db_path = self._db_path_edit.text().strip()
        c.reid_autosave_interval = self._autosave_spin.value()
        c.reid_vote_threshold    = self._vote_threshold_slider.value() / 100
        c.reid_min_votes         = self._min_votes_spin.value() if self._voting_chk.isChecked() else 0
        c.reid_vote_every_n      = self._vote_every_n_spin.value()
        c.color_prefilter_threshold = self._color_prefilter_slider.value() / 100
        c.reid_adaptive_threshold   = self._adaptive_thresh_chk.isChecked()
        c.reid_adaptive_max_boost   = self._adaptive_boost_spin.value()
        c.transit_time_hint         = self._transit_spin.value()
        QMessageBox.information(
            self, "Сохранено",
            "Настройки ReID применены.\n"
            "Перезапустите камеры, чтобы изменения вступили в силу."
        )
        self.settings_saved.emit()

    # ── ONNX export helper ─────────────────────────────────────────────────────

    def _export_onnx(self):
        onnx_path = app_dir() / ReIDFeatureExtractor._ONNX_NAME
        if onnx_path.exists():
            QMessageBox.information(
                self, "ONNX уже существует",
                f"Файл уже есть:\n{onnx_path}\n\n"
                "Удалите его вручную, чтобы экспортировать заново."
            )
            return

        from PyQt5.QtWidgets import QProgressDialog
        pd_dlg = QProgressDialog("Экспорт OSNet → ONNX…\n(может занять 30–60 сек)", None, 0, 0, self)
        pd_dlg.setWindowModality(Qt.WindowModal)
        pd_dlg.setMinimumDuration(0)
        pd_dlg.show()
        QApplication.processEvents()
        try:
            import torch
            m = ReIDFeatureExtractor._build_osnet()
            m.eval()
            dummy = torch.randn(1, 3, ReIDFeatureExtractor.INPUT_H,
                                ReIDFeatureExtractor.INPUT_W)
            torch.onnx.export(
                m, dummy, str(onnx_path),
                export_params=True,
                opset_version=12,
                do_constant_folding=True,
                input_names=["input"],
                output_names=["output"],
                dynamic_axes={"input": {0: "batch_size"}, "output": {0: "batch_size"}},
            )
            pd_dlg.close()
            QMessageBox.information(
                self, "Готово",
                f"✔  ONNX-модель сохранена:\n{onnx_path}"
            )
        except Exception as exc:
            pd_dlg.close()
            QMessageBox.critical(self, "Ошибка экспорта", str(exc))

    def _browse_reid_model(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Выбрать ReID модель", "",
            "ONNX модели (*.onnx);;Все файлы (*)"
        )
        if path:
            self._reid_model_path_edit.setText(path)

    def update_status(self, text: str):
        """Called by MainWindow after ReID extractor loads."""
        self._status_lbl.setText(text)

    # ── Persistent DB helpers ──────────────────────────────────────────────────

    def _browse_db_path(self):
        p, _ = QFileDialog.getSaveFileName(self, "Файл БД ReID", str(app_dir() / "reid_db.json"), "JSON (*.json)")
        if p:
            self._db_path_edit.setText(p)

    def _save_db_now(self):
        p = self._db_path_edit.text().strip() or str(app_dir() / "reid_db.json")
        self.db_action.emit("save", p)

    def _load_db_now(self):
        p, _ = QFileDialog.getOpenFileName(self, "Загрузить БД ReID", str(app_dir()), "JSON (*.json)")
        if p:
            self._db_path_edit.setText(p)
            self.db_action.emit("load", p)

    def _clear_db(self):
        if QMessageBox.question(self, "Очистить БД", "Удалить все записи из БД ReID?",
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            self.db_action.emit("clear", "")

    def update_db_size(self, n: int):
        self._db_size_lbl.setText(str(n))

    def get_autosave_interval(self) -> int:
        return self._autosave_spin.value()


# ── Devices tab ────────────────────────────────────────────────────────────────

class DevicesTab(QScrollArea):
    """Per-component device assignment: YOLO vs ReID vs Training."""
    settings_saved = pyqtSignal()

    def __init__(self, cfg: "AppConfig", parent=None):
        super().__init__(parent)
        self.setWidgetResizable(True)
        self._cfg = cfg
        container = QWidget()
        self.setWidget(container)
        self._build_ui(container)
        self._load()

    # ── helpers ────────────────────────────────────────────────────────────────

    def _device_combo(self, include_train_fmt=False) -> QComboBox:
        cb = QComboBox()
        cb.addItem("🔄  Авто (cuda:0 если есть, иначе CPU)", "auto")
        cb.addItem("🖥  CPU", "cpu")
        try:
            import torch
            for i in range(torch.cuda.device_count()):
                p = torch.cuda.get_device_properties(i)
                mem = p.total_memory / 1024 ** 3
                cb.addItem(f"⚡  CUDA:{i}  {p.name}  ({mem:.1f} GB)", f"cuda:{i}")
                if include_train_fmt:
                    cb.addItem(f"⚡  GPU:{i}  (YOLO формат)", str(i))
        except Exception:
            pass
        return cb

    # ── UI ─────────────────────────────────────────────────────────────────────

    def _build_ui(self, container: QWidget):
        root = QVBoxLayout(container)
        root.setContentsMargins(16, 16, 16, 16)
        root.setSpacing(14)

        hdr = QLabel("Устройства и бэкенды вычислений")
        hdr.setObjectName("hdr")
        root.addWidget(hdr)

        # ── YOLO ──────────────────────────────────────────────────────────────
        yolo_grp = QGroupBox("YOLO — Обнаружение объектов")
        yf = QFormLayout(yolo_grp)
        yf.setSpacing(10)

        self._yolo_dev = self._device_combo()
        self._yolo_dev.setToolTip(
            "Устройство для инференса YOLO.\n"
            "Авто: выбирает cuda:0 если доступен, иначе CPU.\n"
            "Несколько камер на одном GPU — нормально; используйте BatchEngine."
        )
        yf.addRow("Устройство YOLO:", self._yolo_dev)

        self._yolo_half = QCheckBox("FP16 (половинная точность) — ~2× быстрее на RTX/A-series")
        self._yolo_half.setToolTip("Только при device=cuda:X. На CPU вызывает ошибку.")
        yf.addRow("", self._yolo_half)

        self._imgsz_combo = QComboBox()
        for sz, hint in [(256,"быстро/грубо"), (320,""), (416,"баланс"), (512,""),
                         (640,"по умолч."), (736,""), (800,""), (1280,"точно/медленно")]:
            label = f"{sz} px" + (f"  — {hint}" if hint else "")
            self._imgsz_combo.addItem(label, sz)
        self._imgsz_combo.setToolTip(
            "Размер входа YOLO-модели.\n"
            "320 px ≈ 4× быстрее 640 px, но хуже на мелких объектах.\n"
            "416–512 px — хороший компромисс для багажа."
        )
        yf.addRow("Размер входа (imgsz):", self._imgsz_combo)

        self._tracker_combo = QComboBox()
        self._tracker_combo.addItems(["botsort.yaml", "bytetrack.yaml"])
        self._tracker_combo.setToolTip(
            "BoT-SORT: лучше при окклюзиях и пересечениях (рекомендуется).\n"
            "ByteTrack: быстрее, подходит для высокого FPS без окклюзий."
        )
        yf.addRow("Трекер:", self._tracker_combo)

        root.addWidget(yolo_grp)

        # ── ReID / OSNet ───────────────────────────────────────────────────────
        reid_grp = QGroupBox("OSNet ReID — Re-Identification")
        rf = QFormLayout(reid_grp)
        rf.setSpacing(10)

        self._reid_dev = self._device_combo()
        self._reid_dev.setToolTip(
            "Устройство для OSNet ReID — независимо от YOLO!\n"
            "Можно держать YOLO на GPU, а ReID на CPU (или наоборот).\n"
            "ReID обрабатывает одиночные кропы — нагрузка небольшая."
        )
        rf.addRow("Устройство ReID:", self._reid_dev)

        self._reid_engine_combo = QComboBox()
        self._reid_engine_combo.addItem("PyTorch  (универсальный, CPU/GPU)", "pytorch")
        self._reid_engine_combo.addItem("ONNX Runtime — GPU  (быстрее на NVIDIA, нужен onnxruntime-gpu)", "onnx_gpu")
        self._reid_engine_combo.addItem("ONNX Runtime — CPU  (нужен onnxruntime)", "onnx_cpu")
        self._reid_engine_combo.setToolTip(
            "Бэкенд ReID-модели:\n"
            "PyTorch:  работает везде, использует выбранный device.\n"
            "ONNX GPU: значительно быстрее на NVIDIA; pip install onnxruntime-gpu\n"
            "ONNX CPU: чуть быстрее PyTorch на чистом CPU; pip install onnxruntime"
        )
        rf.addRow("Бэкенд (движок):", self._reid_engine_combo)

        self._reid_half = QCheckBox("FP16 для ReID  (только PyTorch + CUDA)")
        rf.addRow("", self._reid_half)

        root.addWidget(reid_grp)

        # ── GPU Info ───────────────────────────────────────────────────────────
        info_grp = QGroupBox("Информация о GPU")
        inf = QVBoxLayout(info_grp)

        self._gpu_lbl = QLabel()
        self._gpu_lbl.setWordWrap(True)
        self._gpu_lbl.setStyleSheet("color:#a6adc8; font-size:11px; font-family:monospace;")
        inf.addWidget(self._gpu_lbl)

        gpu_btn_row = QHBoxLayout()
        refresh_btn = QPushButton("🔄  Обновить")
        refresh_btn.clicked.connect(self._refresh_gpu)
        install_ort_gpu = QPushButton("📦  pip install onnxruntime-gpu")
        install_ort_gpu.setToolTip("Установить ONNX Runtime GPU в текущее окружение")
        install_ort_gpu.clicked.connect(lambda: self._pip_install("onnxruntime-gpu"))
        install_ort_cpu = QPushButton("📦  pip install onnxruntime")
        install_ort_cpu.clicked.connect(lambda: self._pip_install("onnxruntime"))
        gpu_btn_row.addWidget(refresh_btn)
        gpu_btn_row.addWidget(install_ort_gpu)
        gpu_btn_row.addWidget(install_ort_cpu)
        gpu_btn_row.addStretch()
        inf.addLayout(gpu_btn_row)

        # ONNX проверка (subprocess — чистый процесс, без кеша DLL текущего)
        onnx_check_row = QHBoxLayout()
        check_ort_btn = QPushButton("🔍  Проверить onnxruntime (subprocess)")
        check_ort_btn.setToolTip(
            "Запускает python в отдельном процессе и проверяет onnxruntime.\n"
            "Результат не зависит от DLL-кеша текущего приложения."
        )
        check_ort_btn.clicked.connect(self._check_onnx)
        onnx_check_row.addWidget(check_ort_btn)
        onnx_check_row.addStretch()
        inf.addLayout(onnx_check_row)

        self._onnx_status_lbl = QLabel("")
        self._onnx_status_lbl.setWordWrap(True)
        self._onnx_status_lbl.setStyleSheet("font-size:11px; font-family:monospace;")
        inf.addWidget(self._onnx_status_lbl)

        root.addWidget(info_grp)
        self._refresh_gpu()

        # ── Save / Reset ───────────────────────────────────────────────────────
        btn_row = QHBoxLayout()
        save_btn = QPushButton("💾  Сохранить")
        save_btn.setObjectName("btn_primary")
        save_btn.clicked.connect(self._save)
        reset_btn = QPushButton("↺  Сбросить")
        reset_btn.clicked.connect(self._load)
        btn_row.addWidget(save_btn)
        btn_row.addWidget(reset_btn)
        btn_row.addStretch()
        root.addLayout(btn_row)
        root.addStretch()

    # ── load / save ────────────────────────────────────────────────────────────

    def _load(self):
        c = self._cfg
        self._set_combo_data(self._yolo_dev, c.yolo_device)
        self._yolo_half.setChecked(c.yolo_half)
        self._set_combo_data(self._imgsz_combo, c.infer_imgsz)
        idx = self._tracker_combo.findText(c.tracking_config)
        if idx >= 0:
            self._tracker_combo.setCurrentIndex(idx)
        self._set_combo_data(self._reid_dev, c.reid_device)
        self._set_combo_data(self._reid_engine_combo, c.reid_engine)
        self._reid_half.setChecked(c.reid_half)

    def _save(self):
        c = self._cfg
        c.yolo_device    = self._yolo_dev.currentData()
        c.yolo_half      = self._yolo_half.isChecked()
        c.infer_imgsz    = self._imgsz_combo.currentData()
        c.tracking_config = self._tracker_combo.currentText()
        c.reid_device    = self._reid_dev.currentData()
        c.reid_engine    = self._reid_engine_combo.currentData()
        c.reid_half      = self._reid_half.isChecked()

        if c.yolo_half and c.yolo_device == "cpu":
            QMessageBox.warning(self, "FP16", "YOLO FP16 не работает на CPU.")
            return
        if c.reid_half and c.reid_device == "cpu":
            QMessageBox.warning(self, "FP16", "ReID FP16 не работает на CPU.")
            return

        QMessageBox.information(self, "Сохранено",
            "Настройки устройств применены.\nПерезапустите камеры.")
        self.settings_saved.emit()

    # ── helpers ────────────────────────────────────────────────────────────────

    @staticmethod
    def _set_combo_data(cb: QComboBox, value):
        idx = cb.findData(value)
        if idx >= 0:
            cb.setCurrentIndex(idx)

    def _refresh_gpu(self):
        lines = []
        try:
            import torch
            lines.append(f"PyTorch {torch.__version__}  |  CUDA: {torch.version.cuda or 'N/A'}")
            if torch.cuda.is_available():
                for i in range(torch.cuda.device_count()):
                    p = torch.cuda.get_device_properties(i)
                    used  = torch.cuda.memory_allocated(i) / 1024**3
                    total = p.total_memory / 1024**3
                    lines.append(
                        f"  cuda:{i}  {p.name}  {total:.1f} GB  "
                        f"({used:.2f} GB используется)"
                    )
            else:
                lines.append("  CUDA недоступен")
        except ImportError:
            lines.append("torch не установлен")

        try:
            import onnxruntime as ort
            providers = ort.get_available_providers()
            lines.append(f"ONNX Runtime {ort.__version__}  |  провайдеры: {', '.join(providers)}")
        except ImportError:
            lines.append("onnxruntime не установлен")

        self._gpu_lbl.setText("\n".join(lines))

    def _check_onnx(self):
        """Проверяет onnxruntime в отдельном subprocess (без DLL-кеша текущего процесса)."""
        import subprocess, sys
        self._onnx_status_lbl.setStyleSheet("font-size:11px; font-family:monospace; color:#a6adc8;")
        self._onnx_status_lbl.setText("⏳  Проверяем onnxruntime…")
        QApplication.processEvents()

        code = (
            "import onnxruntime as ort; "
            "v = ort.__version__; "
            "p = ort.get_available_providers(); "
            "print(f'VERSION={v}'); "
            "print(f'PROVIDERS={chr(44).join(p)}')"
        )
        r = subprocess.run(
            [sys.executable, "-c", code],
            capture_output=True, text=True, timeout=15
        )

        if r.returncode == 0:
            lines = {ln.split("=")[0]: ln.split("=", 1)[1]
                     for ln in r.stdout.strip().splitlines() if "=" in ln}
            ver = lines.get("VERSION", "?")
            provs = lines.get("PROVIDERS", "?")

            has_cuda = "CUDAExecutionProvider" in provs
            has_cpu  = "CPUExecutionProvider" in provs

            detail = f"onnxruntime {ver}\nПровайдеры: {provs}"
            if has_cuda:
                detail += "\n✅  GPU (CUDA) — доступен"
            if has_cpu:
                detail += "\n✅  CPU — доступен"
            if not has_cuda:
                detail += (
                    "\n⚠  CUDAExecutionProvider отсутствует.\n"
                    "   Для GPU нужен: pip install onnxruntime-gpu"
                )

            self._onnx_status_lbl.setStyleSheet(
                "font-size:11px; font-family:monospace; color:#a6e3a1;"
            )
            self._onnx_status_lbl.setText(detail)
        else:
            err = (r.stderr or r.stdout or "неизвестная ошибка").strip()
            if "DLL" in err or "WinError" in err or "_pybind_state" in err:
                advice = (
                    "🔴  DLL-конфликт: onnxruntime и PyTorch CUDA несовместимы.\n\n"
                    "Решение — установить CPU-версию onnxruntime:\n"
                    "  pip uninstall onnxruntime-gpu onnxruntime -y\n"
                    "  pip install onnxruntime\n\n"
                    "После этого перезапустите приложение и выберите:\n"
                    "  «ONNX Runtime — CPU»"
                )
            elif "No module named" in err or "ModuleNotFoundError" in err:
                advice = (
                    "🔴  onnxruntime не установлен.\n\n"
                    "Установите нужную версию:\n"
                    "  CPU:  pip install onnxruntime\n"
                    "  GPU:  pip install onnxruntime-gpu\n\n"
                    "Затем перезапустите приложение."
                )
            else:
                advice = f"🔴  Ошибка:\n{err[:500]}"

            self._onnx_status_lbl.setStyleSheet(
                "font-size:11px; font-family:monospace; color:#f38ba8;"
            )
            self._onnx_status_lbl.setText(advice)

    def _pip_install(self, package: str):
        import subprocess, sys
        reply = QMessageBox.question(
            self, "Установить пакет",
            f"Установить  {package}  через pip?",
            QMessageBox.Yes | QMessageBox.No
        )
        if reply != QMessageBox.Yes:
            return
        self._gpu_lbl.setText(f"Устанавливаем {package}…")
        QApplication.processEvents()
        r = subprocess.run(
            [sys.executable, "-m", "pip", "install", package,
             "--quiet", "--disable-pip-version-check"],
            capture_output=True, text=True
        )
        if r.returncode == 0:
            QMessageBox.information(self, "Готово", f"✔  {package} установлен.")
        else:
            QMessageBox.critical(self, "Ошибка", f"pip ошибка:\n{r.stderr[-1000:]}")
        self._refresh_gpu()


# ── Training tab ────────────────────────────────────────────────────────────────

class TrainingTab(QScrollArea):
    """YOLO training launcher + dataset statistics."""

    def __init__(self, cfg: "AppConfig", parent=None):
        super().__init__(parent)
        self.setWidgetResizable(True)
        self._cfg = cfg
        self._train_proc = None
        self._reid_proc = None
        container = QWidget()
        self.setWidget(container)
        self._build_ui(container)
        self._load()

    def _build_ui(self, container: QWidget):
        root = QVBoxLayout(container)
        root.setContentsMargins(16, 16, 16, 16)
        root.setSpacing(14)

        hdr = QLabel("Обучение и управление датасетом")
        hdr.setObjectName("hdr")
        root.addWidget(hdr)

        # ── YOLO Training ──────────────────────────────────────────────────────
        yolo_grp = QGroupBox("Обучение YOLO (fine-tune / дообучение)")
        yf = QFormLayout(yolo_grp)
        yf.setSpacing(10)

        # Data YAML
        data_row = QHBoxLayout()
        self._data_edit = QLineEdit()
        self._data_edit.setPlaceholderText("dataset.yaml  или  /путь/к/data.yaml")
        data_browse = QPushButton("📂")
        data_browse.setFixedWidth(36)
        data_browse.clicked.connect(self._browse_data)
        gen_btn = QPushButton("⚙  Создать dataset.yaml")
        gen_btn.setToolTip(
            "Автоматически создаёт dataset.yaml из папок датасета в настройках."
        )
        gen_btn.clicked.connect(self._generate_yaml)
        data_row.addWidget(self._data_edit)
        data_row.addWidget(data_browse)
        data_row.addWidget(gen_btn)
        yf.addRow("Data YAML:", data_row)

        # Base model
        base_row = QHBoxLayout()
        self._base_model_edit = QLineEdit()
        self._base_model_edit.setPlaceholderText("yolo11s.pt")
        base_browse = QPushButton("📂")
        base_browse.setFixedWidth(36)
        base_browse.clicked.connect(self._browse_base_model)
        base_row.addWidget(self._base_model_edit)
        base_row.addWidget(base_browse)
        yf.addRow("Базовая модель:", base_row)

        # Numeric params
        self._epochs_spin = QSpinBox()
        self._epochs_spin.setRange(1, 2000)
        self._epochs_spin.setToolTip("Количество эпох обучения. 50–100 — дообучение, 300+ — обучение с нуля.")
        yf.addRow("Эпохи:", self._epochs_spin)

        self._batch_spin = QSpinBox()
        self._batch_spin.setRange(1, 256)
        self._batch_spin.setToolTip(
            "Размер батча. Уменьшите при нехватке VRAM (ошибка OOM).\n"
            "8 — безопасно для 4 ГБ VRAM; 16–32 — для 8+ ГБ."
        )
        yf.addRow("Batch size:", self._batch_spin)

        self._train_imgsz_combo = QComboBox()
        for sz in [320, 416, 512, 640, 832, 1280]:
            self._train_imgsz_combo.addItem(f"{sz} px", sz)
        yf.addRow("Image size:", self._train_imgsz_combo)

        self._lr_spin = QDoubleSpinBox()
        self._lr_spin.setRange(0.0001, 0.1)
        self._lr_spin.setSingleStep(0.001)
        self._lr_spin.setDecimals(4)
        self._lr_spin.setToolTip(
            "Начальная скорость обучения (lr0).\n"
            "0.01 — стандарт; 0.001–0.005 — дообучение предобученных весов."
        )
        yf.addRow("LR (lr0):", self._lr_spin)

        self._patience_spin = QSpinBox()
        self._patience_spin.setRange(0, 500)
        self._patience_spin.setToolTip(
            "Early stopping: остановить если нет улучшения N эпох.\n0 = отключить."
        )
        yf.addRow("Patience (early stop):", self._patience_spin)

        self._workers_spin = QSpinBox()
        self._workers_spin.setRange(0, 16)
        self._workers_spin.setToolTip(
            "DataLoader workers. 0 = загрузка в главном потоке (без параллелизма).\n"
            "2–4 — обычно оптимально. На Windows уменьшите если зависает."
        )
        yf.addRow("Workers:", self._workers_spin)

        # Device for training
        self._train_dev_combo = QComboBox()
        self._train_dev_combo.addItem("0  — первый GPU (YOLO формат)", "0")
        self._train_dev_combo.addItem("cpu  — CPU", "cpu")
        try:
            import torch
            for i in range(torch.cuda.device_count()):
                p = torch.cuda.get_device_properties(i)
                self._train_dev_combo.addItem(f"{i}  — {p.name}", str(i))
        except Exception:
            pass
        yf.addRow("Устройство обучения:", self._train_dev_combo)

        # Output project
        proj_row = QHBoxLayout()
        self._project_edit = QLineEdit()
        self._project_edit.setPlaceholderText("runs/train")
        proj_browse = QPushButton("📂")
        proj_browse.setFixedWidth(36)
        proj_browse.clicked.connect(lambda: self._browse_dir(self._project_edit))
        proj_row.addWidget(self._project_edit)
        proj_row.addWidget(proj_browse)
        yf.addRow("Папка результатов:", proj_row)

        root.addWidget(yolo_grp)

        # ── Train control ──────────────────────────────────────────────────────
        ctrl_row = QHBoxLayout()
        self._train_btn = QPushButton("▶  Начать обучение")
        self._train_btn.setObjectName("btn_primary")
        self._train_btn.clicked.connect(self._start_training)
        self._stop_btn = QPushButton("⏹  Остановить")
        self._stop_btn.setEnabled(False)
        self._stop_btn.clicked.connect(self._stop_training)
        self._open_results_btn = QPushButton("📁  Открыть папку результатов")
        self._open_results_btn.clicked.connect(self._open_results)
        ctrl_row.addWidget(self._train_btn)
        ctrl_row.addWidget(self._stop_btn)
        ctrl_row.addWidget(self._open_results_btn)
        ctrl_row.addStretch()
        root.addLayout(ctrl_row)

        # Training log
        self._train_log = QPlainTextEdit()
        self._train_log.setReadOnly(True)
        self._train_log.setMaximumBlockCount(500)
        self._train_log.setStyleSheet(
            "QPlainTextEdit { background:#11111b; color:#cdd6f4; "
            "font-family:Consolas,monospace; font-size:11px; border-radius:4px; }"
        )
        self._train_log.setFixedHeight(180)
        root.addWidget(self._train_log)

        # ── Dataset Statistics ─────────────────────────────────────────────────
        ds_grp = QGroupBox("Статистика датасета")
        df = QFormLayout(ds_grp)
        df.setSpacing(8)

        self._stat_images = QLabel("—")
        self._stat_labels = QLabel("—")
        self._stat_reid   = QLabel("—")
        df.addRow("Изображений YOLO:", self._stat_images)
        df.addRow("Разметок YOLO:",    self._stat_labels)
        df.addRow("Кропов ReID:",      self._stat_reid)

        ds_btn_row = QHBoxLayout()
        refresh_ds = QPushButton("🔄  Обновить статистику")
        refresh_ds.clicked.connect(self._refresh_stats)
        open_yolo = QPushButton("📂  Папка YOLO")
        open_yolo.clicked.connect(lambda: self._open_folder(self._cfg.yolo_images_dir))
        open_reid = QPushButton("📂  Папка ReID")
        open_reid.clicked.connect(lambda: self._open_folder(self._cfg.reid_dir))
        ds_btn_row.addWidget(refresh_ds)
        ds_btn_row.addWidget(open_yolo)
        ds_btn_row.addWidget(open_reid)
        ds_btn_row.addStretch()
        df.addRow("", ds_btn_row)

        root.addWidget(ds_grp)

        # ── ReID Fine-Tuning ──────────────────────────────────────────────────
        reid_grp = QGroupBox("Дообучение ReID (OSNet x1.0 → fine-tune)")
        rdf = QFormLayout(reid_grp)
        rdf.setSpacing(10)

        reid_data_row = QHBoxLayout()
        self._reid_data_edit = QLineEdit()
        self._reid_data_edit.setPlaceholderText("Dataset/datasetReID  — папка с кропами")
        reid_data_browse = QPushButton("📂")
        reid_data_browse.setFixedWidth(36)
        reid_data_browse.clicked.connect(lambda: self._browse_dir(self._reid_data_edit))
        reid_data_row.addWidget(self._reid_data_edit)
        reid_data_row.addWidget(reid_data_browse)
        rdf.addRow("Папка ReID датасета:", reid_data_row)

        reid_output_row = QHBoxLayout()
        self._reid_output_edit = QLineEdit()
        self._reid_output_edit.setPlaceholderText("runs/reid_train  — куда сохранить результат")
        reid_output_browse = QPushButton("📂")
        reid_output_browse.setFixedWidth(36)
        reid_output_browse.clicked.connect(lambda: self._browse_dir(self._reid_output_edit))
        reid_output_row.addWidget(self._reid_output_edit)
        reid_output_row.addWidget(reid_output_browse)
        rdf.addRow("Папка результатов:", reid_output_row)

        self._reid_epochs_spin = QSpinBox()
        self._reid_epochs_spin.setRange(1, 500)
        self._reid_epochs_spin.setValue(50)
        rdf.addRow("Эпохи:", self._reid_epochs_spin)

        self._reid_batch_spin = QSpinBox()
        self._reid_batch_spin.setRange(4, 256)
        self._reid_batch_spin.setValue(32)
        rdf.addRow("Batch size:", self._reid_batch_spin)

        self._reid_lr_spin = QDoubleSpinBox()
        self._reid_lr_spin.setRange(0.00001, 0.01)
        self._reid_lr_spin.setValue(0.0003)
        self._reid_lr_spin.setDecimals(6)
        rdf.addRow("Learning rate:", self._reid_lr_spin)

        reid_hint = QLabel(
            "⚠  Требует: pip install torchreid\n"
            "Структура датасета: reid_dir/00001/img.jpg, reid_dir/00002/img.jpg, …\n"
            "После обучения экспортируйте ONNX во вкладке «ReID» → «Экспорт ONNX»."
        )
        reid_hint.setStyleSheet("color:#6c7086; font-size:11px;")
        reid_hint.setWordWrap(True)
        rdf.addRow("", reid_hint)

        reid_ctrl_row = QHBoxLayout()
        self._reid_train_btn = QPushButton("🧠  Дообучить OSNet")
        self._reid_train_btn.setObjectName("btn_primary")
        self._reid_train_btn.clicked.connect(self._start_reid_training)
        self._reid_stop_btn = QPushButton("⏹  Остановить")
        self._reid_stop_btn.setEnabled(False)
        self._reid_stop_btn.clicked.connect(self._stop_reid_training)
        self._open_reid_results_btn = QPushButton("📁  Открыть результаты")
        self._open_reid_results_btn.clicked.connect(
            lambda: self._open_folder(self._reid_output_edit.text().strip() or "runs/reid_train")
        )
        reid_ctrl_row.addWidget(self._reid_train_btn)
        reid_ctrl_row.addWidget(self._reid_stop_btn)
        reid_ctrl_row.addWidget(self._open_reid_results_btn)
        reid_ctrl_row.addStretch()
        rdf.addRow("", reid_ctrl_row)

        self._reid_train_log = QPlainTextEdit()
        self._reid_train_log.setReadOnly(True)
        self._reid_train_log.setMaximumBlockCount(300)
        self._reid_train_log.setStyleSheet(
            "QPlainTextEdit { background:#11111b; color:#cdd6f4; "
            "font-family:Consolas,monospace; font-size:11px; border-radius:4px; }"
        )
        self._reid_train_log.setFixedHeight(120)
        rdf.addRow("", self._reid_train_log)

        root.addWidget(reid_grp)
        root.addStretch()

    # ── load / save ────────────────────────────────────────────────────────────

    def _load(self):
        c = self._cfg
        self._base_model_edit.setText(c.model_path)
        self._epochs_spin.setValue(c.train_epochs)
        self._batch_spin.setValue(c.train_batch)
        self._patience_spin.setValue(c.train_patience)
        self._workers_spin.setValue(c.train_workers)
        self._project_edit.setText(c.train_project)
        self._lr_spin.setValue(c.train_lr0)
        # imgsz
        idx = self._train_imgsz_combo.findData(c.train_imgsz)
        self._train_imgsz_combo.setCurrentIndex(max(0, idx))
        # device
        idx2 = self._train_dev_combo.findData(c.train_device)
        self._train_dev_combo.setCurrentIndex(max(0, idx2))

    def _save_params(self):
        """Сохраняет параметры обучения в cfg без диалога."""
        c = self._cfg
        c.train_epochs   = self._epochs_spin.value()
        c.train_batch    = self._batch_spin.value()
        c.train_imgsz    = self._train_imgsz_combo.currentData()
        c.train_lr0      = self._lr_spin.value()
        c.train_patience = self._patience_spin.value()
        c.train_workers  = self._workers_spin.value()
        c.train_device   = self._train_dev_combo.currentData()
        c.train_project  = self._project_edit.text().strip() or "runs/train"

    # ── Training control ───────────────────────────────────────────────────────

    def _start_training(self):
        data_yaml = self._data_edit.text().strip()
        if not data_yaml:
            QMessageBox.warning(self, "Ошибка", "Укажите путь к data YAML файлу.")
            return
        if not Path(data_yaml).exists():
            QMessageBox.warning(self, "Ошибка",
                f"Файл не найден:\n{data_yaml}\n\nИспользуйте '⚙ Создать dataset.yaml'.")
            return

        self._save_params()
        c = self._cfg

        cmd = [
            sys.executable, "-m", "ultralytics",
            "train",
            f"model={self._base_model_edit.text().strip() or c.model_path}",
            f"data={data_yaml}",
            f"epochs={c.train_epochs}",
            f"batch={c.train_batch}",
            f"imgsz={c.train_imgsz}",
            f"lr0={c.train_lr0}",
            f"patience={c.train_patience}",
            f"workers={c.train_workers}",
            f"device={c.train_device}",
            f"project={c.train_project}",
        ]

        self._train_log.clear()
        self._train_log.appendPlainText("Запуск обучения YOLO...\n> " + " ".join(cmd))

        import subprocess
        self._train_proc = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
        )
        self._train_btn.setEnabled(False)
        self._stop_btn.setEnabled(True)

        # Читаем вывод в QTimer
        self._train_timer = QTimer(self)
        self._train_timer.timeout.connect(self._poll_training)
        self._train_timer.start(200)

    def _poll_training(self):
        if self._train_proc is None:
            return
        try:
            line = self._train_proc.stdout.readline()
            if line:
                self._train_log.appendPlainText(line.rstrip())
                sb = self._train_log.verticalScrollBar()
                sb.setValue(sb.maximum())
        except Exception:
            pass

        if self._train_proc.poll() is not None:
            self._train_timer.stop()
            rc = self._train_proc.returncode
            self._train_proc = None
            self._train_btn.setEnabled(True)
            self._stop_btn.setEnabled(False)
            if rc == 0:
                self._train_log.appendPlainText("\n✔  Обучение завершено успешно!")
            else:
                self._train_log.appendPlainText(f"\n✘  Обучение завершено с кодом {rc}")

    def _stop_training(self):
        if self._train_proc:
            self._train_proc.terminate()
            self._train_log.appendPlainText("\n⏹  Обучение остановлено пользователем.")

    # ── Dataset helpers ────────────────────────────────────────────────────────

    def _refresh_stats(self):
        def _count(folder, ext):
            p = Path(folder)
            if not p.exists():
                return 0
            return sum(1 for _ in p.rglob(f"*.{ext}"))

        imgs  = _count(self._cfg.yolo_images_dir, "jpg") + _count(self._cfg.yolo_images_dir, "png")
        lbls  = _count(self._cfg.yolo_labels_dir, "txt")
        crops = _count(self._cfg.reid_dir, "jpg") + _count(self._cfg.reid_dir, "png")
        self._stat_images.setText(f"{imgs:,}")
        self._stat_labels.setText(f"{lbls:,}")
        self._stat_reid.setText(f"{crops:,}")

    def _generate_yaml(self):
        """Создаёт простой dataset.yaml для YOLO из настроек датасета."""
        c = self._cfg
        names = {cid: name for cid, name in __import__("tracker_core").COCO_NAMES.items()
                 if cid in c.classes}
        data = {
            "path": ".",
            "train": c.yolo_images_dir,
            "val":   c.yolo_images_dir,  # пользователь заменит вручную
            "names": {i: n for i, n in enumerate(names.values())},
        }
        out_path = Path("dataset.yaml")
        import yaml as _yaml
        with open(out_path, "w", encoding="utf-8") as f:
            _yaml.dump(data, f, allow_unicode=True, sort_keys=False)
        self._data_edit.setText(str(out_path.resolve()))
        QMessageBox.information(
            self, "Готово",
            f"dataset.yaml создан:\n{out_path.resolve()}\n\n"
            "⚠️  Проверьте поля 'train' и 'val' перед обучением!"
        )

    def _open_results(self):
        self._open_folder(self._cfg.train_project)

    @staticmethod
    def _open_folder(path: str):
        import webbrowser
        p = Path(path)
        p.mkdir(parents=True, exist_ok=True)
        webbrowser.open(p.resolve().as_uri())

    def _browse_data(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Выбрать data YAML", "", "YAML (*.yaml *.yml);;Все файлы (*)"
        )
        if path:
            self._data_edit.setText(path)

    def _browse_base_model(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "Выбрать базовую модель", "", "PyTorch (*.pt);;Все файлы (*)"
        )
        if path:
            self._base_model_edit.setText(path)

    def _browse_dir(self, edit: QLineEdit):
        d = QFileDialog.getExistingDirectory(self, "Выбрать папку", edit.text() or ".")
        if d:
            edit.setText(d)

    def reload_cfg(self):
        """Вызывается когда cfg изменился снаружи."""
        self._load()

    # ── ReID Fine-tuning methods ───────────────────────────────────────────────

    def _start_reid_training(self):
        data_dir = self._reid_data_edit.text().strip() or self._cfg.reid_dir
        output_dir = self._reid_output_edit.text().strip() or "runs/reid_train"
        epochs = self._reid_epochs_spin.value()
        batch  = self._reid_batch_spin.value()
        lr     = self._reid_lr_spin.value()

        # Build command: use embedded reid_train_helper.py or inline script
        script = str(app_dir() / "reid_train_helper.py")
        if not Path(script).exists():
            # Inline: show instructions
            QMessageBox.information(
                self, "Дообучение ReID",
                f"Для дообучения OSNet запустите:\n\n"
                f"python reid_train_helper.py \\\n"
                f"  --data-dir \"{data_dir}\" \\\n"
                f"  --output-dir \"{output_dir}\" \\\n"
                f"  --epochs {epochs} --batch {batch} --lr {lr}\n\n"
                "Файл reid_train_helper.py будет создан автоматически при первом запуске,\n"
                "или обратитесь к документации torchreid."
            )
            return

        import subprocess
        cmd = [
            sys.executable, script,
            "--data-dir", data_dir,
            "--output-dir", output_dir,
            "--epochs", str(epochs),
            "--batch", str(batch),
            "--lr", str(lr),
        ]
        self._reid_proc = subprocess.Popen(
            cmd, stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
            text=True, bufsize=1, encoding="utf-8", errors="replace"
        )
        self._reid_train_btn.setEnabled(False)
        self._reid_stop_btn.setEnabled(True)
        self._reid_train_log.clear()
        self._reid_train_log.appendPlainText(f"▶  Запуск: {' '.join(cmd)}")

        # Poll subprocess output
        self._reid_poll_timer = QTimer(self)
        self._reid_poll_timer.timeout.connect(self._poll_reid_training)
        self._reid_poll_timer.start(200)

    def _poll_reid_training(self):
        if self._reid_proc is None:
            return
        line = self._reid_proc.stdout.readline()
        if line:
            self._reid_train_log.appendPlainText(line.rstrip())
        if self._reid_proc.poll() is not None:
            # process finished
            self._reid_poll_timer.stop()
            rc = self._reid_proc.returncode
            self._reid_train_btn.setEnabled(True)
            self._reid_stop_btn.setEnabled(False)
            msg = "✔  Обучение завершено!" if rc == 0 else f"✘  Завершено с кодом {rc}"
            self._reid_train_log.appendPlainText(msg)
            self._reid_proc = None

    def _stop_reid_training(self):
        if self._reid_proc and self._reid_proc.poll() is None:
            self._reid_proc.terminate()
        self._reid_train_btn.setEnabled(True)
        self._reid_stop_btn.setEnabled(False)


# ── Statistics tab ─────────────────────────────────────────────────────────────

class StatisticsTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._cameras: List[CameraEntry] = []
        self._processors: List[Optional[CameraProcessor]] = []
        self._match_storage = None
        self._build_ui()
        self._timer = QTimer(self)
        self._timer.timeout.connect(self.refresh)
        self._timer.start(2000)

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(10)

        hdr_row = QHBoxLayout()
        lbl = QLabel("Статистика сбора данных")
        lbl.setObjectName("hdr")
        hdr_row.addWidget(lbl)
        hdr_row.addStretch()

        refresh_btn = QPushButton("⟳  Обновить")
        refresh_btn.setObjectName("btn_primary")
        refresh_btn.clicked.connect(self.refresh)
        hdr_row.addWidget(refresh_btn)

        open_btn = QPushButton("📂  Открыть папку датасета")
        open_btn.clicked.connect(self._open_dataset)
        hdr_row.addWidget(open_btn)

        root.addLayout(hdr_row)

        # Summary cards row
        self._total_active = self._card("0", "Активных предметов")
        self._total_seen   = self._card("0", "Уникальных треков")
        self._total_yolo   = self._card("0", "Кадров YOLO")
        self._total_reid   = self._card("0", "Кропов ReID")
        cards_row = QHBoxLayout()
        for c in (self._total_active, self._total_seen,
                  self._total_yolo, self._total_reid):
            cards_row.addWidget(c)
        root.addLayout(cards_row)

        # Per-camera table
        self._table = QTableWidget(0, 7)
        self._table.setHorizontalHeaderLabels(
            ["Камера", "Стол", "Статус", "FPS",
             "Активно", "Всего треков", "YOLO", "ReID"]
        )
        self._table.setColumnCount(8)
        self._table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self._table.setEditTriggers(QTableWidget.NoEditTriggers)
        self._table.verticalHeader().setVisible(False)
        self._table.setAlternatingRowColors(True)
        self._table.setStyleSheet(
            "QTableWidget { alternate-background-color: #181825; }"
        )
        root.addWidget(self._table, 1)

        # ── ReID совпадения (из SQLite) ──────────────────────────────────────
        sep = QFrame(); sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet("color:#313244;")
        root.addWidget(sep)

        reid_hdr = QLabel("Статистика совпадений ReID")
        reid_hdr.setObjectName("hdr")
        root.addWidget(reid_hdr)

        reid_cards = QHBoxLayout()
        self._r_total   = self._card("—", "Совпадений всего")
        self._r_today   = self._card("—", "Сегодня")
        self._r_sim     = self._card("—", "Ср. сходство %")
        self._r_transit = self._card("—", "Ср. время пути, сек")
        for c in (self._r_total, self._r_today, self._r_sim, self._r_transit):
            reid_cards.addWidget(c)
        root.addLayout(reid_cards)

        verdict_row = QHBoxLayout()
        self._r_high = self._card("—", "Точно тот же  ✔")
        self._r_mid  = self._card("—", "Вероятно тот же  ?")
        self._r_low  = self._card("—", "Другой  ✘")
        self._r_high._value_lbl.setStyleSheet(
            "font-size:24px; font-weight:bold; color:#a6e3a1;")
        self._r_mid._value_lbl.setStyleSheet(
            "font-size:24px; font-weight:bold; color:#f9e2af;")
        self._r_low._value_lbl.setStyleSheet(
            "font-size:24px; font-weight:bold; color:#f38ba8;")
        for c in (self._r_high, self._r_mid, self._r_low):
            verdict_row.addWidget(c)
        root.addLayout(verdict_row)

        # Активность по часам (последние 24ч)
        hourly_lbl = QLabel("Совпадения по часам (последние 24ч):")
        hourly_lbl.setStyleSheet("color:#6c7086; font-size:11px;")
        root.addWidget(hourly_lbl)
        self._r_hourly = QLabel("—")
        self._r_hourly.setStyleSheet(
            "background:#1e1e2e; padding:6px; border-radius:4px; "
            "font-family: monospace; font-size:11px;")
        self._r_hourly.setWordWrap(True)
        root.addWidget(self._r_hourly)

    def _card(self, value: str, label: str) -> QFrame:
        f = QFrame()
        f.setStyleSheet(
            "QFrame { background:#313244; border-radius:8px; padding:8px; }"
        )
        lay = QVBoxLayout(f)
        lay.setSpacing(2)
        v_lbl = QLabel(value)
        v_lbl.setStyleSheet("font-size:28px; font-weight:bold; color:#89b4fa;")
        v_lbl.setAlignment(Qt.AlignCenter)
        n_lbl = QLabel(label)
        n_lbl.setStyleSheet("font-size:11px; color:#6c7086;")
        n_lbl.setAlignment(Qt.AlignCenter)
        lay.addWidget(v_lbl)
        lay.addWidget(n_lbl)
        f._value_lbl = v_lbl
        return f

    def set_data(self, cameras: List[CameraEntry],
                 processors: List[Optional[CameraProcessor]],
                 match_storage=None):
        self._cameras       = cameras
        self._processors    = processors
        self._match_storage = match_storage
        self.refresh()

    @pyqtSlot()
    def refresh(self):
        try:
            self._refresh_impl()
        except Exception as exc:
            logger.debug("StatisticsTab.refresh error: %s", exc)

    def _refresh_impl(self):
        tot_active = tot_seen = tot_yolo = tot_reid = 0
        self._table.setRowCount(len(self._cameras))

        for i, cam in enumerate(self._cameras):
            proc  = self._processors[i] if i < len(self._processors) else None
            stats = proc.stats if proc else None

            def cell(v): return QTableWidgetItem(str(v))

            self._table.setItem(i, 0, cell(cam.name))
            self._table.setItem(i, 1, cell(cam.counter_id))
            self._table.setItem(i, 2, cell(
                STATUS_LABELS.get(stats.status, "—") if stats else "—"
            ))
            self._table.setItem(i, 3, cell(f"{stats.fps:.1f}" if stats else "—"))
            self._table.setItem(i, 4, cell(stats.active_count if stats else 0))
            self._table.setItem(i, 5, cell(stats.total_seen   if stats else 0))
            self._table.setItem(i, 6, cell(stats.yolo_saved   if stats else 0))
            self._table.setItem(i, 7, cell(stats.reid_saved   if stats else 0))

            if stats:
                tot_active += stats.active_count
                tot_seen   += stats.total_seen
                tot_yolo   += stats.yolo_saved
                tot_reid   += stats.reid_saved

        self._total_active._value_lbl.setText(str(tot_active))
        self._total_seen._value_lbl.setText(str(tot_seen))
        self._total_yolo._value_lbl.setText(str(tot_yolo))
        self._total_reid._value_lbl.setText(str(tot_reid))

        # ── ReID статистика из SQLite ─────────────────────────────────────────
        if self._match_storage is not None:
            s = self._match_storage.stats_summary()
            self._r_total._value_lbl.setText(str(s["total"]))
            self._r_today._value_lbl.setText(str(s["today"]))
            self._r_sim._value_lbl.setText(
                f"{s['avg_sim']:.1f}" if s["total"] else "—")
            self._r_transit._value_lbl.setText(
                f"{s['avg_transit']:.1f}" if s["total"] else "—")
            self._r_high._value_lbl.setText(str(s["verdict_high"]))
            self._r_mid._value_lbl.setText(str(s["verdict_mid"]))
            self._r_low._value_lbl.setText(str(s["verdict_low"]))
            if s["hourly"]:
                max_cnt = max((c for _, c in s["hourly"]), default=1) or 1
                lines = []
                for h, cnt in s["hourly"]:
                    bar = "█" * int(cnt / max_cnt * 20)
                    lines.append(f"{h}  {bar} {cnt}")
                self._r_hourly.setText("\n".join(lines))
            else:
                self._r_hourly.setText("Нет данных за последние 24ч")

    def _open_dataset(self):
        path = Path("Dataset")
        path.mkdir(exist_ok=True)
        if sys.platform == "win32":
            os.startfile(str(path))
        elif sys.platform == "darwin":
            subprocess.Popen(["open", str(path)])
        else:
            subprocess.Popen(["xdg-open", str(path)])


# ── Matches tab ────────────────────────────────────────────────────────────────

def _crop_to_pixmap(crop: np.ndarray, size: int) -> QPixmap:
    if crop is None or crop.size == 0 or len(crop.shape) < 2:
        pm = QPixmap(size, size)
        pm.fill(QColor("#313244"))
        return pm
    try:
        rgb  = cv2.cvtColor(crop, cv2.COLOR_BGR2RGB)
        h, w = rgb.shape[:2]
        qimg = QImage(rgb.data, w, h, w * 3, QImage.Format_RGB888)
        return QPixmap.fromImage(qimg).scaled(
            QSize(size, size), Qt.KeepAspectRatio, Qt.SmoothTransformation
        )
    except Exception:
        pm = QPixmap(size, size)
        pm.fill(QColor("#313244"))
        return pm


class MatchCard(QFrame):
    """One match event: source crop ↔ query crop + verdict."""

    def __init__(self, mr: MatchResult, parent=None):
        super().__init__(parent)
        self.match_result = mr
        self.setFrameShape(QFrame.StyledPanel)
        color = mr.verdict_color
        self.setStyleSheet(
            f"MatchCard {{ background:#1e1e2e; border:2px solid {color};"
            f"border-radius:8px; }}"
        )
        self._build(mr, color)

    def _build(self, mr: MatchResult, color: str):
        root = QHBoxLayout(self)
        root.setContentsMargins(10, 8, 10, 8)
        root.setSpacing(12)

        # Source photo
        src_col = QVBoxLayout()
        src_img = QLabel()
        src_img.setPixmap(_crop_to_pixmap(mr.source_entry.crop, 100))
        src_img.setAlignment(Qt.AlignCenter)
        src_lbl = QLabel(f"📤 {mr.source_entry.cam_name}\n"
                         f"Стол #{mr.source_entry.counter_id}  "
                         f"Трек #{mr.source_entry.track_id}")
        src_lbl.setStyleSheet("color:#a6adc8; font-size:10px;")
        src_lbl.setAlignment(Qt.AlignCenter)
        src_col.addWidget(src_img)
        src_col.addWidget(src_lbl)
        root.addLayout(src_col)

        # Verdict column
        v_col = QVBoxLayout()
        v_col.setAlignment(Qt.AlignCenter)
        arrow = QLabel("→")
        arrow.setStyleSheet("font-size:28px; color:#45475a;")
        arrow.setAlignment(Qt.AlignCenter)
        sim_lbl = QLabel(f"{mr.similarity * 100:.1f}%")
        sim_lbl.setStyleSheet(
            f"font-size:22px; font-weight:bold; color:{color};"
        )
        sim_lbl.setAlignment(Qt.AlignCenter)
        verdict_lbl = QLabel(mr.verdict)
        verdict_lbl.setStyleSheet(
            f"font-size:12px; font-weight:bold; color:{color};"
        )
        verdict_lbl.setAlignment(Qt.AlignCenter)
        ts_lbl = QLabel(
            datetime.fromtimestamp(mr.timestamp).strftime("%H:%M:%S")
        )
        ts_lbl.setStyleSheet("color:#585b70; font-size:10px;")
        ts_lbl.setAlignment(Qt.AlignCenter)
        transit_lbl = QLabel(f"⏱ {mr.transit_seconds:.0f} сек в пути")
        transit_lbl.setStyleSheet("color:#6c7086; font-size:10px;")
        transit_lbl.setAlignment(Qt.AlignCenter)
        for w in (arrow, sim_lbl, verdict_lbl, ts_lbl, transit_lbl):
            v_col.addWidget(w)
        root.addLayout(v_col)

        # Query photo
        q_col = QVBoxLayout()
        q_img = QLabel()
        q_img.setPixmap(_crop_to_pixmap(mr.query_crop, 100))
        q_img.setAlignment(Qt.AlignCenter)
        q_lbl = QLabel(f"📥 {mr.query_cam_name}\n"
                        f"Стол #{mr.query_counter_id}  "
                        f"Трек #{mr.query_track_id}")
        q_lbl.setStyleSheet("color:#a6adc8; font-size:10px;")
        q_lbl.setAlignment(Qt.AlignCenter)
        q_col.addWidget(q_img)
        q_col.addWidget(q_lbl)
        root.addLayout(q_col)


# ── Analytics tab ──────────────────────────────────────────────────────────────

class AnalyticsTab(QWidget):
    """Real-time analytics charts using pyqtgraph (if available) or plain text fallback."""

    def __init__(self, parent=None):
        super().__init__(parent)
        self._processors = []
        self._match_storage = None
        self._history_matches: List[float] = []   # timestamps of matches
        self._history_sims: List[float]    = []   # similarity values at each match
        self._db_size_history: List[int]   = []   # DB size samples
        self._db_size_times:  List[float]  = []
        self._start_time = time.time()

        self._pg_available = False
        try:
            import pyqtgraph as pg
            self._pg_available = True
            self._pg = pg
        except ImportError:
            pass

        self._build_ui()

        self._refresh_timer = QTimer(self)
        self._refresh_timer.timeout.connect(self._refresh)
        self._refresh_timer.start(2000)

    def set_data(self, processors, match_storage, reid_db=None):
        self._processors    = processors or []
        self._match_storage = match_storage
        self._reid_db       = reid_db

    def add_match(self, match_result):
        """Called when a new match is found. Pass MatchResult."""
        self._history_matches.append(time.time())
        try:
            self._history_sims.append(float(match_result.similarity))
        except Exception:
            self._history_sims.append(0.0)
        if len(self._history_matches) > 500:
            self._history_matches = self._history_matches[-500:]
            self._history_sims    = self._history_sims[-500:]

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(12, 12, 12, 12)
        root.setSpacing(10)

        hdr = QLabel("📊  Аналитика в реальном времени")
        hdr.setObjectName("hdr")
        root.addWidget(hdr)

        # ── Summary row ──────────────────────────────────────────────────────
        summary_row = QHBoxLayout()
        self._lbl_total_matches = self._stat_box("Совпадений", "0")
        self._lbl_avg_sim       = self._stat_box("Ср. сходство", "—")
        self._lbl_active_tracks = self._stat_box("Активных треков", "0")
        self._lbl_db_size       = self._stat_box("Записей в БД", "0")
        self._lbl_fps_avg       = self._stat_box("Ср. FPS", "0")
        for w in [self._lbl_total_matches, self._lbl_avg_sim,
                  self._lbl_active_tracks, self._lbl_db_size, self._lbl_fps_avg]:
            summary_row.addWidget(w)
        root.addLayout(summary_row)

        if self._pg_available:
            self._build_charts(root)
        else:
            no_pg = QLabel(
                "ℹ  Для графиков установите pyqtgraph:\n"
                "    pip install pyqtgraph\n\n"
                "Цифровая статистика обновляется каждые 2 секунды."
            )
            no_pg.setStyleSheet("color:#6c7086; font-size:12px;")
            root.addWidget(no_pg)

        # ── Per-camera FPS table ──────────────────────────────────────────────
        fps_grp = QGroupBox("FPS по камерам")
        fps_layout = QVBoxLayout(fps_grp)
        self._fps_table = QTableWidget(0, 4)
        self._fps_table.setHorizontalHeaderLabels(["Камера", "Роль", "FPS", "Треков"])
        self._fps_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self._fps_table.setEditTriggers(QTableWidget.NoEditTriggers)
        self._fps_table.setMaximumHeight(180)
        fps_layout.addWidget(self._fps_table)
        root.addWidget(fps_grp)

        root.addStretch()

    def _stat_box(self, label: str, value: str) -> QWidget:
        w = QFrame()
        w.setFrameShape(QFrame.StyledPanel)
        w.setStyleSheet(
            "QFrame { background:#313244; border-radius:8px; padding:8px; }"
        )
        vl = QVBoxLayout(w)
        vl.setContentsMargins(12, 8, 12, 8)
        vl.setSpacing(2)
        lbl = QLabel(label)
        lbl.setStyleSheet("color:#a6adc8; font-size:10px;")
        val = QLabel(value)
        val.setStyleSheet("color:#89b4fa; font-size:20px; font-weight:bold;")
        val.setObjectName("val")
        vl.addWidget(lbl)
        vl.addWidget(val)
        return w

    def _set_stat(self, box: QWidget, value: str):
        lbl = box.findChild(QLabel, "val")
        if lbl:
            lbl.setText(value)

    def _build_charts(self, root: QVBoxLayout):
        pg = self._pg
        pg.setConfigOption("background", "#1e1e2e")
        pg.setConfigOption("foreground", "#cdd6f4")

        splitter = QSplitter(Qt.Vertical)

        # Match timeline (matches per minute over last 30 min)
        self._plot_timeline = pg.PlotWidget(title="Совпадения во времени")
        self._plot_timeline.setLabel("left", "Совпадений / мин")
        self._plot_timeline.setLabel("bottom", "Время (мин назад)")
        self._curve_timeline = self._plot_timeline.plot(pen=pg.mkPen("#89b4fa", width=2))
        splitter.addWidget(self._plot_timeline)

        # Similarity histogram
        self._plot_sim_hist = pg.PlotWidget(title="Распределение сходства")
        self._plot_sim_hist.setLabel("left", "Частота")
        self._plot_sim_hist.setLabel("bottom", "Сходство (cosine)")
        self._bar_sim = pg.BarGraphItem(x=[], height=[], width=0.04, brush="#a6e3a1")
        self._plot_sim_hist.addItem(self._bar_sim)
        splitter.addWidget(self._plot_sim_hist)

        root.addWidget(splitter, 1)

    def _refresh(self):
        try:
            self._refresh_impl()
        except Exception as exc:
            logger.debug("AnalyticsTab._refresh error: %s", exc)

    def _refresh_impl(self):
        # Update summary boxes
        total = len(self._history_matches)
        self._set_stat(self._lbl_total_matches, str(total))

        if self._history_sims:
            avg = sum(self._history_sims) / len(self._history_sims)
            self._set_stat(self._lbl_avg_sim, f"{avg:.3f}")

        active = sum(
            p.stats.active_count for p in self._processors if p and hasattr(p, "stats")
        )
        self._set_stat(self._lbl_active_tracks, str(active))

        db_size = 0
        try:
            reid_db = getattr(self, '_reid_db', None)
            if reid_db:
                db_size = reid_db.count()
        except Exception:
            pass
        self._set_stat(self._lbl_db_size, str(db_size))

        fps_vals = [p.stats.fps for p in self._processors if p and hasattr(p, "stats")]
        avg_fps = sum(fps_vals) / len(fps_vals) if fps_vals else 0
        self._set_stat(self._lbl_fps_avg, f"{avg_fps:.1f}")

        # Update FPS table
        self._fps_table.setRowCount(len(self._processors))
        for i, proc in enumerate(self._processors):
            if proc is None:
                continue
            try:
                cam = proc.cam
                s   = proc.stats
                self._fps_table.setItem(i, 0, QTableWidgetItem(cam.name))
                self._fps_table.setItem(i, 1, QTableWidgetItem(cam.role))
                self._fps_table.setItem(i, 2, QTableWidgetItem(f"{s.fps:.1f}"))
                self._fps_table.setItem(i, 3, QTableWidgetItem(str(s.active_count)))
            except Exception:
                pass

        if not self._pg_available:
            return

        # Update charts
        now = time.time()
        # Match timeline: bucket by minute, last 30 min
        N_BINS = 30
        bins = [0] * N_BINS
        for t in self._history_matches:
            age_min = (now - t) / 60.0
            idx = int(age_min)
            if 0 <= idx < N_BINS:
                bins[idx] += 1
        x = list(range(N_BINS))
        self._curve_timeline.setData(x=x, y=bins[::-1])

        # Similarity histogram
        if self._history_sims:
            import numpy as np
            arr = np.array(self._history_sims)
            counts, edges = np.histogram(arr, bins=20, range=(0.5, 1.0))
            centers = (edges[:-1] + edges[1:]) / 2
            self._bar_sim.setOpts(x=centers, height=counts, width=float(edges[1] - edges[0]) * 0.8)


class MatchesTab(QWidget):
    MAX_CARDS = 100

    def __init__(self, parent=None):
        super().__init__(parent)
        self._cards: List[MatchCard] = []
        self._current_filter = "all"
        self._storage: Optional["MatchStorage"] = None
        self._build_ui()

    def set_storage(self, storage: "MatchStorage"):
        self._storage = storage

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(10, 10, 10, 10)
        root.setSpacing(8)

        # Header
        hdr_row = QHBoxLayout()
        title = QLabel("Результаты сопоставления багажа")
        title.setObjectName("hdr")
        hdr_row.addWidget(title, 1)

        self._count_lbl = QLabel("0 совпадений")
        self._count_lbl.setStyleSheet("color:#6c7086;")
        hdr_row.addWidget(self._count_lbl)

        export_btn = QPushButton("📊  Экспорт CSV")
        export_btn.clicked.connect(self._export_csv)
        hdr_row.addWidget(export_btn)

        export_xlsx_btn = QPushButton("📗  Экспорт Excel")
        export_xlsx_btn.setToolTip("Экспорт всей истории совпадений из БД в .xlsx")
        export_xlsx_btn.clicked.connect(self._export_xlsx)
        hdr_row.addWidget(export_xlsx_btn)

        clear_btn = QPushButton("🗑  Очистить")
        clear_btn.clicked.connect(self.clear)
        hdr_row.addWidget(clear_btn)
        root.addLayout(hdr_row)

        # Filter buttons
        flt_row = QHBoxLayout()
        flt_row.setSpacing(6)
        flt_lbl = QLabel("Фильтр:")
        flt_lbl.setStyleSheet("color:#6c7086; font-size:12px;")
        flt_row.addWidget(flt_lbl)
        self._flt_btns: Dict[str, QPushButton] = {}
        _flt_specs = [
            ("all",      "Все",        ""),
            ("same",     "✔ Тот же",   "background:#a6e3a1; color:#1e1e2e; border-color:#a6e3a1;"),
            ("probable", "? Вероятно", "background:#f9e2af; color:#1e1e2e; border-color:#f9e2af;"),
            ("different","✘ Другой",   "background:#f38ba8; color:#1e1e2e; border-color:#f38ba8;"),
        ]
        for key, text, checked_style in _flt_specs:
            btn = QPushButton(text)
            btn.setCheckable(True)
            btn.setChecked(key == "all")
            btn.setFixedHeight(26)
            btn.setProperty("flt_key", key)
            btn.setProperty("checked_style", checked_style)
            btn.clicked.connect(lambda _, k=key: self._set_filter(k))
            self._flt_btns[key] = btn
            flt_row.addWidget(btn)
        flt_row.addStretch()
        root.addLayout(flt_row)

        # Mode banner (shown only in training mode)
        self._mode_banner = QLabel(
            "⚠  Переключитесь в  🔍 Рабочий режим  в настройках, "
            "чтобы видеть результаты сопоставления."
        )
        self._mode_banner.setStyleSheet(
            "background:#313244; color:#f9e2af; border-radius:6px; padding:8px 12px;"
        )
        self._mode_banner.setWordWrap(True)
        root.addWidget(self._mode_banner)

        # Scroll area with cards
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        inner = QWidget()
        self._cards_layout = QVBoxLayout(inner)
        self._cards_layout.setAlignment(Qt.AlignTop)
        self._cards_layout.setSpacing(6)
        scroll.setWidget(inner)
        root.addWidget(scroll, 1)

    def set_mode(self, mode: str):
        self._mode_banner.setVisible(mode != "production")

    def _set_filter(self, key: str):
        self._current_filter = key
        for k, btn in self._flt_btns.items():
            active = (k == key)
            btn.setChecked(active)
            cs = btn.property("checked_style")
            btn.setStyleSheet(cs if (active and cs) else "")
        for card in self._cards:
            card.setVisible(key == "all" or self._card_key(card) == key)

    @staticmethod
    def _card_key(card: "MatchCard") -> str:
        v = card.match_result.verdict
        if "Тот же"   in v: return "same"
        if "Вероятно" in v: return "probable"
        return "different"

    @pyqtSlot(object)
    def add_match(self, mr: MatchResult):
        card = MatchCard(mr)
        visible = (self._current_filter == "all" or
                   self._card_key(card) == self._current_filter)
        card.setVisible(visible)
        self._cards_layout.insertWidget(0, card)
        self._cards.insert(0, card)

        while len(self._cards) > self.MAX_CARDS:
            old = self._cards.pop()
            self._cards_layout.removeWidget(old)
            old.deleteLater()

        total   = len(self._cards)
        visible_n = sum(1 for c in self._cards if c.isVisible())
        self._count_lbl.setText(
            f"{visible_n}/{total} совпадений" if self._current_filter != "all"
            else f"{total} совпадений"
        )

    def _export_csv(self):
        if not self._cards:
            QMessageBox.information(self, "Экспорт", "Нет данных для экспорта.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить CSV", f"matches_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv",
            "CSV файлы (*.csv)"
        )
        if not path:
            return
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([
                "Время", "Вердикт", "Сходство %",
                "Источник (камера)", "Источник (стол)", "Источник (трек)",
                "Запрос (камера)",   "Запрос (стол)",   "Запрос (трек)",
            ])
            for card in self._cards:
                mr = card.match_result
                writer.writerow([
                    datetime.fromtimestamp(mr.timestamp).strftime("%Y-%m-%d %H:%M:%S"),
                    mr.verdict,
                    f"{mr.similarity * 100:.2f}",
                    mr.source_entry.cam_name,
                    mr.source_entry.counter_id,
                    mr.source_entry.track_id,
                    mr.query_cam_name,
                    mr.query_counter_id,
                    mr.query_track_id,
                ])
        QMessageBox.information(self, "Экспорт завершён", f"Сохранено: {path}")

    def _export_xlsx(self):
        if self._storage is None:
            QMessageBox.information(self, "Экспорт", "Хранилище не инициализировано.")
            return
        if self._storage.count() == 0:
            QMessageBox.information(self, "Экспорт", "Нет данных для экспорта.")
            return
        path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить Excel",
            f"matches_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel файлы (*.xlsx)"
        )
        if not path:
            return
        try:
            n = self._storage.export_xlsx(path)
            QMessageBox.information(self, "Экспорт завершён",
                                    f"Сохранено {n} записей:\n{path}")
        except RuntimeError as e:
            QMessageBox.critical(self, "Ошибка", str(e))
        except Exception as e:
            QMessageBox.critical(self, "Ошибка экспорта", str(e))

    def clear(self):
        for c in self._cards:
            self._cards_layout.removeWidget(c)
            c.deleteLater()
        self._cards.clear()
        self._current_filter = "all"
        for k, btn in self._flt_btns.items():
            btn.setChecked(k == "all")
            btn.setStyleSheet("")
        self._count_lbl.setText("0 совпадений")


# ── Log tab ────────────────────────────────────────────────────────────────────

class LogTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._autoscroll = True
        self._build_ui()
        self._setup_handler()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(8, 8, 8, 8)
        root.setSpacing(6)

        btn_row = QHBoxLayout()
        lbl = QLabel("Журнал событий")
        lbl.setObjectName("hdr")
        btn_row.addWidget(lbl)
        btn_row.addStretch()

        self._autoscroll_chk = QCheckBox("Авто-прокрутка")
        self._autoscroll_chk.setChecked(True)
        self._autoscroll_chk.toggled.connect(lambda v: setattr(self, "_autoscroll", v))
        btn_row.addWidget(self._autoscroll_chk)

        clear_btn = QPushButton("🗑  Очистить")
        clear_btn.clicked.connect(self._log_view.clear if hasattr(self, "_log_view") else lambda: None)
        save_btn  = QPushButton("💾  Сохранить")
        save_btn.clicked.connect(self._save_log)
        btn_row.addWidget(clear_btn)
        btn_row.addWidget(save_btn)
        root.addLayout(btn_row)

        self._log_view = QPlainTextEdit()
        self._log_view.setReadOnly(True)
        self._log_view.setMaximumBlockCount(5000)
        root.addWidget(self._log_view)

        # Fix clear button closure
        clear_btn.clicked.disconnect()
        clear_btn.clicked.connect(self._log_view.clear)

    def _setup_handler(self):
        class QtLogHandler(logging.Handler):
            def __init__(self, slot):
                super().__init__()
                self._slot = slot
            def emit(self, record):
                self._slot(self.format(record))

        handler = QtLogHandler(self.append)
        handler.setFormatter(logging.Formatter("%(asctime)s  %(levelname)-8s  %(message)s"))
        logging.getLogger("BaggageTracker").addHandler(handler)

    @pyqtSlot(str)
    def append(self, msg: str):
        self._log_view.appendPlainText(msg)
        if self._autoscroll:
            sb = self._log_view.verticalScrollBar()
            sb.setValue(sb.maximum())

    def _save_log(self):
        path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить журнал", f"log_{datetime.now():%Y%m%d_%H%M%S}.txt",
            "Текстовые файлы (*.txt);;Все файлы (*)"
        )
        if path:
            Path(path).write_text(self._log_view.toPlainText(), encoding="utf-8")
            QMessageBox.information(self, "Сохранено", f"Журнал сохранён:\n{path}")


# ── Network / ONVIF scanner dialog ────────────────────────────────────────────

_RTSP_PATHS = [
    "/stream1", "/stream2", "/h264", "/live", "/live/main",
    "/live/ch00_0", "/cam/realmonitor?channel=1&subtype=0",
    "/Streaming/Channels/101", "/video1", "/ch1/main/av_stream",
    "/11", "/12", "/axis-media/media.amp",
]

class _NetworkScanDialog(QDialog):
    camera_selected = pyqtSignal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Сканирование сети — поиск камер")
        self.setMinimumSize(620, 460)
        self._build()

    def _build(self):
        root = QVBoxLayout(self)
        root.setSpacing(10)

        info = QLabel(
            "Введите диапазон адресов и порт RTSP. "
            "Сканер проверит доступность порта и попробует типовые пути потока."
        )
        info.setWordWrap(True)
        info.setStyleSheet("color:#a6adc8; font-size:12px;")
        root.addWidget(info)

        form = QHBoxLayout()
        form.addWidget(QLabel("Подсеть (пример: 192.168.1):"))
        self._subnet_edit = QLineEdit()
        self._subnet_edit.setPlaceholderText("192.168.1")
        self._subnet_edit.setFixedWidth(160)
        form.addWidget(self._subnet_edit)
        form.addWidget(QLabel("  Порт RTSP:"))
        self._port_spin = QSpinBox()
        self._port_spin.setRange(1, 65535)
        self._port_spin.setValue(554)
        self._port_spin.setFixedWidth(80)
        form.addWidget(self._port_spin)
        form.addStretch()
        self._scan_btn = QPushButton("▶  Сканировать")
        self._scan_btn.setObjectName("btn_primary")
        self._scan_btn.clicked.connect(self._start_scan)
        form.addWidget(self._scan_btn)
        root.addLayout(form)

        self._progress = QLabel("Готов.")
        self._progress.setStyleSheet("color:#6c7086; font-size:11px;")
        root.addWidget(self._progress)

        self._result_list = QTableWidget(0, 2)
        self._result_list.setHorizontalHeaderLabels(["RTSP URL", "Статус"])
        self._result_list.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self._result_list.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        self._result_list.setEditTriggers(QTableWidget.NoEditTriggers)
        self._result_list.setSelectionBehavior(QTableWidget.SelectRows)
        root.addWidget(self._result_list, 1)

        btns = QHBoxLayout()
        add_btn = QPushButton("＋  Добавить выбранную камеру")
        add_btn.setObjectName("btn_success")
        add_btn.clicked.connect(self._add_selected)
        btns.addWidget(add_btn)
        btns.addStretch()
        close_btn = QPushButton("Закрыть")
        close_btn.clicked.connect(self.reject)
        btns.addWidget(close_btn)
        root.addLayout(btns)

    def _start_scan(self):
        import socket
        subnet = self._subnet_edit.text().strip()
        if not subnet:
            QMessageBox.warning(self, "Ошибка", "Введите подсеть (например: 192.168.1)")
            return
        port = self._port_spin.value()
        self._scan_btn.setEnabled(False)
        self._result_list.setRowCount(0)
        QApplication.processEvents()

        found = 0
        for last in range(1, 255):
            ip = f"{subnet}.{last}"
            self._progress.setText(f"Проверяю {ip}:{port} …")
            QApplication.processEvents()
            if self._check_port(ip, port, timeout=0.3):
                for path in _RTSP_PATHS:
                    url = f"rtsp://{ip}:{port}{path}"
                    ok = self._check_rtsp(url)
                    r = self._result_list.rowCount()
                    self._result_list.insertRow(r)
                    self._result_list.setItem(r, 0, QTableWidgetItem(url))
                    self._result_list.setItem(
                        r, 1, QTableWidgetItem("✔ Отвечает" if ok else "· Порт открыт"))
                    if ok:
                        self._result_list.item(r, 1).setForeground(QColor("#a6e3a1"))
                    found += 1
                    QApplication.processEvents()

        self._progress.setText(
            f"Готово. Найдено хостов: {found}." if found
            else "Не найдено. Проверьте подсеть и порт."
        )
        self._scan_btn.setEnabled(True)

    @staticmethod
    def _check_port(ip: str, port: int, timeout: float = 0.3) -> bool:
        import socket
        try:
            with socket.create_connection((ip, port), timeout=timeout):
                return True
        except OSError:
            return False

    @staticmethod
    def _check_rtsp(url: str) -> bool:
        try:
            cap = cv2.VideoCapture(url, cv2.CAP_FFMPEG)
            cap.set(cv2.CAP_PROP_BUFFERSIZE, 1)
            ok = cap.isOpened()
            cap.release()
            return ok
        except Exception:
            return False

    def _add_selected(self):
        rows = self._result_list.selectedItems()
        if not rows:
            QMessageBox.information(self, "Выбор", "Выберите строку в таблице.")
            return
        url = self._result_list.item(self._result_list.currentRow(), 0).text()
        self.camera_selected.emit(url)
        self.accept()


# ── Session report dialog ──────────────────────────────────────────────────────

class _SessionReportDialog(QDialog):
    def __init__(self, duration: float, matches: int, rows: list, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Итоги смены")
        self.setMinimumWidth(600)
        self._rows = rows
        self._duration = duration
        self._matches  = matches
        self._build(duration, matches, rows)

    def _build(self, duration, matches, rows):
        root = QVBoxLayout(self)
        root.setSpacing(12)

        h, m, s = int(duration//3600), int((duration%3600)//60), int(duration%60)
        summary = QLabel(
            f"Длительность смены:  {h:02d}:{m:02d}:{s:02d}   |   "
            f"Совпадений багажа: {matches}"
        )
        summary.setObjectName("hdr")
        root.addWidget(summary)

        if rows:
            tbl = QTableWidget(len(rows), len(rows[0]))
            tbl.setHorizontalHeaderLabels(list(rows[0].keys()))
            tbl.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            tbl.setEditTriggers(QTableWidget.NoEditTriggers)
            for r, row in enumerate(rows):
                for c, val in enumerate(row.values()):
                    tbl.setItem(r, c, QTableWidgetItem(str(val)))
            root.addWidget(tbl)

        btn_row = QHBoxLayout()
        export_btn = QPushButton("📊  Экспорт Excel")
        export_btn.setObjectName("btn_primary")
        export_btn.clicked.connect(self._export_excel)
        btn_row.addWidget(export_btn)
        btn_row.addStretch()
        close_btn = QPushButton("Закрыть")
        close_btn.clicked.connect(self.accept)
        btn_row.addWidget(close_btn)
        root.addLayout(btn_row)

    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            QMessageBox.critical(self, "Ошибка",
                "Установите openpyxl:\n  pip install openpyxl")
            return

        path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить отчёт",
            f"shift_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
            "Excel (*.xlsx)"
        )
        if not path:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Итоги смены"

        # Header info
        h, m, s = int(self._duration//3600), int((self._duration%3600)//60), int(self._duration%60)
        ws.append(["Airport Baggage Tracker — Итоги смены"])
        ws.append([f"Дата: {datetime.now().strftime('%d.%m.%Y')}",
                   f"Длительность: {h:02d}:{m:02d}:{s:02d}",
                   f"Совпадений: {self._matches}"])
        ws.append([])

        if self._rows:
            headers = list(self._rows[0].keys())
            ws.append(headers)
            hdr_row = ws.max_row
            hdr_fill = PatternFill("solid", fgColor="313244")
            hdr_font = Font(bold=True, color="89b4fa")
            for col, _ in enumerate(headers, 1):
                cell = ws.cell(hdr_row, col)
                cell.fill = hdr_fill
                cell.font = hdr_font
                cell.alignment = Alignment(horizontal="center")
            for row in self._rows:
                ws.append(list(row.values()))

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        wb.save(path)
        QMessageBox.information(self, "Экспорт", f"Отчёт сохранён:\n{path}")


# ── Main window ────────────────────────────────────────────────────────────────

class MainWindow(QMainWindow):
    def __init__(self, cfg_path: str):
        super().__init__()
        self._cfg_path  = cfg_path
        self._cfg       = load_config(cfg_path)
        self._processors: List[Optional[CameraProcessor]] = []
        self._running_all = False

        # Shared ReID objects (production mode)
        self._reid_db:        Optional[ReIDDatabase]         = None
        self._reid_extractor: Optional[ReIDFeatureExtractor] = None
        self._training_registry = None  # TrainingIdentityRegistry (training mode)

        # Shared batch inference engine (shared_yolo mode, 8+ cameras)
        self._batch_engine = None        # BatchInferenceEngine | None
        self._cam_slots: Dict[int, int] = {}   # camera list index → engine slot

        # Signal-loss alert tracking
        self._error_start: Dict[str, float] = {}
        self._alerted: set = set()

        # Session tracking
        self._session_start: Optional[float] = None
        self._session_matches: int = 0

        # Persistent match storage (Feature 3)
        self._match_storage = MatchStorage()

        # Missing-baggage alert tracking (Feature 6)
        self._matched_source_keys: set = set()   # (cam_name, track_id) that got a match
        self._alerted_missing: set = set()        # keys already alerted so we don't repeat

        # Web dashboard
        self._web = WebDashboard(port=self._cfg.web_port)
        self._web.start()

        self.setWindowTitle("✈  Airport Baggage Tracker")
        self.resize(1280, 800)
        self.setMinimumSize(900, 600)

        self._build_ui()
        self._matches_tab.set_storage(self._match_storage)
        self._sync_processors()
        self._refresh_cameras_tab()
        self._refresh_monitor()

        # Status bar timer
        self._sb_timer = QTimer(self)
        self._sb_timer.timeout.connect(self._update_statusbar)
        self._sb_timer.start(1000)

        # Alert timer — check camera health every 5 s
        self._alert_timer = QTimer(self)
        self._alert_timer.timeout.connect(self._check_alerts)
        self._alert_timer.start(5000)

        # Missing-baggage alert timer — check every 60 s (Feature 6)
        self._missing_timer = QTimer(self)
        self._missing_timer.timeout.connect(self._check_missing_baggage)
        self._missing_timer.start(60_000)

        # ReID autosave timer
        self._reid_autosave_timer = QTimer(self)
        self._reid_autosave_timer.timeout.connect(self._reid_autosave)

        # System tray icon
        self._tray: Optional[QSystemTrayIcon] = None
        self._setup_tray()

    # ── Tray icon ──────────────────────────────────────────────────────────────

    def _setup_tray(self):
        if not QSystemTrayIcon.isSystemTrayAvailable():
            return
        self._tray = QSystemTrayIcon(self)
        self._tray.setIcon(self.style().standardIcon(
            self.style().SP_ComputerIcon
        ))
        self._tray.setToolTip("Airport Baggage Tracker")
        tray_menu = QMenu(self)
        tray_menu.addAction("Показать", self.showNormal)
        tray_menu.addAction(f"Открыть дашборд ({self._web.url})",
                            lambda: self._open_web())
        tray_menu.addSeparator()
        tray_menu.addAction("Выход", self.close)
        self._tray.setContextMenu(tray_menu)
        self._tray.activated.connect(
            lambda r: self.showNormal() if r == QSystemTrayIcon.DoubleClick else None
        )
        self._tray.show()

    def _open_web(self):
        import webbrowser
        webbrowser.open(self._web.url)

    # ── Config Profiles ───────────────────────────────────────────────────────

    def _profiles_dir(self) -> Path:
        d = app_dir() / "profiles"
        d.mkdir(exist_ok=True)
        return d

    def _refresh_profiles(self):
        self._profile_combo.blockSignals(True)
        self._profile_combo.clear()
        self._profile_combo.addItem("— выбрать профиль —")
        for p in sorted(self._profiles_dir().glob("*.yaml")):
            self._profile_combo.addItem(p.stem)
        self._profile_combo.blockSignals(False)

    def _load_profile(self, name: str):
        if name.startswith("—"):
            return
        path = self._profiles_dir() / f"{name}.yaml"
        if not path.exists():
            return
        try:
            new_cfg = load_config(str(path))
            # Preserve cameras from current config
            new_cfg.cameras = self._cfg.cameras
            self._cfg.__dict__.update(new_cfg.__dict__)
            self._cfg.cameras = new_cfg.cameras if new_cfg.cameras else self._cfg.cameras
        except Exception as e:
            QMessageBox.critical(self, "Ошибка загрузки профиля", str(e))
            return
        # Reload all UI tabs
        try:
            self._settings_tab._load()
        except Exception:
            pass
        try:
            self._reid_tab._load()
        except Exception:
            pass
        try:
            self._devices_tab._load()
        except Exception:
            pass
        self._statusbar.showMessage(f"✔  Профиль «{name}» загружен")

    def _save_new_profile(self):
        name, ok = QInputDialog.getText(self, "Сохранить профиль", "Имя профиля:")
        if not ok or not name.strip():
            return
        name = name.strip().replace(" ", "_")
        path = self._profiles_dir() / f"{name}.yaml"
        try:
            save_config(self._cfg, str(path))
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", str(e))
            return
        self._refresh_profiles()
        idx = self._profile_combo.findText(name)
        if idx >= 0:
            self._profile_combo.blockSignals(True)
            self._profile_combo.setCurrentIndex(idx)
            self._profile_combo.blockSignals(False)
        self._statusbar.showMessage(f"✔  Профиль «{name}» сохранён")

    def _delete_profile(self):
        name = self._profile_combo.currentText()
        if name.startswith("—"):
            return
        path = self._profiles_dir() / f"{name}.yaml"
        if QMessageBox.question(self, "Удалить профиль",
                                f"Удалить профиль «{name}»?",
                                QMessageBox.Yes | QMessageBox.No) == QMessageBox.Yes:
            path.unlink(missing_ok=True)
            self._refresh_profiles()
            self._statusbar.showMessage(f"✔  Профиль «{name}» удалён")

    # ── Session management ────────────────────────────────────────────────────

    def _toggle_session(self):
        if self._session_start is None:
            self._session_start    = time.time()
            self._session_matches  = 0
            self._session_act.setText("■  Завершить смену")
            # Reset missing-baggage tracking for new shift
            self._matched_source_keys.clear()
            self._alerted_missing.clear()
            # Clear persistent DB for new session (optional — keeps only current shift)
            self._match_storage.clear_session()
            logger.info("Смена начата в %s", datetime.now().strftime("%H:%M:%S"))
            if self._tray:
                self._tray.showMessage("Смена начата",
                    datetime.now().strftime("Начало: %H:%M:%S"),
                    QSystemTrayIcon.Information, 3000)
        else:
            self._end_session()

    def _scan_network(self):
        dlg = _NetworkScanDialog(self)
        dlg.camera_selected.connect(self._add_camera_from_url)
        dlg.exec_()

    def _next_cam_id(self) -> int:
        used = {c.cam_id for c in self._cfg.cameras}
        i = 1
        while i in used:
            i += 1
        return i

    def _add_camera_from_url(self, url: str):
        from tracker_core import CameraEntry
        entry = CameraEntry(
            name="Camera " + str(len(self._cfg.cameras) + 1),
            cam_id=self._next_cam_id(),
            mode="rtsp", rtsp_url=url,
        )
        dlg = AddEditCameraDialog(cam=entry, all_cameras=self._cfg.cameras, parent=self)
        if dlg.exec_() == QDialog.Accepted:
            self._cfg.cameras.append(dlg.result_entry())
            self._processors.append(None)
            self._save_config()
            self._refresh_all()

    def _end_session(self):
        if self._session_start is None:
            return
        duration = time.time() - self._session_start
        self._session_start = None
        self._session_act.setText("▶  Начать смену")

        # Собираем статистику
        stats_rows = []
        for cam, proc in zip(self._cfg.cameras, self._processors):
            if proc:
                s = proc.stats
                stats_rows.append({
                    "Камера": cam.name, "Стол": cam.counter_id,
                    "Роль": cam.role, "Активно": s.active_count,
                    "Треков за смену": s.total_seen,
                    "YOLO сохранено": s.yolo_saved,
                    "ReID сохранено": s.reid_saved,
                    "FPS": round(s.fps, 1),
                })

        dlg = _SessionReportDialog(
            duration, self._session_matches, stats_rows, self
        )
        dlg.exec_()

    # ── Camera health alerts ───────────────────────────────────────────────────

    @pyqtSlot(str, object)
    def _on_cam_stats(self, cam_name: str, stats: ProcessorStats):
        if stats.status == "error":
            if cam_name not in self._error_start:
                self._error_start[cam_name] = time.time()
        else:
            self._error_start.pop(cam_name, None)
            self._alerted.discard(cam_name)

    def _check_alerts(self):
        now = time.time()
        for cam_name, start in list(self._error_start.items()):
            if now - start > 30 and cam_name not in self._alerted:
                self._alerted.add(cam_name)
                msg = f"Камера «{cam_name}» не отвечает более 30 секунд"
                logger.warning(msg)
                if self._tray:
                    self._tray.showMessage(
                        "⚠  Потеря сигнала камеры", msg,
                        QSystemTrayIcon.Warning, 6000
                    )

    @pyqtSlot()
    def _check_missing_baggage(self):
        """Feature 6: alert when a source-side track was not matched within 60 % of TTL."""
        if self._reid_db is None or self._cfg.app_mode != "production":
            return
        alert_age = self._cfg.reid_ttl_minutes * 60 * 0.6
        try:
            old_entries = self._reid_db.get_old_entries(alert_age)
        except Exception:
            return
        for entry in old_entries:
            key = (entry.cam_name, entry.track_id)
            if key in self._matched_source_keys or key in self._alerted_missing:
                continue
            self._alerted_missing.add(key)
            mins = alert_age / 60
            msg = (
                f"Трек #{entry.track_id} со стойки #{entry.counter_id} "
                f"({entry.cam_name}) не появился на сортировке "
                f"уже {mins:.0f} мин."
            )
            logger.warning("⚠ Возможно потерянный багаж: %s", msg)
            if self._tray:
                self._tray.showMessage(
                    "⚠  Возможно потерянный багаж", msg,
                    QSystemTrayIcon.Warning, 8000
                )

    # ── UI construction ────────────────────────────────────────────────────────

    def _build_ui(self):
        self._create_toolbar()

        self._tabs = QTabWidget()
        self._tabs.setDocumentMode(True)
        self.setCentralWidget(self._tabs)

        self._monitor_tab   = MonitorTab()
        self._tracking_tab  = TrackingTab()
        self._cameras_tab   = CamerasTab()
        self._settings_tab  = SettingsTab(self._cfg)
        self._devices_tab   = DevicesTab(self._cfg)
        self._reid_tab      = ReIDTab(self._cfg)
        self._training_tab  = TrainingTab(self._cfg)
        self._stats_tab     = StatisticsTab()
        self._analytics_tab = AnalyticsTab()
        self._matches_tab   = MatchesTab()
        self._log_tab       = LogTab()

        self._tabs.addTab(self._monitor_tab,   "📷  Мониторинг")
        self._tabs.addTab(self._tracking_tab,  "🎯  Треки")
        self._tabs.addTab(self._cameras_tab,   "🎛  Камеры")
        self._tabs.addTab(self._settings_tab,  "⚙  Настройки")
        self._tabs.addTab(self._devices_tab,   "🖥  Устройства")
        self._tabs.addTab(self._reid_tab,      "🧠  ReID")
        self._tabs.addTab(self._training_tab,  "🎓  Обучение")
        self._tabs.addTab(self._stats_tab,     "📊  Статистика")
        self._tabs.addTab(self._analytics_tab, "📈  Аналитика")
        self._tabs.addTab(self._matches_tab,   "🔍  Совпадения")
        self._tabs.addTab(self._log_tab,       "📋  Журнал")

        # Sync matches tab banner with current mode
        self._matches_tab.set_mode(self._cfg.app_mode)

        # Settings → notify mode change
        self._settings_tab.settings_saved.connect(self._on_settings_saved)
        # Devices tab → same handler
        self._devices_tab.settings_saved.connect(self._on_settings_saved)
        # ReID tab → same handler
        self._reid_tab.settings_saved.connect(self._on_settings_saved)
        # ReID DB action
        self._reid_tab.db_action.connect(self._on_reid_db_action)

        # Camera reorder from monitor drag-and-drop
        self._monitor_tab.cameras_reordered.connect(self._on_cameras_reordered)
        # ROI changed from monitor tile
        self._monitor_tab.roi_changed.connect(self._on_roi_changed)

        # Wire camera-tab buttons
        ct = self._cameras_tab
        ct._add_btn.clicked.connect(self._add_camera)
        ct._edit_btn.clicked.connect(self._edit_camera)
        ct._del_btn.clicked.connect(self._delete_camera)
        ct._start_btn.clicked.connect(self._start_selected)
        ct._stop_btn.clicked.connect(self._stop_selected)
        ct._test_btn.clicked.connect(self._test_selected)
        ct._scan_btn.clicked.connect(self._scan_network)

        self._statusbar = QStatusBar()
        self.setStatusBar(self._statusbar)
        self._statusbar.showMessage("Готов к работе")

    def _create_toolbar(self):
        tb = QToolBar("Главная панель")
        tb.setMovable(False)
        tb.setIconSize(QSize(16, 16))
        self.addToolBar(tb)

        def _act(text, tip, slot):
            a = QAction(text, self)
            a.setToolTip(tip)
            a.triggered.connect(slot)
            tb.addAction(a)
            return a

        _act("＋  Добавить камеру",  "Добавить новую камеру",  self._add_camera)
        _act("✖  Удалить камеру",   "Удалить выбранную камеру (из вкладки Камеры)", self._delete_camera)
        tb.addSeparator()
        _act("▶  Запустить все",    "Запустить все камеры",   self._start_all)
        _act("■  Остановить все",   "Остановить все камеры",  self._stop_all)
        tb.addSeparator()
        _act("💾  Сохранить конфиг", "Сохранить config.yaml",  self._save_config)
        _act("📂  Открыть датасет",  "Открыть папку Dataset",  lambda: self._stats_tab._open_dataset())
        tb.addSeparator()
        _act("📷  Снимок всех камер","Сохранить снимки всех камер", self._snapshot_all)
        tb.addSeparator()
        self._session_act = _act("▶  Начать смену", "Начать смену (сброс счётчиков + запись времени)",
                                 self._toggle_session)
        tb.addSeparator()
        _act(f"🌐  Веб-дашборд", f"Открыть дашборд в браузере ({self._web.url})", self._open_web)
        tb.addSeparator()
        tb.addWidget(QLabel("  Профиль: "))
        self._profile_combo = QComboBox()
        self._profile_combo.setFixedWidth(160)
        self._profile_combo.setToolTip("Профили конфигурации — сохранённые наборы настроек")
        self._profile_combo.currentTextChanged.connect(self._load_profile)
        tb.addWidget(self._profile_combo)
        save_profile_btn = QPushButton("💾")
        save_profile_btn.setFixedSize(28, 28)
        save_profile_btn.setToolTip("Сохранить текущие настройки как новый профиль")
        save_profile_btn.clicked.connect(self._save_new_profile)
        tb.addWidget(save_profile_btn)
        del_profile_btn = QPushButton("✖")
        del_profile_btn.setFixedSize(28, 28)
        del_profile_btn.setToolTip("Удалить текущий профиль")
        del_profile_btn.clicked.connect(self._delete_profile)
        tb.addWidget(del_profile_btn)
        self._refresh_profiles()

    # ── Processor management ───────────────────────────────────────────────────

    def _sync_processors(self):
        """Ensure _processors list length matches cameras list."""
        n = len(self._cfg.cameras)
        while len(self._processors) < n:
            self._processors.append(None)
        while len(self._processors) > n:
            p = self._processors.pop()
            if p and p.isRunning():
                p.stop(); p.wait(3000)

    def _ensure_reid_objects(self):
        """Создаём/пересоздаём shared ReID DB и extractor при необходимости."""
        if self._cfg.app_mode == "training":
            if self._training_registry is None:
                from tracker_core import TrainingIdentityRegistry
                self._training_registry = TrainingIdentityRegistry(
                    link_timeout=self._cfg.training_link_timeout
                )
            return
        if self._cfg.app_mode != "production":
            return
        ttl = self._cfg.reid_ttl_minutes * 60
        if self._reid_db is None or abs(self._reid_db.ttl - ttl) > 1:
            self._reid_db = ReIDDatabase(ttl_seconds=ttl)
            logger.info("ReID DB создана (TTL %.0f мин)", self._cfg.reid_ttl_minutes)
        if self._reid_extractor is None:
            self._load_reid_extractor()
        self._matches_tab.set_mode("production")

    def _load_reid_extractor(self) -> None:
        """
        Загружает ReIDFeatureExtractor синхронно.
        При первом запуске torchreid скачивает веса (~6 МБ) через gdown.
        Прогресс виден в консоли/терминале — GUI обновляет статус бар.
        """
        from tracker_core import resolve_device

        device   = resolve_device(self._cfg, component="reid")
        engine   = self._cfg.reid_engine
        use_half = self._cfg.reid_half

        logger.info("ReID: инициализация OSNet x1.0  engine=%s  device=%s", engine, device)
        self._reid_tab.update_status("⏳  Загрузка OSNet…")
        if self._statusbar:
            self._statusbar.showMessage("⏳  Загрузка OSNet ReID — следите за консолью…")
        QApplication.processEvents()

        try:
            logger.debug("ReID: создаём ReIDFeatureExtractor...")
            self._reid_extractor = ReIDFeatureExtractor(
                device=device, engine=engine, use_half=use_half,
                onnx_model_path=self._cfg.reid_model_path,
            )
        except Exception as exc:
            logger.error("ReID: ошибка загрузки OSNet: %s", exc, exc_info=True)
            self._reid_tab.update_status(f"✘  Ошибка: {exc}")
            if self._statusbar:
                self._statusbar.showMessage(f"✘  OSNet не загружен: {exc}")
            QMessageBox.critical(
                self, "Ошибка загрузки ReID",
                f"Не удалось инициализировать OSNet:\n\n{exc}\n\n"
                "Запустите  setup_models.py  для предзагрузки весов:\n"
                "  python setup_models.py"
            )
            return

        actual_engine = self._reid_extractor.effective_engine

        # Проверяем fallback: запросили ONNX, получили PyTorch
        if self._reid_extractor.fallback_occurred:
            reason = self._reid_extractor.fallback_reason or ""
            logger.warning("ReID: fallback с %s на PyTorch. Причина: %s", engine, reason)

            # Определяем направление fallback: pytorch→onnx или onnx→pytorch
            pytorch_to_onnx = actual_engine in ("onnx_gpu", "onnx_cpu")

            if pytorch_to_onnx:
                # torchreid недоступен, но ONNX работает — это нормальная ситуация
                hint = (
                    f"torchreid не установлен (Python 3.12+ несовместим).\n"
                    f"Приложение автоматически использует {actual_engine.upper()} — "
                    "производительность такая же или лучше.\n\n"
                    "Это предупреждение можно убрать, выбрав ONNX движок в «Устройства»."
                )
                status = (
                    f"✔  OSNet x1.0  |  движок: {actual_engine}  |  устройство: {device}"
                    + ("  [FP16]" if use_half else "")
                    + "  (автопереключение с pytorch)"
                )
            else:
                # onnx→pytorch fallback: ONNX не смог загрузиться
                if "DLL" in reason or "WinError" in reason or "_pybind_state" in reason or "OSError" in reason:
                    hint = (
                        "Конфликт DLL: onnxruntime и PyTorch CUDA.\n\n"
                        "✅  Решение:\n"
                        "1. Закройте приложение\n"
                        "2. Выполните в терминале:\n"
                        "     pip uninstall onnxruntime-gpu onnxruntime -y\n"
                        "     pip install onnxruntime\n"
                        "3. Запустите приложение снова\n"
                        "4. В «Устройства» выберите  ONNX Runtime — CPU"
                    )
                elif "onnxruntime не установлен" in reason or "No module named 'onnxruntime'" in reason:
                    hint = (
                        "onnxruntime не установлен.\n\n"
                        "✅  Как установить:\n"
                        "  GPU:  pip install onnxruntime-gpu\n"
                        "  CPU:  pip install onnxruntime\n\n"
                        "После установки перезапустите приложение."
                    )
                else:
                    hint = (
                        f"Причина: {reason}\n\n"
                        "После устранения перезапустите приложение."
                    )
                status = (
                    f"⚠  OSNet x1.0  |  ONNX недоступен → PyTorch  |  устройство: {device}"
                    + ("  [FP16]" if use_half else "")
                )

            self._reid_tab.update_status(status)
            if self._statusbar:
                self._statusbar.showMessage(
                    f"⚠  ONNX ({engine}) недоступен — используется {actual_engine}. "
                    "Откройте «Устройства» для деталей.", 15000
                )

            dlg_title = (
                "PyTorch → ONNX (автопереключение)" if pytorch_to_onnx
                else "ONNX недоступен — используется PyTorch"
            )
            QMessageBox.warning(
                self, dlg_title,
                f"Запрошен движок: {engine}\n"
                f"Фактически используется: {actual_engine}\n\n"
                + hint
            )
        else:
            status = (
                f"✔  OSNet x1.0  |  движок: {actual_engine}  |  устройство: {device}"
                + ("  [FP16]" if use_half else "")
            )
            logger.info("ReID: %s", status)
            self._reid_tab.update_status(status)
            if self._statusbar:
                self._statusbar.showMessage(status, 5000)

    def _ensure_batch_engine(self, _skip_restart: bool = False):
        """
        Создаём/пересоздаём BatchInferenceEngine при необходимости.

        Движок пересоздаётся если:
        - включён shared_yolo и движок ещё не создан
        - количество слотов не совпадает с числом камер (добавили/удалили камеру)

        При пересоздании работающие процессоры перезапускаются автоматически,
        иначе они остаются с ссылкой на остановленный старый движок.
        """
        n_cams = len(self._cfg.cameras)

        if not self._cfg.shared_yolo:
            self._destroy_batch_engine()
            return

        need_rebuild = (
            self._batch_engine is None
            or self._batch_engine.n_slots != n_cams
        )
        if not need_rebuild:
            return

        # Запомним, какие процессоры сейчас работают — нужно перезапустить
        # их с новым движком (_skip_restart=True чтобы не рекурсить)
        restart_idxs = [] if _skip_restart else [
            i for i, p in enumerate(self._processors)
            if p and p.isRunning() and i < n_cams
        ]
        for i in restart_idxs:
            self._stop_processor(i)

        self._destroy_batch_engine()

        from tracker_core import resolve_device
        device = resolve_device(self._cfg)
        engine = BatchInferenceEngine(
            model_path      = self._cfg.model_path,
            tracking_config = self._cfg.tracking_config,
            confidence      = self._cfg.confidence,
            iou             = self._cfg.iou,
            classes         = self._cfg.classes,
            device          = device,
            half            = self._cfg.half,
        )
        self._cam_slots = {}
        for i in range(n_cams):
            self._cam_slots[i] = engine.register()

        if engine.start():
            self._batch_engine = engine
            logger.info(
                "BatchEngine создан: %d камер, device=%s", n_cams, device
            )
            # Перезапускаем процессоры с новым движком
            for i in restart_idxs:
                if i < len(self._cfg.cameras):
                    self._start_processor(i, _skip_engine_rebuild=True)
        else:
            self._batch_engine = None
            self._cam_slots    = {}
            logger.error("BatchEngine: не удалось запустить движок")

    def _destroy_batch_engine(self):
        if self._batch_engine is not None:
            self._batch_engine.stop()
            self._batch_engine = None
        self._cam_slots = {}

    def _start_processor(self, idx: int, _skip_engine_rebuild: bool = False):
        self._stop_processor(idx)
        self._ensure_reid_objects()
        self._ensure_batch_engine(_skip_restart=_skip_engine_rebuild)
        cam    = self._cfg.cameras[idx]
        engine = self._batch_engine
        slot   = self._cam_slots.get(idx, 0) if engine else 0
        proc = CameraProcessor(
            cam, self._cfg,
            reid_db           = self._reid_db        if self._cfg.app_mode == "production" else None,
            reid_extractor    = self._reid_extractor if self._cfg.app_mode == "production" else None,
            engine            = engine,
            slot              = slot,
            training_registry = self._training_registry if self._cfg.app_mode == "training" else None,
        )
        proc.log_msg.connect(self._log_tab.append)
        proc.match_found.connect(self._matches_tab.add_match)
        proc.match_found.connect(self._on_match_found)
        # Monitor stats for signal-loss alerts
        cam_name = cam.name
        proc.stats_updated.connect(
            lambda stats, n=cam_name: self._on_cam_stats(n, stats)
        )
        self._processors[idx] = proc
        proc.start()

    @pyqtSlot(object)
    def _on_match_found(self, mr: MatchResult):
        """Flash the Совпадения tab and push event to web dashboard."""
        self._tracking_tab.on_match(mr)
        tab_idx = self._tabs.indexOf(self._matches_tab)
        self._tabs.setTabText(tab_idx, "🔍  Совпадения  ●")
        QTimer.singleShot(3000, lambda: self._tabs.setTabText(tab_idx, "🔍  Совпадения"))
        if self._session_start is not None:
            self._session_matches += 1
        # Feature 3: persist to SQLite
        self._match_storage.save(mr)
        # Feature 6: mark source track as matched (no missing-baggage alert for it)
        self._matched_source_keys.add((mr.source_entry.cam_name, mr.source_entry.track_id))
        # Push to web dashboard
        self._web.push_match(mr)
        # Analytics tab
        self._analytics_tab.add_match(mr)
        # Tray balloon for confirmed matches
        if self._tray and "Тот же" in mr.verdict:
            self._tray.showMessage(
                "✔  Багаж совпал",
                f"{mr.source_entry.cam_name} → {mr.query_cam_name}  "
                f"({mr.similarity * 100:.1f}%)",
                QSystemTrayIcon.Information, 4000
            )

    @pyqtSlot(object)
    def _on_roi_changed(self, _cam: CameraEntry):
        self._save_config()

    @pyqtSlot(list)
    def _on_cameras_reordered(self, new_cams: list):
        """Sync cfg.cameras and processors after monitor drag-and-drop swap."""
        cam_to_proc = {id(c): p for c, p in zip(self._cfg.cameras, self._processors)}
        self._cfg.cameras = list(new_cams)
        self._processors  = [cam_to_proc.get(id(c)) for c in new_cams]
        # Batch engine uses fixed slot per list-index — rebuild so tracker
        # states stay aligned with the new camera order.
        if self._batch_engine is not None:
            self._destroy_batch_engine()
        self._save_config()
        self._refresh_cameras_tab()

    def _stop_processor(self, idx: int):
        p = self._processors[idx] if idx < len(self._processors) else None
        if p and p.isRunning():
            p.stop()
            p.wait(5000)
        if idx < len(self._processors):
            self._processors[idx] = None

    # ── Camera CRUD actions ────────────────────────────────────────────────────

    def _add_camera(self):
        from tracker_core import CameraEntry
        new_entry = CameraEntry(
            name="Camera " + str(len(self._cfg.cameras) + 1),
            cam_id=self._next_cam_id(),
        )
        dlg = AddEditCameraDialog(cam=new_entry, all_cameras=self._cfg.cameras, parent=self)
        if dlg.exec_() != QDialog.Accepted:
            return
        entry = dlg.result_entry()
        self._cfg.cameras.append(entry)
        self._processors.append(None)
        self._save_config()
        self._refresh_all()

        # Ask to start immediately
        if QMessageBox.question(
            self, "Запустить?",
            f"Камера «{entry.name}» добавлена. Запустить её сейчас?",
            QMessageBox.Yes | QMessageBox.No
        ) == QMessageBox.Yes:
            self._start_processor(len(self._cfg.cameras) - 1)
            self._refresh_all()

    def _edit_camera(self):
        row = self._cameras_tab.selected_row()
        if row < 0:
            QMessageBox.information(self, "Выбор", "Выберите камеру в таблице.")
            return
        dlg = AddEditCameraDialog(
            cam=self._cfg.cameras[row],
            all_cameras=self._cfg.cameras,
            parent=self,
        )
        if dlg.exec_() != QDialog.Accepted:
            return
        self._cfg.cameras[row] = dlg.result_entry()
        self._save_config()
        self._refresh_all()

    def _delete_camera(self):
        row = self._cameras_tab.selected_row()
        if row < 0:
            if not self._cfg.cameras:
                QMessageBox.information(self, "Удаление", "Нет камер для удаления.")
                return
            row = len(self._cfg.cameras) - 1  # delete last if none selected

        name = self._cfg.cameras[row].name
        if QMessageBox.question(
            self, "Удалить камеру",
            f"Удалить камеру «{name}»?",
            QMessageBox.Yes | QMessageBox.No
        ) != QMessageBox.Yes:
            return

        self._stop_processor(row)
        del self._cfg.cameras[row]
        del self._processors[row]
        self._save_config()
        self._refresh_all()

    def _start_selected(self):
        row = self._cameras_tab.selected_row()
        if row < 0:
            QMessageBox.information(self, "Выбор", "Выберите камеру в таблице.")
            return
        self._start_processor(row)
        self._refresh_all()

    def _stop_selected(self):
        row = self._cameras_tab.selected_row()
        if row < 0:
            QMessageBox.information(self, "Выбор", "Выберите камеру в таблице.")
            return
        self._stop_processor(row)
        self._refresh_all()

    def _test_selected(self):
        row = self._cameras_tab.selected_row()
        if row < 0:
            QMessageBox.information(self, "Выбор", "Выберите камеру в таблице.")
            return
        cam = self._cfg.cameras[row]
        src = cam.source
        if not src:
            QMessageBox.warning(self, "Ошибка", "Источник не задан.")
            return

        pb = QMessageBox(self)
        pb.setWindowTitle("Проверка подключения")
        pb.setText(f"Подключение к:\n{src}\n\nПожалуйста, подождите…")
        pb.setStandardButtons(QMessageBox.NoButton)
        pb.show()
        QApplication.processEvents()

        backend = cv2.CAP_FFMPEG if cam.mode == "rtsp" else cv2.CAP_ANY
        cap = cv2.VideoCapture(src, backend)
        ok  = cap.isOpened()
        if ok:
            ok2, _ = cap.read()
            ok = ok and ok2
        cap.release()
        pb.close()

        if ok:
            QMessageBox.information(self, "Успех", f"✔  Камера «{cam.name}» отвечает!")
        else:
            QMessageBox.critical(self, "Ошибка", f"✘  Не удалось подключиться к «{cam.name}».")

    @pyqtSlot()
    def _on_settings_saved(self):
        """Вызывается после сохранения настроек — синхронизируем UI и сбрасываем объекты ReID."""
        # Предупреждение если камеры запущены — настройки применятся после перезапуска
        running = [p for p in self._processors if p and p.isRunning()]
        if running:
            names = ", ".join(p.cam.name for p in running[:3])
            if len(running) > 3:
                names += f" и ещё {len(running) - 3}"
            QMessageBox.warning(
                self, "Камеры запущены",
                f"Настройки сохранены, но следующие камеры нужно перезапустить вручную:\n"
                f"{names}\n\n"
                "Нажмите «Остановить все» → «Запустить все» чтобы применить изменения."
            )
        self._matches_tab.set_mode(self._cfg.app_mode)
        # Сброс: при смене режима или device нужно пересоздать extractor/DB/engine
        # Только если нет запущенных камер — иначе они всё ещё используют старый extractor/DB
        if not running:
            self._reid_db        = None
            self._reid_extractor = None
            self._training_registry = None
        self._reid_tab.update_status("Загрузка при запуске камеры…")
        self._destroy_batch_engine()
        self._web.reset()
        # Обновить заголовок окна
        mode_str = "🔍 РАБОТА" if self._cfg.app_mode == "production" else "🎓 ОБУЧЕНИЕ"
        self.setWindowTitle(f"✈  Airport Baggage Tracker  [{mode_str}]")
        # Restart autosave timer — берём из cfg (уже сохранено в _save())
        interval = self._cfg.reid_autosave_interval
        if interval > 0:
            self._reid_autosave_timer.start(interval * 1000)
        else:
            self._reid_autosave_timer.stop()

    def _on_reid_db_action(self, action: str, path: str):
        if self._reid_db is None:
            QMessageBox.warning(self, "БД не инициализирована", "Сначала запустите хотя бы одну камеру в режиме production.")
            return
        try:
            if action == "save":
                self._reid_db.save_to_file(path)
                self._statusbar.showMessage(f"✔  БД ReID сохранена: {path}")
            elif action == "load":
                self._reid_db.load_from_file(path)
                n = len(self._reid_db._entries)
                self._reid_tab.update_db_size(n)
                self._statusbar.showMessage(f"✔  БД ReID загружена: {n} записей")
            elif action == "clear":
                self._reid_db.clear()
                self._reid_tab.update_db_size(0)
                self._statusbar.showMessage("✔  БД ReID очищена")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка БД", str(e))

    def _reid_autosave(self):
        if self._reid_db is None:
            return
        cfg = self._cfg
        path = getattr(cfg, 'reid_db_path', '') or str(app_dir() / "reid_db.json")
        try:
            self._reid_db.save_to_file(path)
            n = len(self._reid_db._entries)
            self._reid_tab.update_db_size(n)
        except Exception as e:
            logger.warning("Автосохранение БД: %s", e)

    def _start_all(self):
        for i in range(len(self._cfg.cameras)):
            if self._cfg.cameras[i].enabled:
                self._start_processor(i)
        self._refresh_all()

    def _stop_all(self):
        for i in range(len(self._cfg.cameras)):
            self._stop_processor(i)
        # Очищаем ReID базу при полной остановке
        if self._reid_db is not None:
            self._reid_db.clear()
        self._refresh_all()

    # ── Snapshot all ───────────────────────────────────────────────────────────

    def _snapshot_all(self):
        if not self._monitor_tab._widgets:
            QMessageBox.information(self, "Снимок", "Нет активных камер.")
            return
        SNAPSHOT_DIR.mkdir(exist_ok=True)
        saved = []
        for fw in self._monitor_tab._widgets:
            if fw._last_raw_frame is not None:
                ts   = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
                path = SNAPSHOT_DIR / f"{fw.cam.name.replace(' ','_')}_{ts}.jpg"
                cv2.imwrite(str(path), fw._last_raw_frame, [cv2.IMWRITE_JPEG_QUALITY, 95])
                saved.append(str(path))
        if saved:
            QMessageBox.information(
                self, "Снимки сохранены",
                "Сохранено:\n" + "\n".join(saved)
            )
        else:
            QMessageBox.information(self, "Снимок", "Нет кадров для сохранения.")

    # ── Refresh helpers ────────────────────────────────────────────────────────

    def _refresh_monitor(self):
        self._monitor_tab.rebuild(self._cfg.cameras, self._processors)
        self._tracking_tab.rebuild(self._cfg.cameras, self._processors)

    def _refresh_cameras_tab(self):
        self._cameras_tab.populate(self._cfg.cameras, self._processors)

    def _refresh_stats(self):
        self._stats_tab.set_data(self._cfg.cameras, self._processors, self._match_storage)
        self._analytics_tab.set_data(self._processors, self._match_storage, self._reid_db)

    def _refresh_all(self):
        self._refresh_monitor()
        self._refresh_cameras_tab()
        self._refresh_stats()

    def _save_config(self):
        try:
            save_config(self._cfg, self._cfg_path)
            logger.info("Config saved to %s", self._cfg_path)
        except Exception as e:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить конфиг:\n{e}")

    # ── Status bar ─────────────────────────────────────────────────────────────

    @pyqtSlot()
    def _update_statusbar(self):
        running = sum(1 for p in self._processors if p and p.isRunning())
        active  = sum(
            p.stats.active_count for p in self._processors
            if p and p.isRunning()
        )
        total_yolo = sum(
            p.stats.yolo_saved for p in self._processors
            if p and p.isRunning()
        )
        mode_icon = "🔍 РАБОТА" if self._cfg.app_mode == "production" else "🎓 ОБУЧЕНИЕ"
        db_part   = (f"  │  БД: {self._reid_db.count()} записей"
                     if self._reid_db and self._cfg.app_mode == "production" else "")
        yolo_part = (f"  │  YOLO: {total_yolo} кадров"
                     if self._cfg.app_mode == "training" else "")
        self._statusbar.showMessage(
            f"{mode_icon}  │  "
            f"Камер: {running}/{len(self._cfg.cameras)}  │  "
            f"Предметов: {active}"
            f"{yolo_part}{db_part}  │  "
            f"{datetime.now():%H:%M:%S}"
        )
        self._refresh_cameras_tab()

    # ── Close event ────────────────────────────────────────────────────────────

    def closeEvent(self, event):
        reply = QMessageBox.question(
            self, "Выход",
            "Остановить все камеры и выйти?",
            QMessageBox.Yes | QMessageBox.No,
        )
        if reply != QMessageBox.Yes:
            event.ignore()
            return

        self._sb_timer.stop()
        self._alert_timer.stop()
        self._missing_timer.stop()
        self._stats_tab._timer.stop()
        self._web.stop()
        if self._tray:
            self._tray.hide()

        for i in range(len(self._processors)):
            p = self._processors[i]
            if p and p.isRunning():
                p.stop()
        for p in self._processors:
            if p:
                p.wait(5000)

        self._destroy_batch_engine()
        self._match_storage.close()
        self._save_config()
        event.accept()


# ── Entry point ────────────────────────────────────────────────────────────────

def main():
    # ── Logging ────────────────────────────────────────────────────────────────
    _fmt = "%(asctime)s [%(levelname)-8s] [%(threadName)s] %(name)s: %(message)s"
    _handlers = [
        logging.StreamHandler(sys.stdout),
        logging.FileHandler("baggage_tracker.log", encoding="utf-8"),
    ]
    for h in _handlers:
        h.setFormatter(logging.Formatter(_fmt))

    root_log = logging.getLogger()
    root_log.setLevel(logging.DEBUG)
    for h in _handlers:
        root_log.addHandler(h)

    # Заглушаем слишком шумные сторонние библиотеки
    for noisy in ("urllib3", "PIL", "ultralytics", "torch", "matplotlib",
                  "asyncio", "aiohttp", "werkzeug"):
        logging.getLogger(noisy).setLevel(logging.WARNING)

    # Перехватываем все необработанные исключения (включая из потоков)
    def _exc_hook(exctype, value, tb):
        logging.critical("UNHANDLED EXCEPTION", exc_info=(exctype, value, tb))
    sys.excepthook = _exc_hook

    import threading as _threading
    _orig_excepthook = getattr(_threading, "excepthook", None)
    def _thread_exc_hook(args):
        logging.critical(
            "UNHANDLED EXCEPTION in thread %s", args.thread,
            exc_info=(args.exc_type, args.exc_value, args.exc_traceback)
        )
        if _orig_excepthook:
            _orig_excepthook(args)
    _threading.excepthook = _thread_exc_hook

    logging.info("=" * 60)
    logging.info("Airport Baggage Tracker starting")
    logging.info("Python %s | PID %s", sys.version.split()[0], os.getpid())
    logging.info("=" * 60)

    # ── Qt platform plugin ─────────────────────────────────────────────────────
    # Выставляем путь к platforms/qwindows.dll явно из Python,
    # чтобы не зависеть от переменной окружения из bat-файла.
    # Qt platform plugins: сначала рядом с exe/py, потом внутри PyQt5-пакета
    _platforms = app_dir() / "platforms"
    if not _platforms.exists():
        try:
            import PyQt5
            _platforms = Path(PyQt5.__file__).parent / "Qt5" / "plugins" / "platforms"
            if not _platforms.exists():
                _platforms = Path(PyQt5.__file__).parent / "Qt" / "plugins" / "platforms"
        except Exception:
            pass
    if _platforms.exists():
        os.environ["QT_QPA_PLATFORM_PLUGIN_PATH"] = str(_platforms)
        logging.info("Qt platforms: %s", _platforms)
    else:
        logging.warning("Папка platforms/ не найдена — Qt может не запуститься")

    parser = argparse.ArgumentParser(description="Airport Baggage Tracker GUI")
    parser.add_argument("--config", default=CONFIG_PATH, help="Путь к config.yaml")
    args = parser.parse_args()

    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    app.setStyleSheet(DARK_QSS)

    win = MainWindow(args.config)
    win.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
